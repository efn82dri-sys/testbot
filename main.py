# -*- coding: utf-8 -*-
"""
====================================================================
 ربات تلگرام «رواق» — مرجع فایل‌های معماری و عمران
نسخهٔ VIP با اشتراک کامل و دسته‌بندی‌های نمایشی
====================================================================
"""

import asyncio
import json
import logging
import os
import uuid
import zipfile
from datetime import datetime, timedelta
from html import escape as html_escape
from io import BytesIO
from pathlib import Path

import jdatetime
import pytz
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ChatMemberStatus, ParseMode
from aiogram.filters import Command, CommandObject, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.base import BaseStorage, StorageKey
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    ChatJoinRequest,
    ChatMemberUpdated,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    PollAnswer,
    Update,
    InputMediaPhoto,
)
from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
from aiohttp import web
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------
# ۱) تنظیمات اولیه
# --------------------------------------------------------------
BOT_TOKEN = os.environ["BOT_TOKEN"]
GROUP_CHAT_ID = int(os.environ["GROUP_CHAT_ID"])

MESSAGE_EFFECT_PARTY_POPPER = "5046509860389126442"
GROUP_INVITE_LINK = os.environ.get("GROUP_INVITE_LINK", "")
NOTIFY_CHAT_ID = os.environ.get("NOTIFY_CHAT_ID", "").strip()
ADMIN_IDS = {
    int(x) for x in os.environ.get("ADMIN_IDS", "").replace(" ", "").split(",") if x
}
WEBHOOK_HOST = os.environ["WEBHOOK_HOST"].rstrip("/")
WEBHOOK_PATH = "/webhook"
WEBHOOK_URL = f"{WEBHOOK_HOST}{WEBHOOK_PATH}"
PORT = int(os.environ.get("PORT", 8080))
PING_INTERVAL_SECONDS = int(os.environ.get("PING_INTERVAL_SECONDS", 10 * 60))
AUTO_BACKUP_INTERVAL_SECONDS = int(os.environ.get("AUTO_BACKUP_INTERVAL_SECONDS", 15 * 60))
JOIN_REMINDER_AFTER_HOURS = int(os.environ.get("JOIN_REMINDER_AFTER_HOURS", 24))
JOIN_ESCALATE_AFTER_HOURS = int(os.environ.get("JOIN_ESCALATE_AFTER_HOURS", 72))

DATA_FILE = Path(__file__).parent / "data" / "submissions.jsonl"
DATA_FILE.parent.mkdir(exist_ok=True)
STATS_FILE = Path(__file__).parent / "data" / "stats.json"
VERIFIED_FILE = Path(__file__).parent / "data" / "verified_humans.json"
FUNNEL_USERS_FILE = Path(__file__).parent / "data" / "funnel_users.json"
ATTENDANCE_FILE = Path(__file__).parent / "data" / "attendance.json"
BOT_STATE_FILE = Path(__file__).parent / "data" / "bot_state.json"
MENU_CONFIG_FILE = Path(__file__).parent / "data" / "menu_config.json"
FSM_STATE_FILE = Path(__file__).parent / "data" / "fsm_state.json"
PENDING_JOIN_FILE = Path(__file__).parent / "data" / "pending_join_requests.json"

# ---------- تنظیمات گروه VIP ----------
VIP_GROUP_CHAT_ID_RAW = os.environ.get("VIP_GROUP_CHAT_ID", "").strip()
try:
    VIP_GROUP_CHAT_ID: int | None = int(VIP_GROUP_CHAT_ID_RAW) if VIP_GROUP_CHAT_ID_RAW else None
except ValueError:
    VIP_GROUP_CHAT_ID = None

VIP_CARD_NUMBER = os.environ.get("VIP_CARD_NUMBER", "6219861963810246").strip()
VIP_CARD_HOLDER = os.environ.get("VIP_CARD_HOLDER", "عرفان دارائی").strip()
VIP_INTRO_DELAY_MINUTES = int(os.environ.get("VIP_INTRO_DELAY_MINUTES", 30))

VIP_CATEGORIES_FILE = Path(__file__).parent / "data" / "vip_categories.json"
VIP_SUBSCRIPTIONS_FILE = Path(__file__).parent / "data" / "vip_subscriptions.json"
VIP_PAYMENTS_FILE = Path(__file__).parent / "data" / "vip_payments.json"
VIP_GLOBAL_SETTINGS_FILE = Path(__file__).parent / "data" / "vip_global_settings.json"

VIP_MONTHS_TO_DAYS = {3: 90, 6: 180, 12: 365}
_vip_intro_tasks: dict[int, asyncio.Task] = {}

REFERRAL_LABELS = {
    "instagram": "📷 اینستاگرام",
    "friends": "👥 معرفی دوستان",
    "other_groups": "💬 سایر گروه‌ها و کانال‌ها",
    "search": "🔍 جستجوی اینترنتی",
    "other": "✨ سایر موارد",
}

EDUCATION_OPTIONS: list[tuple[str, str]] = [
    ("diploma", "دیپلم / پیش‌دانشگاهی"),
    ("associate", "کاردانی"),
    ("bachelor", "کارشناسی"),
    ("master", "کارشناسی ارشد"),
    ("phd", "دکتری"),
    ("other", "سایر"),
]

INTERESTS: list[str] = [
    "اتاق پرامپت",
    "فرصت‌های شغلی",
    "پرزانته و پرتفولیو",
    "آکادمی آنلاین",
    "کتابخانه و ضوابط ملی",
    "رادیو معماری",
    "بانک پروژه",
    "معماری جهان",
    "فایل‌های گرافیکی",
    "دنیای نرم‌افزار و پلاگین",
    "آبجکت، فمیلی و متریال",
    "پلان و نقشه‌های اجرایی",
]
MAX_INTERESTS = 3

GROUP_NAME = "رواق"
SIGNATURE = f"\n\n— <i>تیمِ {GROUP_NAME}</i> 🏛"

# ---------- متنِ قوانینِ گروه (ارسال هنگامِ درخواستِ عضویت) ----------
GROUP_RULES_TEXT = (
    "<b>سلام دوست من 🌱</b>\n\n"
    "<blockquote><b>این فضا رو VIP کردیم که یه پاتوق زنده و پرانرژی برای معمارای واقعی باشه، نه فقط یه انبار فایل ساکت!</b></blockquote>\n\n"
    "<b>🤖 قانون فعالیت:</b>\n"
    "اگه ۷ روز توی گروه هیچ تعاملی نداشته باشید، ربات به‌طور خودکار حذفتون می‌کنه.\n\n"
    "پس همراهمون باشید و توی بحث‌های <a href='https://t.me/c/4388421316/95'>کافه معماری</a> فعال بمونید ❤️☕️\n\n"
    "<blockquote><b>@IRarchit</b></blockquote>"
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ==============================================================
#  استوریجِ پایدارِ FSM (جایگزینِ MemoryStorage)
#  MemoryStorage وضعیتِ گفتگو (مثلاً وسطِ پرکردنِ پروفایل یا وسطِ
#  آپلودِ رسیدِ پرداختِ VIP) را فقط توی RAM نگه می‌داشت؛ با هر
#  ری‌استارت/دیپلویِ Render این وضعیت کامل از بین می‌رفت و کاربر
#  وسطِ کار جواب بی‌ربط می‌گرفت یا گیر می‌کرد. اینجا همان وضعیت را
#  روی دیسک (داخل DATA_DIR) نگه می‌داریم تا:
#   ۱) با ری‌استارتِ معمولی از بین نرود (نوشتنِ فوری و اتمیک روی هر تغییر)
#   ۲) داخلِ همان مکانیزمِ بکاپ/بازیابیِ تلگرام هم بیفتد (چون داخلِ
#      DATA_DIR است) و با بکاپِ دوره‌ای که پایین‌تر اضافه شده، حتی جلوی
#      ری‌استارت‌های سنگین/دیپلوی هم تا حدِ زیادی محافظت شود.
# ==============================================================

class JSONFSMStorage(BaseStorage):
    def __init__(self, path: Path):
        self._path = path
        self._data: dict[str, dict] = {}
        self._lock = asyncio.Lock()
        self._load()

    def _load(self) -> None:
        if self._path.exists():
            try:
                self._data = json.loads(self._path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError) as e:
                logger.warning(f"خواندنِ فایلِ وضعیتِ FSM ناموفق بود، با حالتِ خالی شروع می‌کنیم: {e}")
                self._data = {}

    def _persist(self) -> None:
        try:
            self._path.parent.mkdir(exist_ok=True)
            tmp_path = self._path.with_suffix(".tmp")
            tmp_path.write_text(json.dumps(self._data, ensure_ascii=False), encoding="utf-8")
            tmp_path.replace(self._path)  # جایگزینیِ اتمیک، برای جلوگیری از فایلِ نصفه‌نوشته
        except OSError as e:
            logger.error(f"ذخیره‌ی وضعیتِ FSM روی دیسک ناموفق بود: {e}")

    @staticmethod
    def _key_str(key: StorageKey) -> str:
        parts = [str(key.bot_id), str(key.chat_id), str(key.user_id)]
        for extra_attr in ("thread_id", "business_connection_id", "destiny"):
            parts.append(str(getattr(key, extra_attr, None)))
        return ":".join(parts)

    async def set_state(self, key: StorageKey, state=None) -> None:
        async with self._lock:
            state_str = state.state if isinstance(state, State) else state
            k = self._key_str(key)
            entry = self._data.get(k, {})
            if state_str is None:
                entry.pop("state", None)
            else:
                entry["state"] = state_str
            if entry:
                self._data[k] = entry
            else:
                self._data.pop(k, None)
            self._persist()

    async def get_state(self, key: StorageKey):
        return self._data.get(self._key_str(key), {}).get("state")

    async def set_data(self, key: StorageKey, data: dict) -> None:
        async with self._lock:
            k = self._key_str(key)
            entry = self._data.get(k, {})
            if data:
                entry["data"] = data
            else:
                entry.pop("data", None)
            if entry:
                self._data[k] = entry
            else:
                self._data.pop(k, None)
            self._persist()

    async def get_data(self, key: StorageKey) -> dict:
        return self._data.get(self._key_str(key), {}).get("data", {}) or {}

    async def close(self) -> None:
        self._persist()

    def reload(self) -> None:
        """وضعیت را دوباره از دیسک می‌خواند — لازم است بعد از بازیابیِ بکاپ در
        on_startup صدا زده شود، چون این آبجکت قبل از آن (زمانِ import) ساخته
        و از دیسک خوانده شده و از تغییراتِ restore بی‌خبر می‌ماند."""
        self._load()

storage = JSONFSMStorage(FSM_STATE_FILE)
bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML, protect_content=True),
)
dp = Dispatcher(storage=storage)

_write_lock = asyncio.Lock()
_pending_leave_polls: dict[str, int] = {}
_pending_admin_replies: dict[int, int] = {}
_attendance_tasks: dict[str, asyncio.Task] = {}
_user_cache: dict[str, dict] = {}

try:
    NOTIFY_CHAT_ID_INT = int(NOTIFY_CHAT_ID) if NOTIFY_CHAT_ID else None
except ValueError:
    NOTIFY_CHAT_ID_INT = None

# ---------- تنظیمات بکاپ (چون فایل‌سیستم Render ناپایدار است) ----------
BACKUP_CHAT_ID_RAW = os.environ.get("BACKUP_CHAT_ID", "").strip()
try:
    BACKUP_CHAT_ID: int | None = int(BACKUP_CHAT_ID_RAW) if BACKUP_CHAT_ID_RAW else NOTIFY_CHAT_ID_INT
except ValueError:
    BACKUP_CHAT_ID = NOTIFY_CHAT_ID_INT
DATA_DIR = DATA_FILE.parent

# ==============================================================
#  بکاپِ دستیِ پوشه‌ی data روی تلگرام
#  (Render فایل‌سیستم ناپایدار دارد؛ هر ری‌استارت/دیپلوی فایل‌های محلی را
#   پاک می‌کند، پس با دکمه‌ی «📥 گرفتن بکاپ» در پنل ادمین یک نسخه‌ی پشتیبان
#   در یک پیامِ پین‌شده نگه می‌داریم و در استارتاپ خودکار بازیابی می‌شود)
# ==============================================================

def _zip_data_dir() -> BytesIO:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in DATA_DIR.rglob("*"):
            if file_path.is_file():
                zf.write(file_path, arcname=str(file_path.relative_to(DATA_DIR)))
    buf.seek(0)
    return buf

async def _notify_backup_admin(text: str) -> None:
    """گزارشِ وضعیتِ بکاپ/بازیابی را برای ادمین ارسال می‌کند تا خطاها دیگر بی‌صدا گم نشوند."""
    if not NOTIFY_CHAT_ID_INT:
        return
    try:
        await bot.send_message(chat_id=NOTIFY_CHAT_ID_INT, text=text, disable_notification=True)
    except Exception as e:
        logger.error(f"ارسال گزارشِ بکاپ به ادمین ممکن نشد: {e}")

async def backup_data_dir_to_telegram() -> tuple[bool, str]:
    """بکاپ می‌گیرد و روی تلگرام پین می‌کند. خروجی: (موفقیت, پیامِ توضیحی)."""
    if not BACKUP_CHAT_ID:
        msg = "BACKUP_CHAT_ID تنظیم نشده — گرفتنِ بکاپ ممکن نیست."
        logger.warning(msg)
        return False, msg
    if not DATA_DIR.exists() or not any(f.is_file() for f in DATA_DIR.rglob("*")):
        msg = "پوشه‌ی data خالی است — چیزی برای بکاپ‌گیری وجود ندارد."
        logger.info(msg)
        return False, msg
    try:
        buf = _zip_data_dir()
        filename = f"ravaq_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
        message = await bot.send_document(
            chat_id=BACKUP_CHAT_ID,
            document=BufferedInputFile(buf.read(), filename=filename),
            caption=f"🗄 بکاپ دیتا — {format_jalali_datetime(datetime.utcnow())}",
            disable_notification=True,
        )
        try:
            await bot.unpin_all_chat_messages(BACKUP_CHAT_ID)
        except Exception as e:
            # اگر آنپین ناموفق باشد، خودِ pin_chat_message در ادامه پیامِ جدید را پین می‌کند؛
            # فقط لاگ می‌کنیم که بی‌صدا گم نشود.
            logger.warning(f"آنپین کردنِ بکاپِ قبلی ناموفق بود (ادامه می‌دهیم): {e}")
        await bot.pin_chat_message(BACKUP_CHAT_ID, message.message_id, disable_notification=True)
        logger.info("بکاپ دیتا با موفقیت ارسال و پین شد.")
        return True, "بکاپ با موفقیت گرفته و پین شد."
    except Exception as e:
        msg = f"ارسال بکاپ ناموفق بود: {e}"
        logger.error(msg, exc_info=True)
        return False, msg

def _clear_data_dir_files() -> None:
    """همه‌ی فایل‌های محلیِ data را پاک می‌کند تا بازیابی از بکاپ واقعاً «جایگزین» شود،
    نه اینکه با فایل‌های قدیمی/ناقصِ باقی‌مانده قاطی شود."""
    if not DATA_DIR.exists():
        return
    for file_path in DATA_DIR.rglob("*"):
        if file_path.is_file():
            try:
                file_path.unlink()
            except Exception as e:
                logger.warning(f"حذفِ فایلِ محلیِ {file_path} قبل از بازیابی ناموفق بود: {e}")

async def restore_data_dir_from_telegram(force: bool = False) -> tuple[bool, str]:
    """
    دیتا را از بکاپِ پین‌شده در تلگرام بازیابی می‌کند.
    خروجی: (موفقیت, پیامِ توضیحی/دلیلِ شکست) — تا دیگر شکست‌ها بی‌صدا گم نشوند.

    توجه: چون فایل‌سیستمِ Render ناپایدار است، منبعِ حقیقتِ داده همیشه بکاپِ تلگرام است؛
    پس این تابع در صورتِ پیدا کردنِ بکاپِ معتبر، فایل‌های محلی را کامل جایگزین می‌کند
    (نه فقط زمانی که پوشه‌ی data خالی باشد).
    """
    if not BACKUP_CHAT_ID:
        msg = "BACKUP_CHAT_ID تنظیم نشده — بازیابیِ خودکار از تلگرام غیرفعال است."
        logger.info(msg)
        return False, msg

    has_local_data = DATA_DIR.exists() and any(f.is_file() for f in DATA_DIR.rglob("*"))
    if has_local_data and not force:
        logger.info(
            "دیتای محلی از قبل موجود است؛ برای اطمینان همچنان تلاش می‌کنیم آخرین بکاپِ تلگرام را بخوانیم "
            "و در صورتِ پیدا شدن، جایگزینِ دیتای محلی می‌کنیم."
        )

    try:
        logger.info("در حال گرفتنِ اطلاعاتِ چتِ بکاپ (chat_id=%s)...", BACKUP_CHAT_ID)
        chat = await bot.get_chat(BACKUP_CHAT_ID)
    except Exception as e:
        msg = (
            f"دسترسی به چتِ بکاپ (chat_id={BACKUP_CHAT_ID}) ناموفق بود: {e}\n"
            "بررسی کنید که ربات هنوز عضوِ آن چت است و BACKUP_CHAT_ID درست است."
        )
        logger.error(msg, exc_info=True)
        return False, msg

    pinned = chat.pinned_message
    if not pinned:
        msg = "هیچ پیامِ پین‌شده‌ای در چتِ بکاپ پیدا نشد."
        logger.info(msg)
        return False, msg
    if not pinned.document:
        msg = (
            f"پیامِ پین‌شده در چتِ بکاپ فاقدِ فایل (document) است "
            f"(message_id={pinned.message_id}) — احتمالاً پیامِ دیگری غیر از بکاپ پین شده."
        )
        logger.warning(msg)
        return False, msg

    try:
        logger.info(
            "در حال دانلودِ فایلِ بکاپِ پین‌شده: %s (%s بایت)",
            pinned.document.file_name, pinned.document.file_size,
        )
        file_bytes = await bot.download(pinned.document.file_id)
    except Exception as e:
        msg = f"دانلودِ فایلِ بکاپ ناموفق بود: {e}"
        logger.error(msg, exc_info=True)
        return False, msg

    try:
        DATA_DIR.mkdir(exist_ok=True)
        with zipfile.ZipFile(file_bytes) as zf:
            bad_file = zf.testzip()
            if bad_file:
                raise zipfile.BadZipFile(f"فایلِ خراب در آرشیو: {bad_file}")
            # ابتدا دیتای محلیِ فعلی را کامل پاک می‌کنیم تا محتوایِ بکاپ واقعاً
            # جایگزینِ آن شود، نه اینکه با فایل‌های قدیمی قاطی/ادغام شود.
            _clear_data_dir_files()
            zf.extractall(DATA_DIR)
        restored_files = [str(p.relative_to(DATA_DIR)) for p in DATA_DIR.rglob("*") if p.is_file()]
        msg = f"دیتا با موفقیت از بکاپِ تلگرام بازیابی شد ({len(restored_files)} فایل)."
        logger.info(msg)
        return True, msg
    except zipfile.BadZipFile as e:
        msg = f"فایلِ دانلودشده یک ZIP معتبر نیست: {e}"
        logger.error(msg, exc_info=True)
        return False, msg
    except Exception as e:
        msg = f"استخراجِ فایلِ بکاپ ناموفق بود: {e}"
        logger.error(msg, exc_info=True)
        return False, msg

LEAVE_REASONS: list[tuple[str, str]] = [
    (
        "فایل‌ها و محتوای گروه به‌دردم نخورد",
        "حیف شد! اگر دقیقاً بگویی دنبالِ چه فایلی بودی، حتماً در انبارِ این "
        "رواق گم‌شده‌ای پیدا می‌شود که به‌کارت بیاید.\n\n"
        "به ادمین‌ها پیام بده، شاید درِ گنج‌خانه‌ای تازه باز شود 🙏",
    ),
    (
        "پیام‌های زیاد گروه رو شلوغ می‌کرد",
        "راستی؟ می‌دونی که می‌تونی گروه رو روی حالتِ سکوت بذاری و فقط گاهی "
        "سراغِ «پیام‌های سنجاق‌شده» (همون فایل‌های طلایی) بیای؟\n\n"
        "بدونِ اینکه اعلان‌ها اذیتت کنن 🔕",
    ),
    (
        "فعلاً به این موضوع نیاز ندارم",
        "کاملاً درک می‌کنم. بساطِ معماری گاهی خلوت‌شدن هم می‌خواد.\n\n"
        "هر وقت دوباره خواستی قدم بذاری، درِ رواق به رویت باز است 🙌",
    ),
    (
        "دلیل دیگه‌ای دارم",
        "ممنون که وقت گذاشتی.\n\n"
        "اگه حرفِ دلت رو مستقیم با ادمین‌ها در میون بذاری، به ما در مرمتِ این فضا کمکِ بزرگی کردی 🙏",
    ),
]

# ---------- زمان و تاریخ شمسی ----------
TEHRAN_TZ = pytz.timezone('Asia/Tehran')

def utc_to_tehran(utc_dt: datetime) -> datetime:
    return utc_dt.astimezone(TEHRAN_TZ)

def to_jalali(utc_dt: datetime) -> jdatetime.datetime:
    tehran_dt = utc_to_tehran(utc_dt)
    return jdatetime.datetime.fromgregorian(datetime=tehran_dt.replace(tzinfo=None))

def format_jalali_datetime(utc_dt: datetime) -> str:
    jalali = to_jalali(utc_dt)
    return jalali.strftime("%Y/%m/%d %H:%M:%S")

# ---------- توابع کمکی ----------
def to_persian_num(num) -> str:
    mapping = {
        '0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴',
        '5': '۵', '6': '۶', '7': '۷', '8': '۸', '9': '۹'
    }
    return ''.join(mapping.get(ch, ch) for ch in str(num))

def greet_user(user, suffix="عزیز") -> str:
    name = html_escape(user.first_name or "کاربر")
    return f"{name} {suffix}"

def sign(text: str) -> str:
    return f"{text}{SIGNATURE}"

def progress_bar(step: int, total: int = 3) -> str:
    step = max(0, min(step, total))
    return ("🟩" * step) + ("⬜" * (total - step))

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def load_stats() -> dict:
    if not STATS_FILE.exists():
        return {"total_joined": 0, "total_left": 0}
    try:
        return json.loads(STATS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"total_joined": 0, "total_left": 0}

async def increment_stat(field: str) -> None:
    async with _write_lock:
        stats = load_stats()
        stats[field] = stats.get(field, 0) + 1
        STATS_FILE.write_text(json.dumps(stats, ensure_ascii=False), encoding="utf-8")

def load_verified() -> dict:
    if not VERIFIED_FILE.exists():
        return {}
    try:
        return json.loads(VERIFIED_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

async def mark_verified(user_id: int) -> None:
    async with _write_lock:
        verified = load_verified()
        verified[str(user_id)] = datetime.utcnow().isoformat()
        VERIFIED_FILE.write_text(json.dumps(verified, ensure_ascii=False), encoding="utf-8")

def is_verified(user_id: int) -> bool:
    return str(user_id) in load_verified()

def load_funnel_users() -> set[int]:
    if not FUNNEL_USERS_FILE.exists():
        return set()
    try:
        return set(json.loads(FUNNEL_USERS_FILE.read_text(encoding="utf-8")))
    except (json.JSONDecodeError, OSError):
        return set()

async def mark_funnel_entry(user_id: int) -> None:
    async with _write_lock:
        users = load_funnel_users()
        if user_id in users:
            return
        users.add(user_id)
        FUNNEL_USERS_FILE.write_text(json.dumps(list(users)), encoding="utf-8")

def collect_form_user_ids() -> set[int]:
    user_ids: set[int] = set()
    if not DATA_FILE.exists():
        return user_ids
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
                user_ids.add(int(record["user_id"]))
            except (json.JSONDecodeError, KeyError, ValueError):
                continue
    return user_ids

async def is_user_member(user_id: int) -> bool:
    try:
        member = await bot.get_chat_member(GROUP_CHAT_ID, user_id)
        return member.status in [ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR]
    except Exception:
        return False

def collect_form_user_ids_by_interest(interest: str) -> set[int]:
    user_ids: set[int] = set()
    if not DATA_FILE.exists():
        return user_ids
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
                if interest in (record.get("interests") or []):
                    user_ids.add(int(record["user_id"]))
            except (json.JSONDecodeError, KeyError, ValueError):
                continue
    return user_ids

def is_form_completed(user_id: int) -> bool:
    return str(user_id) in _user_cache

def cache_users():
    if not DATA_FILE.exists():
        return
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
                _user_cache[str(record["user_id"])] = record
            except (json.JSONDecodeError, KeyError):
                continue

# نکته: cache_users() دیگر اینجا (زمان import) صدا زده نمی‌شود؛
# چون باید بعد از بازیابیِ احتمالیِ بکاپ از تلگرام در on_startup اجرا شود
# (وگرنه با فایل‌سیستم خالیِ تازه‌ری‌استارت‌شده کش خالی می‌ماند).

# ==============================================================
#  توابع کمکی داده‌های VIP
# ==============================================================

def load_vip_categories() -> list[dict]:
    if not VIP_CATEGORIES_FILE.exists():
        return []
    try:
        return json.loads(VIP_CATEGORIES_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []

async def save_vip_categories(categories: list[dict]) -> None:
    async with _write_lock:
        VIP_CATEGORIES_FILE.parent.mkdir(exist_ok=True)
        VIP_CATEGORIES_FILE.write_text(json.dumps(categories, ensure_ascii=False, indent=2), encoding="utf-8")

def get_vip_category(cat_id: str) -> dict | None:
    for cat in load_vip_categories():
        if cat["id"] == cat_id:
            return cat
    return None

def load_vip_subscriptions() -> dict:
    if not VIP_SUBSCRIPTIONS_FILE.exists():
        return {}
    try:
        return json.loads(VIP_SUBSCRIPTIONS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

async def save_vip_subscriptions(data: dict) -> None:
    async with _write_lock:
        VIP_SUBSCRIPTIONS_FILE.parent.mkdir(exist_ok=True)
        VIP_SUBSCRIPTIONS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def get_user_vip_status(user_id: int, subs: dict | None = None) -> dict:
    """
    خلاصه‌ی وضعیتِ VIP یک کاربر را برمی‌گرداند.
    چون هر تمدید یک رکوردِ جدید به لیستِ اشتراک‌های کاربر اضافه می‌کند، وضعیتِ
    واقعیِ کاربر باید بر اساسِ «آخرین تاریخِ پایان» در میانِ تمامِ رکوردهای
    active/renewed محاسبه شود، نه صرفاً یک رکورد؛ این از قطعِ زودهنگامِ
    دسترسی در صورتِ وجودِ چند رکوردِ هم‌پوشان جلوگیری می‌کند.
    """
    subs = subs if subs is not None else load_vip_subscriptions()
    user_subs = subs.get(str(user_id), [])
    now = datetime.utcnow()

    latest_end: datetime | None = None
    for sub in user_subs:
        if sub.get("status") not in ("active", "renewed"):
            continue
        try:
            end = datetime.fromisoformat(sub["end"])
        except (KeyError, ValueError):
            continue
        if latest_end is None or end > latest_end:
            latest_end = end

    if latest_end is None:
        return {"has_subscription": False, "is_active": False, "end": None, "remaining_days": 0}

    remaining_seconds = (latest_end - now).total_seconds()
    remaining_days = max(0, int(remaining_seconds // 86400) + (1 if remaining_seconds % 86400 > 0 else 0))
    return {
        "has_subscription": True,
        "is_active": remaining_seconds > 0,
        "end": latest_end,
        "remaining_days": remaining_days,
    }

def load_vip_payments() -> dict:
    if not VIP_PAYMENTS_FILE.exists():
        return {}
    try:
        return json.loads(VIP_PAYMENTS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

async def save_vip_payments(data: dict) -> None:
    async with _write_lock:
        VIP_PAYMENTS_FILE.parent.mkdir(exist_ok=True)
        VIP_PAYMENTS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def load_vip_global_settings() -> dict:
    if not VIP_GLOBAL_SETTINGS_FILE.exists():
        return {
            "prices": {"3": 150000, "6": 250000, "12": 400000},
            "discount_percent": 0,
            "updated_at": datetime.utcnow().isoformat()
        }
    try:
        return json.loads(VIP_GLOBAL_SETTINGS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"prices": {"3": 150000, "6": 250000, "12": 400000}, "discount_percent": 0}

async def save_vip_global_settings(settings: dict) -> None:
    settings["updated_at"] = datetime.utcnow().isoformat()
    async with _write_lock:
        VIP_GLOBAL_SETTINGS_FILE.parent.mkdir(exist_ok=True)
        VIP_GLOBAL_SETTINGS_FILE.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")

def format_toman(amount: int) -> str:
    return f"{to_persian_num(f'{amount:,}')} تومان"

# ---------- توابع آمار ----------
async def build_stats_text() -> str:
    try:
        member_count = await bot.get_chat_member_count(GROUP_CHAT_ID)
        member_count_str = to_persian_num(member_count)
    except Exception as e:
        logger.warning("گرفتن تعداد اعضا ممکن نشد: %s", e)
        member_count_str = "نامشخص"

    form_count = 0
    if DATA_FILE.exists():
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            form_count = sum(1 for line in f if line.strip())

    stats = load_stats()
    funnel_count = len(load_funnel_users())
    verified_count = len(load_verified())
    form_joined_count = stats.get("form_completed_and_joined", 0)

    verified_rate = (verified_count / funnel_count * 100) if funnel_count else 0
    joined_rate = (form_joined_count / verified_count * 100) if verified_count else 0

    return (
        "📐 <b>داشبورد رواق (آمار لحظه‌ای)</b>\n\n"
        f"👥 ساکنینِ فعلی گروه: <b>{member_count_str}</b>\n\n"
        "<b>قیفِ عضویت:</b>\n"
        f"1️⃣ استارت ربات / پیامِ درخواست عضویت: <b>{to_persian_num(funnel_count)}</b>\n"
        f"2️⃣ تاییدِ عدمِ ربات‌بودن: <b>{to_persian_num(verified_count)}</b> ({verified_rate:.0f}٪)\n"
        f"3️⃣ فرمِ تکمیل‌شده + ورود به گروه: <b>{to_persian_num(form_joined_count)}</b> ({joined_rate:.0f}٪)\n\n"
        f"📝 کل فرم‌های ثبت‌شده (شامل موارد تأییدنشده): <b>{to_persian_num(form_count)}</b>\n\n"
        "<i>این آمار از زمانی که دروازه‌ی الکترونیکی نصب شده، ثبت می‌شود.</i>"
    )

async def build_admin_dashboard_text() -> str:
    status_text = "روشن ✅" if load_bot_state().get("enabled", True) else "خاموش 🔴"
    stats_text = await build_stats_text()
    return f"🛠 <b>پنل مدیریت</b>\nوضعیت ربات: {status_text}\n\n{stats_text}"

async def build_stats_detail_text() -> str:
    if not DATA_FILE.exists():
        return "هنوز هیچ فرمی ثبت نشده است."

    educations: dict[str, int] = {}
    referrals: dict[str, int] = {}
    interests: dict[str, int] = {}
    form_count = 0

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue

            form_count += 1

            edu_label = record.get("education_label") or record.get("education") or "نامشخص"
            educations[edu_label] = educations.get(edu_label, 0) + 1

            ref = record.get("referral") or "نامشخص"
            ref_label = REFERRAL_LABELS.get(ref, ref)
            referrals[ref_label] = referrals.get(ref_label, 0) + 1

            for interest in record.get("interests") or []:
                interests[interest] = interests.get(interest, 0) + 1

    if form_count == 0:
        return "هنوز هیچ فرمی ثبت نشده است."

    lines = [
        f"📊 <b>آمارِ تفصیلیِ ساکنانِ رواق</b>\n"
        f"از میانِ <b>{to_persian_num(form_count)}</b> نفری که احرازِ هویت را کامل کرده‌اند:\n"
    ]

    lines.append("<b>مقطعِ تحصیلی:</b>")
    for label, count in sorted(educations.items(), key=lambda x: -x[1]):
        lines.append(f"▪️ {label}: <b>{to_persian_num(count)}</b> نفر")

    lines.append("\n<b>نحوه‌ی آشنایی:</b>")
    for label, count in sorted(referrals.items(), key=lambda x: -x[1]):
        lines.append(f"▪️ {label}: <b>{to_persian_num(count)}</b> نفر")

    lines.append("\n<b>علایق:</b>")
    if interests:
        for label, count in sorted(interests.items(), key=lambda x: -x[1]):
            lines.append(f"▪️ {label}: <b>{to_persian_num(count)}</b> نفر")
    else:
        lines.append("هنوز کسی علایقش را ثبت نکرده.")

    return "\n".join(lines)

def build_export_file() -> BufferedInputFile | None:
    verified = load_verified()
    if not verified and not DATA_FILE.exists():
        return None

    form_records = {}
    if DATA_FILE.exists():
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                    uid = str(record.get("user_id"))
                    form_records[uid] = record
                except json.JSONDecodeError:
                    continue

    all_user_ids = set(verified.keys()) | set(form_records.keys())
    if not all_user_ids:
        return None

    rows = []
    for uid_str in all_user_ids:
        try:
            uid_int = int(uid_str)
        except ValueError:
            continue
        record = form_records.get(uid_str, {})

        username = record.get("username")
        full_name = record.get("full_name", "")
        submitted_at = record.get("submitted_at", "")
        education = record.get("education_label") or record.get("education") or "-"
        referral = REFERRAL_LABELS.get(record.get("referral"), record.get("referral") or "-")
        interests_list = record.get("interests", [])
        interests_str = "، ".join(interests_list) if interests_list else "-"

        form_status = "تکمیل شده" if record else "تکمیل نشده"

        rows.append([
            uid_str,
            f"@{username}" if username else "-",
            full_name or "-",
            submitted_at[:16] if submitted_at else "-",
            education,
            referral,
            interests_str,
            form_status,
        ])

    rows.sort(key=lambda r: (r[3] == "-", r[3]), reverse=False)

    headers = [
        "آیدی عددی",
        "نام کاربری",
        "نام کامل",
        "تاریخ و ساعت ثبت (UTC)",
        "مقطع تحصیلی",
        "نحوه آشنایی",
        "علایق انتخاب‌شده",
        "وضعیت فرم",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "همه‌ی تأییدشده‌ها"
    sheet.sheet_view.rightToLeft = True

    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="14532F", end_color="14532F", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append(row)

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_index, _ in enumerate(headers, start=1):
        max_len = len(headers[col_index - 1])
        for row in rows:
            cell_value = row[col_index - 1]
            max_len = max(max_len, len(str(cell_value)))
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max_len + 4, 42)

    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return BufferedInputFile(buffer.read(), filename="همه‌ی تأییدشده‌ها.xlsx")

# ---------- مدیریت منوی پویا ----------
def migrate_menu_config(config: dict) -> dict:
    if "menu_items" not in config:
        config["menu_items"] = {}
    if "settings" not in config:
        config["settings"] = {}

    if "join" in config["menu_items"]:
        if config["menu_items"]["join"].get("label") == "📝 عضویت در گروه":
            config["menu_items"]["join"]["label"] = "👥 دعوت از دوستان"

    defaults = {
        "vip": {"label": "🌟 گروه VIP", "response": ""},
        "join": {"label": "👥 دعوت از دوستان", "response": "🔗 لینک دعوت گروه:\n{invite_link}"},
        "topics": {"label": "📚 راهنمای تاپیک‌ها", "response": "لطفاً یکی از تاپیک‌های زیر را انتخاب کنید:"},
        "contact_admin": {"label": "📞 ارتباط با ادمین", "response": "پیام خود را تایپ کنید تا برای ادمین ارسال شود."},
        "my_status": {"label": "📊 وضعیت عضویت من", "response": "وضعیت شما: {status}"},
        "announcements": {"label": "📢 اطلاعیه‌های جدید", "response": "آخرین اطلاعیه‌ها:\n{announcements}"},
        "faq": {"label": "❓ سوالات متداول", "response": "سوالات پرتکرار:\n{faq_list}"},
        "social": {"label": "🌐 شبکه‌های اجتماعی", "response": "ما را دنبال کنید:\nاینستاگرام: {instagram}\nکانال: {channel}"},
    }
    for key, val in defaults.items():
        if key not in config["menu_items"]:
            config["menu_items"][key] = val
        else:
            if key == "join":
                config["menu_items"][key]["label"] = defaults[key]["label"]
                config["menu_items"][key]["response"] = defaults[key]["response"]

    if "settings" not in config:
        config["settings"] = {}
    settings_defaults = {
        "group_invite_link": GROUP_INVITE_LINK,
        "announcements": [],
        "announcement_files": [],
        "faq": [],
        "faq_files": [],
        "social": {
            "instagram": "https://www.instagram.com/archit.ir/",
            "channel": "https://t.me/irarchit"
        }
    }
    for key, val in settings_defaults.items():
        if key not in config["settings"]:
            config["settings"][key] = val

    return config

def load_menu_config() -> dict:
    if not MENU_CONFIG_FILE.exists():
        default_config = {
            "menu_items": {
                "vip": {"label": "🌟 گروه VIP", "response": ""},
                "join": {"label": "👥 دعوت از دوستان", "response": "🔗 لینک دعوت گروه:\n{invite_link}"},
                "topics": {"label": "📚 راهنمای تاپیک‌ها", "response": "لطفاً یکی از تاپیک‌های زیر را انتخاب کنید:"},
                "contact_admin": {"label": "📞 ارتباط با ادمین", "response": "پیام خود را تایپ کنید تا برای ادمین ارسال شود."},
                "my_status": {"label": "📊 وضعیت عضویت من", "response": "وضعیت شما: {status}"},
                "announcements": {"label": "📢 اطلاعیه‌های جدید", "response": "آخرین اطلاعیه‌ها:\n{announcements}"},
                "faq": {"label": "❓ سوالات متداول", "response": "سوالات پرتکرار:\n{faq_list}"},
                "social": {"label": "🌐 شبکه‌های اجتماعی", "response": "ما را دنبال کنید:\nاینستاگرام: {instagram}\nکانال: {channel}"},
            },
            "settings": {
                "group_invite_link": GROUP_INVITE_LINK,
                "announcements": [],
                "announcement_files": [],
                "faq": [],
                "faq_files": [],
                "social": {
                    "instagram": "https://www.instagram.com/archit.ir/",
                    "channel": "https://t.me/irarchit"
                }
            }
        }
        MENU_CONFIG_FILE.parent.mkdir(exist_ok=True)
        MENU_CONFIG_FILE.write_text(json.dumps(default_config, ensure_ascii=False, indent=2), encoding="utf-8")
        return default_config
    try:
        config = json.loads(MENU_CONFIG_FILE.read_text(encoding="utf-8"))
        config = migrate_menu_config(config)
        return config
    except:
        return load_menu_config()

async def save_menu_config(config: dict) -> None:
    async with _write_lock:
        MENU_CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

# ---------- پنل تاپیک‌ها ----------
TOPICS = {
    "🎓 آکادمی آنلاین": "https://t.me/c/4388421316/146",
    "🛠 رفع اشکال تخصصی": "https://t.me/thedaraeii",
    "📐 پلان و نقشه‌های اجرایی": "https://t.me/c/4388421316/143",
    "💻 نرم‌افزار و پلاگین": "https://t.me/c/4388421316/136",
    "🎨 گرافیک و پست‌پرو": "https://t.me/c/4388421316/134",
    "🤖 اتاق پرامپت": "https://t.me/c/4388421316/114",
    "🛋 آبجکت و متریال": "https://t.me/c/4388421316/140",
    "📚 کتابخانه ضوابط ملی": "https://t.me/c/4388421316/123",
    "🖼 پرزانته و پورتفولیو": "https://t.me/c/4388421316/121",
    "📂 بانک پروژه": "https://t.me/c/4388421316/149",
    "💼 فرصت‌های شغلی": "https://t.me/c/4388421316/109",
    "🌐 معماری جهان": "https://t.me/c/4388421316/131",
    "☕️ کافه معماری": "https://t.me/c/4388421316/95",
    "🎙 رادیو معماری": "https://t.me/c/4388421316/125"
}

def topics_panel_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    topic_items = list(TOPICS.items())
    for i in range(0, len(topic_items), 2):
        row = []
        row.append(InlineKeyboardButton(text=topic_items[i][0], url=topic_items[i][1], style="primary"))
        if i+1 < len(topic_items):
            row.append(InlineKeyboardButton(text=topic_items[i+1][0], url=topic_items[i+1][1], style="primary"))
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back", style="primary")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ---------- پنل کاربری ----------
_USER_MENU_STYLES = {
    "profile": "success",
    "vip": "success",
    "join": "primary",
    "topics": "primary",
    "contact_admin": "primary",
    "my_status": "primary",
    "announcements": "primary",
    "faq": "primary",
    "social": "primary",
}

def user_panel_keyboard() -> InlineKeyboardMarkup:
    config = load_menu_config()
    items = config["menu_items"]
    # چیدمانِ ثابتِ پنلِ کاربری — دقیقاً به همین ترتیب و در همین ردیف‌های دوتایی:
    # پروفایل من | گروه VIP
    # راهنمای تاپیک‌ها | دعوت از دوستان
    # وضعیت عضویت من | ارتباط با ادمین
    # سوالات متداول | شبکه‌های اجتماعی
    # بستن پنل
    rows_keys = [
        ("profile", "vip"),
        ("topics", "join"),
        ("my_status", "contact_admin"),
        ("faq", "social"),
    ]
    labels = {
        "profile": "👤 پروفایل من",
        **{k: v["label"] for k, v in items.items()},
    }
    buttons = []
    for key1, key2 in rows_keys:
        buttons.append([
            InlineKeyboardButton(
                text=labels[key1], callback_data=f"menu:{key1}",
                style=_USER_MENU_STYLES.get(key1, "primary"),
            ),
            InlineKeyboardButton(
                text=labels[key2], callback_data=f"menu:{key2}",
                style=_USER_MENU_STYLES.get(key2, "primary"),
            ),
        ])
    buttons.append([InlineKeyboardButton(text="❌ بستن پنل", callback_data="menu:close", style="danger")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ---------- پنل ادمین ----------
def admin_panel_keyboard() -> InlineKeyboardMarkup:
    bot_enabled = load_bot_state().get("enabled", True)
    if bot_enabled:
        toggle_label = "🔴 خاموش کردن ربات"
        toggle_style = "danger"
    else:
        toggle_label = "🟢 روشن کردن ربات"
        toggle_style = "success"

    # صفحه‌ی اصلیِ پنل فقط دسته‌بندی‌هاست، نه همه‌ی ۱۲ اکشن با هم؛ هر دسته
    # زیرمنوی خودش را دارد (همان الگویی که «مدیریت محتوا» و «تنظیمات VIP»
    # قبلاً داشتند) تا صفحه سبک بماند و اکشن‌های حساس (حذف کاربر، بازیابیِ
    # بکاپ) از کارهای روتین جدا و در جای خودشان دیده شوند.
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📊 گزارش‌ها", callback_data="admin:cat_reports", style="primary"),
                InlineKeyboardButton(text="📨 پیام‌رسانی", callback_data="admin:cat_messaging", style="primary"),
            ],
            [
                InlineKeyboardButton(text="👥 کاربران و محتوا", callback_data="admin:cat_users", style="primary"),
                InlineKeyboardButton(text="💎 VIP", callback_data="admin:cat_vip", style="success"),
            ],
            [
                InlineKeyboardButton(text="⚙️ بکاپ و سیستم", callback_data="admin:cat_backup", style="primary"),
            ],
            [
                InlineKeyboardButton(text=toggle_label, callback_data="admin:toggle_bot", style=toggle_style),
            ],
            [
                InlineKeyboardButton(text="❌ بستن", callback_data="admin:close", style="danger"),
            ],
        ]
    )

def admin_reports_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📈 آمار تفصیلی", callback_data="admin:stats_detail", style="primary"),
                InlineKeyboardButton(text="📄 خروجی اکسل", callback_data="admin:export", style="primary"),
            ],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

def admin_messaging_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📢 پیام همگانی", callback_data="admin:broadcast", style="success"),
                InlineKeyboardButton(text="📨 ارسال مستقیم", callback_data="admin:sendmsg", style="primary"),
            ],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

def admin_users_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="🛠 مدیریت محتوا", callback_data="admin:menu_edit", style="primary"),
                InlineKeyboardButton(text="🗑 حذف کاربر", callback_data="admin:delete_user", style="danger"),
            ],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

def admin_vip_category_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="💎 تنظیمات VIP", callback_data="admin:vip_settings", style="success"),
                InlineKeyboardButton(text="💰 تنظیم قیمت اشتراک", callback_data="admin:vip_global_settings", style="primary"),
            ],
            [InlineKeyboardButton(text="📋 مشترکینِ VIP", callback_data="vipadmin:list:0", style="primary")],
            # >>> اینجا <<< ردیفِ جدید برای پاداشِ دستی
            [InlineKeyboardButton(text="🎁 اهدای پاداش VIP", callback_data="vipreward:start", style="success")],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

def admin_backup_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📥 گرفتن بکاپ", callback_data="admin:manual_backup", style="success"),
                InlineKeyboardButton(text="🔁 بازیابی از بکاپ", callback_data="admin:restore_backup", style="danger"),
            ],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

def admin_back_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")]]
    )

def admin_menu_edit_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📢 ویرایش اطلاعیه‌ها", callback_data="admin:edit_announcements", style="primary")],
            [InlineKeyboardButton(text="❓ ویرایش سوالات متداول", callback_data="admin:edit_faq", style="primary")],
            [InlineKeyboardButton(text="🔙 بازگشت به منو", callback_data="admin:menu", style="primary")],
        ]
    )

# ---------- مدیریت وضعیت ربات ----------
def load_bot_state() -> dict:
    if not BOT_STATE_FILE.exists():
        return {"enabled": True, "pending_requests": []}
    try:
        return json.loads(BOT_STATE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"enabled": True, "pending_requests": []}

async def save_bot_state(state: dict) -> None:
    async with _write_lock:
        BOT_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")

# ---------- ردیابیِ درخواست‌های عضویتِ معلق (کسانی که پیامِ قوانین را دیده‌اند
# ولی دکمه‌ی «قبول دارم» را نزده‌اند) — برای یادآوریِ خودکار و اطلاعِ ادمین
# در صورتِ ادامه‌ی سکوت، به‌جایِ چک‌کردنِ دستیِ دوره‌ای. ----------
def load_pending_joins() -> dict:
    if not PENDING_JOIN_FILE.exists():
        return {}
    try:
        return json.loads(PENDING_JOIN_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

async def save_pending_joins(data: dict) -> None:
    async with _write_lock:
        PENDING_JOIN_FILE.parent.mkdir(exist_ok=True)
        PENDING_JOIN_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

async def _track_pending_join(user) -> None:
    pending = load_pending_joins()
    pending[str(user.id)] = {
        "full_name": user.full_name,
        "username": user.username,
        "requested_at": datetime.utcnow().isoformat(),
        "reminded": False,
        "escalated": False,
    }
    await save_pending_joins(pending)

async def _untrack_pending_join(user_id: int) -> None:
    pending = load_pending_joins()
    if str(user_id) in pending:
        pending.pop(str(user_id))
        await save_pending_joins(pending)

async def pending_join_checker_loop() -> None:
    while True:
        try:
            await _check_pending_joins()
        except Exception as e:
            logger.error("خطا در بررسیِ درخواست‌های عضویتِ معلق: %s", e)
        await asyncio.sleep(3 * 3600)

async def _check_pending_joins() -> None:
    """
    برای کسانی که پیامِ قوانین را دریافت کرده‌اند ولی مدتی طولانی روی «قبول
    دارم» نزده‌اند: بعد از JOIN_REMINDER_AFTER_HOURS یک یادآوری خودکار
    می‌فرستیم، و اگر باز هم بعد از JOIN_ESCALATE_AFTER_HOURS اقدامی نکرده
    باشند، به ادمین اطلاع می‌دهیم تا خودش تصمیم بگیرد — به‌جایِ اینکه ادمین
    مجبور باشد هرچند وقت یک‌بار به‌صورتِ دستی لیستِ درخواست‌های معلق را چک کند.
    """
    pending = load_pending_joins()
    if not pending:
        return
    now = datetime.utcnow()
    changed = False

    for user_id_str, info in list(pending.items()):
        try:
            requested_at = datetime.fromisoformat(info["requested_at"])
        except (KeyError, ValueError):
            continue
        hours_passed = (now - requested_at).total_seconds() / 3600
        user_id = int(user_id_str)

        if hours_passed >= JOIN_ESCALATE_AFTER_HOURS and not info.get("escalated"):
            info["escalated"] = True
            changed = True
            if NOTIFY_CHAT_ID_INT:
                try:
                    username_part = f"@{info['username']}" if info.get("username") else "بدونِ‌یوزرنیم"
                    await bot.send_message(
                        chat_id=NOTIFY_CHAT_ID_INT,
                        text=(
                            "🕐 <b>درخواستِ عضویتِ معلق</b>\n\n"
                            f"👤 {html_escape(info.get('full_name', ''))} ({username_part})\n"
                            f"🆔 <code>{user_id}</code>\n\n"
                            f"بیش از {to_persian_num(JOIN_ESCALATE_AFTER_HOURS)} ساعت از ارسالِ قوانین "
                            "می‌گذرد و هنوز «قبول دارم» را نزده. می‌توانید دستی تصمیم بگیرید:"
                        ),
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                            [
                                InlineKeyboardButton(text="✅ تاییدِ عضویت", callback_data=f"adminjoin:approve:{user_id}", style="success"),
                                InlineKeyboardButton(text="❌ ردِ درخواست", callback_data=f"adminjoin:reject:{user_id}", style="danger"),
                            ],
                        ]),
                    )
                except Exception as e:
                    logger.warning("اطلاع‌رسانیِ درخواستِ معلقِ کاربر %s به ادمین ممکن نشد: %s", user_id, e)

        elif hours_passed >= JOIN_REMINDER_AFTER_HOURS and not info.get("reminded"):
            info["reminded"] = True
            changed = True
            try:
                await bot.send_message(
                    chat_id=user_id,
                    text=sign(
                        "⏳ <b>یادآوری</b>\n\n"
                        "هنوز روی دکمه‌ی «قوانین را قبول دارم» نزده‌اید؛ برای تکمیلِ عضویت در رواق، "
                        "کافی‌ست همان دکمه‌ی زیرِ پیامِ قبلی را بزنید."
                    ),
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[[
                            InlineKeyboardButton(
                                text="✅ قوانین را قبول دارم و عضو می‌شوم",
                                callback_data="rules_accept",
                                style="success",
                            )
                        ]]
                    ),
                )
            except Exception as e:
                logger.warning("ارسالِ یادآوریِ عضویت به کاربر %s ممکن نشد: %s", user_id, e)

    if changed:
        await save_pending_joins(pending)

# ---------- ارسالِ متنِ قوانین همراه با دکمهٔ پذیرش ----------
async def send_rules_message(user) -> bool:
    """پیامِ قوانین را برای کاربر می‌فرستد. اگر ربات هرگز از سمتِ این کاربر
    استارت نشده باشد، تلگرام اجازه‌ی شروعِ گفتگو را به ربات نمی‌دهد و ارسال
    شکست می‌خورد — این تابع در آن صورت False برمی‌گرداند تا فراخوان بتواند
    ادمین را مطلع کند (وگرنه درخواستِ عضویت بدونِ هیچ اطلاعی برای همیشه معلق می‌ماند)."""
    try:
        await bot.send_message(
            chat_id=user.id,
            text=GROUP_RULES_TEXT,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[
                    InlineKeyboardButton(
                        text="✅ قوانین را قبول دارم و عضو می‌شوم",
                        callback_data="rules_accept",
                        style="success",
                    )
                ]]
            ),
        )
        return True
    except Exception as e:
        logger.warning("ارسالِ پیامِ قوانین به کاربر %s ممکن نشد: %s", user.id, e)
        return False

async def _finalize_group_approval(user_id: int, notify_user: bool = True) -> bool:
    """پس از تاییدِ درخواستِ عضویت (چه با کلیکِ خودِ کاربر، چه دستیِ ادمین)،
    مراحلِ مشترک را انجام می‌دهد: تاییدِ واقعیِ عضویت در تلگرام، ثبتِ آمار،
    و تلاش برای خوش‌آمدگویی به کاربر."""
    try:
        await bot.approve_chat_join_request(chat_id=GROUP_CHAT_ID, user_id=user_id)
    except Exception as e:
        logger.warning("تاییدِ عضویتِ کاربر %s ممکن نشد: %s", user_id, e)
        return False

    await mark_verified(user_id)
    await increment_stat("form_completed_and_joined")
    await _untrack_pending_join(user_id)

    if notify_user:
        try:
            user = await bot.get_chat(user_id)
            await bot.send_message(
                chat_id=user_id,
                text=sign(f"{greet_user(user)}، به رواق خوش آمدید 🏛\n\nاز پنل زیر یکی از گزینه‌ها را انتخاب کنید:"),
                reply_markup=user_panel_keyboard(),
            )
        except Exception as e:
            logger.warning("ارسالِ پیامِ خوش‌آمدگویی به کاربر %s ممکن نشد: %s", user_id, e)

    _schedule_vip_intro(user_id)
    return True

async def process_pending_requests():
    state = load_bot_state()
    pending = state.get("pending_requests", [])
    if not pending:
        return
    logger.info("شروع پردازش %d درخواست معلق", len(pending))
    for user_id in pending:
        try:
            user = await bot.get_chat(user_id)
            await bot.send_message(
                chat_id=user.id,
                text=sign("🟢 <b>ربات فعال شد</b>\n\nدرخواستِ عضویتِ شما اکنون در حالِ پردازش است."),
            )
            await asyncio.sleep(1)
            await send_rules_message(user)
            logger.info("پیام به کاربر %s ارسال شد", user_id)
        except Exception as e:
            logger.warning("پردازش درخواست معلق برای %s ناموفق: %s", user_id, e)
        await asyncio.sleep(2)
    state["pending_requests"] = []
    await save_bot_state(state)
    logger.info("پردازش درخواست‌های معلق پایان یافت")

# ---------- نشانگر تایپ ----------
async def send_with_action(chat_id: int, action: str = "typing", delay: float = 1.0):
    await bot.send_chat_action(chat_id=chat_id, action=action)
    if delay > 0:
        await asyncio.sleep(delay)

# ---------- باز کردنِ پنل‌ها (برای استارتِ معمولی و لینکِ مستقیم) ----------
async def open_user_panel(chat_id: int) -> None:
    await bot.send_message(
        chat_id=chat_id,
        text="🏛 <b>به رواق خوش آمدید</b>\n\nاز پنل زیر یکی از گزینه‌ها را انتخاب کنید:",
        reply_markup=user_panel_keyboard(),
    )

async def open_vip_panel(chat_id: int) -> None:
    if VIP_GROUP_CHAT_ID is None:
        await bot.send_message(chat_id=chat_id, text="🌟 گروه VIP هنوز راه‌اندازی نشده است.")
        return
    caption, keyboard, image_id = await render_vip_page(0)
    if image_id:
        await bot.send_photo(
            chat_id=chat_id, photo=image_id, caption=caption,
            show_caption_above_media=True, reply_markup=keyboard,
        )
    else:
        await bot.send_message(chat_id=chat_id, text=caption, reply_markup=keyboard)

# ---------- دستور /start ----------
@dp.message(Command("start"))
async def handle_start(message: Message, command: CommandObject):
    user_id = message.from_user.id
    await mark_funnel_entry(user_id)
    await send_with_action(message.chat.id, "typing", 0.5)

    is_member = await is_user_member(user_id)

    if is_member:
        # لینکِ مستقیمِ استارت به دکمه‌ی گروهِ VIP: t.me/<bot_username>?start=vip
        if (command.args or "").strip().lower() == "vip":
            await open_vip_panel(message.chat.id)
        else:
            await open_user_panel(message.chat.id)
        return

    member_count_line = ""
    try:
        member_count = await bot.get_chat_member_count(GROUP_CHAT_ID)
        member_count_str = to_persian_num(member_count)
        member_count_line = f"هم‌اکنون <b>{member_count_str}</b> معمار و مهندس در اینجا حضور دارند.\n\n"
    except Exception:
        pass
    await message.answer(
        sign(
            f"{greet_user(message.from_user)}، به {GROUP_NAME} خوش آمدید.\n\n"
            "اینجا انبارِ تخصصیِ فایل‌های معماری و عمران است.\n"
            f"{member_count_line}"
            "برای عضویت، کافی‌ست درخواستِ پیوستن به گروه را ثبت کنید.\n"
            "مسیرِ بعدی برای شما گشوده خواهد شد."
        ),
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="📝 ثبت درخواست عضویت", url=GROUP_INVITE_LINK, style="success")],
            ]
        )
    )

# ==============================================================
#  درخواستِ عضویت و پذیرشِ قوانین
# ==============================================================

@dp.chat_join_request()
async def handle_join_request(join_request: ChatJoinRequest):
    if join_request.chat.id != GROUP_CHAT_ID:
        return

    user = join_request.from_user
    logger.info("درخواست عضویت جدید از %s (%s)", user.full_name, user.id)
    await mark_funnel_entry(user.id)

    state = load_bot_state()
    if not state.get("enabled", True):
        try:
            await bot.send_message(
                chat_id=user.id,
                text=(
                    "🔴 <b>عضویت موقتاً بسته است</b>\n\n"
                    "درِ رواق برای عضوگیری بسته شده.\n"
                    "به محضِ فعال‌سازی، به شما اطلاع داده خواهد شد."
                )
            )
        except Exception as e:
            logger.warning("ارسال پیام خاموشی به کاربر %s ممکن نشد: %s", user.id, e)
        pending = state.get("pending_requests", [])
        if user.id not in pending:
            pending.append(user.id)
            state["pending_requests"] = pending
            await save_bot_state(state)
        return

    sent_ok = await send_rules_message(user)

    if not sent_ok:
        # اگه کاربر تا حالا هیچ‌وقت ربات را استارت نکرده باشه، تلگرام به ربات
        # اجازه‌ی شروعِ گفتگو رو نمی‌ده و پیامِ قوانین اصلاً ارسال نمی‌شه — یعنی
        # درخواستِ عضویت بدونِ هیچ اطلاعی به کاربر یا ادمین، معلق می‌مونه.
        # برای همین اینجا به ادمین خبر می‌دیم تا بتونه دستی تصمیم بگیره.
        if NOTIFY_CHAT_ID_INT:
            try:
                username_part = f"@{user.username}" if user.username else "بدونِ‌یوزرنیم"
                await bot.send_message(
                    chat_id=NOTIFY_CHAT_ID_INT,
                    text=(
                        "⚠️ <b>پیامِ قوانین برای این کاربر ارسال نشد</b>\n"
                        "(احتمالاً چون تا حالا ربات را استارت نکرده)\n\n"
                        f"👤 {html_escape(user.full_name)} ({username_part})\n"
                        f"🆔 <code>{user.id}</code>\n\n"
                        "می‌توانید دستی تصمیم بگیرید:"
                    ),
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [
                            InlineKeyboardButton(text="✅ تاییدِ عضویت", callback_data=f"adminjoin:approve:{user.id}", style="success"),
                            InlineKeyboardButton(text="❌ ردِ درخواست", callback_data=f"adminjoin:reject:{user.id}", style="danger"),
                        ],
                    ]),
                )
            except Exception as e:
                logger.error("اطلاع‌رسانیِ شکستِ ارسالِ قوانین به ادمین ممکن نشد: %s", e)
        return

    # پیامِ قوانین با موفقیت رسید؛ حالا این درخواست را ردیابی می‌کنیم تا اگر
    # کاربر مدتی طولانی روی دکمه‌ی «قبول دارم» نزد، خودمان یادآوری بفرستیم و
    # در صورتِ ادامه‌ی سکوت، به ادمین اطلاع بدیم — به‌جایِ تکیه بر چک‌کردنِ دستی.
    await _track_pending_join(user)

@dp.callback_query(F.data.startswith("adminjoin:approve:"))
async def cb_adminjoin_approve(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    user_id = int(callback.data.split(":")[2])
    ok = await _finalize_group_approval(user_id, notify_user=True)
    if ok:
        try:
            await callback.message.edit_text(
                callback.message.text + f"\n\n✅ توسطِ {html_escape(callback.from_user.full_name)} تایید شد.",
                reply_markup=None,
            )
        except Exception:
            pass
        await callback.answer("✅ عضویت تایید شد.")
    else:
        await callback.answer("❌ تاییدِ عضویت ناموفق بود (شاید درخواست قبلاً منقضی شده).", show_alert=True)

@dp.callback_query(F.data.startswith("adminjoin:reject:"))
async def cb_adminjoin_reject(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    user_id = int(callback.data.split(":")[2])
    try:
        await bot.decline_chat_join_request(chat_id=GROUP_CHAT_ID, user_id=user_id)
    except Exception as e:
        logger.warning("ردِ درخواستِ عضویتِ کاربر %s ممکن نشد: %s", user_id, e)
    await _untrack_pending_join(user_id)
    try:
        await callback.message.edit_text(
            callback.message.text + f"\n\n❌ توسطِ {html_escape(callback.from_user.full_name)} رد شد.",
            reply_markup=None,
        )
    except Exception:
        pass
    await callback.answer("❌ درخواست رد شد.")

@dp.callback_query(F.data == "rules_accept")
async def cb_rules_accept(callback: CallbackQuery):
    user = callback.from_user
    approved = await _finalize_group_approval(user.id, notify_user=False)

    if approved:
        try:
            await callback.message.edit_text(
                "✅ <b>خوش آمدید!</b>\n\nقوانین پذیرفته شد و عضویتِ شما در رواق تایید شد."
            )
        except Exception:
            pass
        await callback.answer("عضویت تایید شد ✅")
        await bot.send_message(
            chat_id=user.id,
            text=sign(f"{greet_user(user)}، به رواق خوش آمدید 🏛\n\nاز پنل زیر یکی از گزینه‌ها را انتخاب کنید:"),
            reply_markup=user_panel_keyboard(),
        )
    else:
        await callback.answer(
            "❌ تاییدِ عضویت با مشکلی مواجه شد. کمی صبر کنید یا از طریق «ارتباط با ادمین» پیگیری کنید.",
            show_alert=True,
        )

# ==============================================================
#  پروفایلِ من — ارتقا به کاربرِ طلایی
# ==============================================================
_pending_profile: dict[int, dict] = {}

def education_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    for i in range(0, len(EDUCATION_OPTIONS), 2):
        row = [InlineKeyboardButton(text=EDUCATION_OPTIONS[i][1], callback_data=f"profedu:{EDUCATION_OPTIONS[i][0]}", style="primary")]
        if i + 1 < len(EDUCATION_OPTIONS):
            row.append(InlineKeyboardButton(text=EDUCATION_OPTIONS[i+1][1], callback_data=f"profedu:{EDUCATION_OPTIONS[i+1][0]}", style="primary"))
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back", style="danger")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def referral_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    items = list(REFERRAL_LABELS.items())
    for i in range(0, len(items), 2):
        row = [InlineKeyboardButton(text=items[i][1], callback_data=f"profref:{items[i][0]}", style="primary")]
        if i + 1 < len(items):
            row.append(InlineKeyboardButton(text=items[i+1][1], callback_data=f"profref:{items[i+1][0]}", style="primary"))
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="🔙 بازگشت به سوالِ قبل", callback_data="profback:edu", style="danger")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def interests_keyboard(selected: list[str]) -> InlineKeyboardMarkup:
    buttons = []
    for i in range(0, len(INTERESTS), 2):
        row = []
        for item in INTERESTS[i:i + 2]:
            is_sel = item in selected
            row.append(InlineKeyboardButton(
                text=f"✅ {item}" if is_sel else item,
                callback_data=f"profint:{item}",
                style="success" if is_sel else "primary",
            ))
        buttons.append(row)
    buttons.append([InlineKeyboardButton(
        text=f"🏁 ثبتِ نهایی ({to_persian_num(len(selected))}/{to_persian_num(MAX_INTERESTS)})",
        callback_data="profint_done",
        style="success" if selected else "primary",
    )])
    buttons.append([InlineKeyboardButton(text="🔙 بازگشت به سوالِ قبل", callback_data="profback:ref", style="danger")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

async def start_profile_form(user, chat_id: int, message_id: int | None = None) -> None:
    _pending_profile[user.id] = {}
    text = (
        "👤 <b>تکمیلِ پروفایل</b>\n\n"
        "با پاسخ به سه سوالِ کوتاه، پروفایل‌تان تکمیل می‌شود و به «🥇 کاربرِ طلایی» رواق ارتقا پیدا می‌کنید.\n\n"
        f"{progress_bar(1)}  سوال ۱ از ۳ — سطح تحصیلی شما؟"
    )
    keyboard = education_keyboard()
    if message_id:
        try:
            await bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, reply_markup=keyboard)
            return
        except Exception:
            pass
    await bot.send_message(chat_id=chat_id, text=text, reply_markup=keyboard)

@dp.callback_query(F.data == "profile:start")
async def cb_profile_start(callback: CallbackQuery):
    await start_profile_form(callback.from_user, callback.message.chat.id, callback.message.message_id)
    await callback.answer()

@dp.callback_query(F.data == "profile:edit")
async def cb_profile_edit(callback: CallbackQuery):
    await start_profile_form(callback.from_user, callback.message.chat.id, callback.message.message_id)
    await callback.answer()

@dp.callback_query(F.data.startswith("profedu:"))
async def cb_profile_education(callback: CallbackQuery):
    user = callback.from_user
    value = callback.data.split(":", 1)[1]
    label = dict(EDUCATION_OPTIONS).get(value, value)
    _pending_profile[user.id] = {"education": value, "education_label": label}
    await callback.message.edit_text(
        f"✅ گزینه‌ی <b>{label}</b> ثبت شد.\n\n"
        f"{progress_bar(2)}  سوال ۲ از ۳ — چگونه با رواق آشنا شدید؟",
        reply_markup=referral_keyboard(),
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("profref:"))
async def cb_profile_referral(callback: CallbackQuery):
    user = callback.from_user
    value = callback.data.split(":", 1)[1]
    data = _pending_profile.setdefault(user.id, {})
    data["referral"] = value
    data["interests"] = set()
    label = REFERRAL_LABELS.get(value, value)
    await callback.message.edit_text(
        f"✅ گزینه‌ی <b>{label}</b> ثبت شد.\n\n"
        f"{progress_bar(3)}  سوال ۳ از ۳ — کدام بخش از رواق برای شما جذاب‌تر است؟\n"
        f"(حداکثر {to_persian_num(MAX_INTERESTS)} مورد)",
        reply_markup=interests_keyboard([]),
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("profint:"))
async def cb_profile_interest_toggle(callback: CallbackQuery):
    user = callback.from_user
    item = callback.data.split(":", 1)[1]
    data = _pending_profile.setdefault(user.id, {})
    selected: set = data.setdefault("interests", set())
    if item in selected:
        selected.discard(item)
    elif len(selected) >= MAX_INTERESTS:
        await callback.answer(f"حداکثر {to_persian_num(MAX_INTERESTS)} مورد را می‌توانید انتخاب کنید.", show_alert=True)
        return
    else:
        selected.add(item)
    await callback.message.edit_reply_markup(reply_markup=interests_keyboard(list(selected)))
    await callback.answer()

@dp.callback_query(F.data == "profback:edu")
async def cb_profile_back_to_education(callback: CallbackQuery):
    user = callback.from_user
    _pending_profile[user.id] = {}
    await callback.message.edit_text(
        f"{progress_bar(1)}  سوال ۱ از ۳ — سطح تحصیلی شما؟",
        reply_markup=education_keyboard(),
    )
    await callback.answer()

@dp.callback_query(F.data == "profback:ref")
async def cb_profile_back_to_referral(callback: CallbackQuery):
    user = callback.from_user
    data = _pending_profile.setdefault(user.id, {})
    data.pop("referral", None)
    data.pop("interests", None)
    await callback.message.edit_text(
        f"{progress_bar(2)}  سوال ۲ از ۳ — چگونه با رواق آشنا شدید؟",
        reply_markup=referral_keyboard(),
    )
    await callback.answer()

# ---------- ساختِ کارتِ پروفایل ----------
def build_profile_card(user, data: dict, jalali_now: str) -> str:
    display_name = html_escape(user.full_name or user.first_name or "کاربر")
    interests_text = '، '.join(data.get('interests', []))
    card = (
        "🥇 <b>تبریک! شما کاربرِ طلاییِ رواق شدید</b>\n\n"
        f"👤 {display_name}\n"
        f"🎓 {data.get('education_label', '')}\n"
        f"⭐️ علایق: {interests_text}\n"
        f"🗓 {jalali_now}\n\n"
        "هر زمان بخواهید می‌توانید از همین بخش، اطلاعاتِ پروفایل‌تان را دوباره ویرایش کنید."
    )
    return sign(card)

def profile_result_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ ویرایشِ مجدد", callback_data="profile:edit", style="primary")],
        [InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back", style="primary")],
    ])

@dp.callback_query(F.data == "profint_done")
async def cb_profile_submit(callback: CallbackQuery):
    user = callback.from_user
    data = _pending_profile.get(user.id)
    if not data or not data.get("education") or not data.get("referral"):
        await callback.answer("انگار مسیر قطع شده. لطفاً دوباره از «پروفایل من» شروع کنید.", show_alert=True)
        return
    selected = list(data.get("interests") or [])
    if not selected:
        await callback.answer("حداقل یک مورد را انتخاب کنید.", show_alert=True)
        return

    await callback.answer("⏳ در حال ثبت...")

    record = {
        "user_id": user.id,
        "username": user.username,
        "full_name": f"{user.first_name or ''} {user.last_name or ''}".strip(),
        "submitted_at": datetime.utcnow().isoformat(),
        "education": data["education"],
        "education_label": data["education_label"],
        "referral": data["referral"],
        "interests": selected,
    }

    async with _write_lock:
        with open(DATA_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    _user_cache[str(user.id)] = record
    _pending_profile.pop(user.id, None)
    logger.info("پروفایلِ کاربر %s ذخیره شد.", user.id)

    jalali_now = format_jalali_datetime(datetime.utcnow())
    card = build_profile_card(user, record, jalali_now)
    await callback.message.edit_text(card, reply_markup=profile_result_keyboard())

# ---------- رویداد تغییر وضعیت عضو ----------
@dp.chat_member()
async def handle_chat_member_update(update: ChatMemberUpdated):
    if update.chat.id != GROUP_CHAT_ID:
        return

    old_status = update.old_chat_member.status
    new_status = update.new_chat_member.status
    user = update.new_chat_member.user

    if user.is_bot:
        return

    became_member = new_status == ChatMemberStatus.MEMBER and old_status != ChatMemberStatus.MEMBER
    if became_member:
        await increment_stat("total_joined")
        await notify_new_member(user)
        await send_welcome_to_group(user)
        return

    left_group = (
        old_status == ChatMemberStatus.MEMBER
        and new_status in (ChatMemberStatus.LEFT, ChatMemberStatus.KICKED)
    )
    if left_group:
        await increment_stat("total_left")
        await handle_member_left(user)

async def notify_new_member(user) -> None:
    if not NOTIFY_CHAT_ID:
        return
    username_part = f"@{user.username}" if user.username else f"<code>{user.id}</code>"
    try:
        await bot.send_message(
            chat_id=NOTIFY_CHAT_ID,
            text=(
                f"✅ عضو جدید به گروه پیوست:\n"
                f"👤 {user.full_name} ({username_part})"
            ),
        )
    except Exception as e:
        logger.warning("ارسال گزارش عضو جدید ممکن نشد: %s", e)

async def send_welcome_to_group(user) -> None:
    display_name = html_escape(user.full_name or user.first_name or "کاربر")
    user_mention = f"<a href='tg://user?id={user.id}'>{display_name}</a>"
    try:
        sent = await bot.send_message(
            chat_id=GROUP_CHAT_ID,
            text=(
                f"{user_mention} عزیز خوش آمدید 👋\n\n"
                "▫️ اینجا انباری از فایل‌های تخصصی معماری و عمران است.\n"
                "▫️ برای شروع، خود را در تاپیک <a href='https://t.me/c/4388421316/95'>کافه معماری</a> معرفی کنید.\n\n"
                "🏛 آماده‌اید برای پیشرفت؟"
            ),
            parse_mode=ParseMode.HTML,
            disable_notification=True,
        )
    except Exception as e:
        logger.warning("ارسال پیام خوش‌آمدگویی به گروه ممکن نشد: %s", e)
        return

    asyncio.create_task(_delete_message_later(sent.chat.id, sent.message_id, delay=30))

async def _delete_message_later(chat_id: int, message_id: int, delay: int) -> None:
    await asyncio.sleep(delay)
    try:
        await bot.delete_message(chat_id=chat_id, message_id=message_id)
    except Exception as e:
        logger.warning("حذفِ خودکارِ پیامِ خوش‌آمدگویی ممکن نشد: %s", e)

async def handle_member_left(user) -> None:
    if NOTIFY_CHAT_ID:
        username_part = f"@{user.username}" if user.username else f"<code>{user.id}</code>"
        try:
            await bot.send_message(
                chat_id=NOTIFY_CHAT_ID,
                text=f"🚪 یک عضو گروه را ترک کرد:\n👤 {user.full_name} ({username_part})",
            )
        except Exception as e:
            logger.warning("ارسال گزارش ترک عضو ممکن نشد: %s", e)

    try:
        await bot.send_message(
            chat_id=user.id,
            text=(
                "متأسفیم که از جمعِ رواق فاصله گرفتید.\n\n"
                "اگر فرصتی باشد، مایلیم بدانیم چه عاملی باعث ترک شما شد.\n"
                "نظرات شما به ما در بهبود این فضا کمک خواهد کرد."
            ),
        )

        sent_poll = await bot.send_poll(
            chat_id=user.id,
            question="چرا این بنا را ترک کردی؟",
            options=[reason for reason, _ in LEAVE_REASONS],
            is_anonymous=False,
        )
        _pending_leave_polls[sent_poll.poll.id] = user.id

        keyboard = None
        if GROUP_INVITE_LINK:
            keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="بازگشت به گروه ↩️", url=GROUP_INVITE_LINK)]
                ]
            )
        await bot.send_message(
            chat_id=user.id,
            text="هر زمان خواستید، درِ رواق به روی شما باز است 🏛",
            reply_markup=keyboard,
        )
    except Exception as e:
        logger.warning("نمی‌توان به کاربر خارج‌شده %s پیام داد: %s", user.id, e)

@dp.poll_answer()
async def handle_leave_poll_answer(poll_answer: PollAnswer):
    user_id = _pending_leave_polls.pop(poll_answer.poll_id, None)
    if user_id is None or not poll_answer.option_ids:
        return

    option_index = poll_answer.option_ids[0]
    if option_index >= len(LEAVE_REASONS):
        return

    _, reply_text = LEAVE_REASONS[option_index]
    try:
        await bot.send_message(chat_id=user_id, text=reply_text)
    except Exception as e:
        logger.warning("ارسال پاسخ نظرسنجی به کاربر %s ممکن نشد: %s", user_id, e)

# ---------- پنل مدیریت ----------
@dp.message(Command("admin"))
async def handle_admin_panel(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await send_with_action(message.chat.id, "typing", 0.5)
    await message.answer(
        await build_admin_dashboard_text(),
        reply_markup=admin_panel_keyboard(),
    )

@dp.message(Command("stats"))
async def handle_stats(message: Message):
    if not is_admin(message.from_user.id):
        return
    await send_with_action(message.chat.id, "typing", 1.0)
    await message.answer(await build_stats_text())

@dp.message(Command("stats_detail"))
async def handle_stats_detail(message: Message):
    if not is_admin(message.from_user.id):
        return
    await send_with_action(message.chat.id, "typing", 1.0)
    await message.answer(await build_stats_detail_text())

@dp.message(Command("export"))
async def handle_export(message: Message):
    if not is_admin(message.from_user.id):
        return
    await send_with_action(message.chat.id, "upload_document", 0.5)
    file = build_export_file()
    if file is None:
        await message.answer("هنوز هیچ کاربری شماره‌اش را تأیید نکرده است.")
        return
    await message.answer_document(file, caption="📄 خروجی اکسل همه‌ی تأییدشده‌ها")

@dp.message(Command("broadcast"))
async def handle_broadcast(message: Message, command: CommandObject):
    if not is_admin(message.from_user.id):
        return
    text = (command.args or "").strip()
    if not text:
        await message.answer(
            "برای ارسال پیام همگانی به این شکل دستور را بفرستید:\n"
            "<code>/broadcast متن پیام شما</code>\n\n"
            "یا از پنل شیشه‌ای با دستور /admin استفاده کنید (که از عکس و فایل هم پشتیبانی می‌کند)."
        )
        return

    user_ids = load_funnel_users()
    if not user_ids:
        await message.answer("هیچ کاربری برای ارسال پیدا نشد.")
        return

    await message.answer(f"⏳ در حال ارسال پیام به {to_persian_num(len(user_ids))} نفر...")
    sent, failed = await send_broadcast_text(text, user_ids)
    await message.answer(
        f"✅ ارسال همگانی تمام شد.\n"
        f"موفق: <b>{to_persian_num(sent)}</b>\n"
        f"ناموفق: <b>{to_persian_num(failed)}</b>"
    )

async def send_broadcast_text(text: str, user_ids: set[int]) -> tuple[int, int]:
    sent, failed = 0, 0
    for user_id in user_ids:
        try:
            await bot.send_message(chat_id=user_id, text=text)
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.05)
    return sent, failed

# ---------- هندلر واحد برای تمام کالبک‌های ادمین ----------
@dp.callback_query(F.data.startswith("admin:"))
async def handle_all_admin_callbacks(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return

    action = callback.data.split(":", 1)[1]
    logger.info(f"ادمین {callback.from_user.id} درخواست {action} کرد")

    if action == "menu":
        await state.clear()
        await callback.message.edit_text(
            await build_admin_dashboard_text(),
            reply_markup=admin_panel_keyboard(),
        )
        await callback.answer()
        return

    if action == "cat_reports":
        await callback.answer()
        await callback.message.edit_text(
            "📊 <b>گزارش‌ها</b>\n\nیکی از گزینه‌ها را انتخاب کنید.",
            reply_markup=admin_reports_keyboard(),
        )
        return

    if action == "cat_messaging":
        await callback.answer()
        await callback.message.edit_text(
            "📨 <b>پیام‌رسانی</b>\n\nیکی از گزینه‌ها را انتخاب کنید.",
            reply_markup=admin_messaging_keyboard(),
        )
        return

    if action == "cat_users":
        await callback.answer()
        await callback.message.edit_text(
            "👥 <b>کاربران و محتوا</b>\n\nیکی از گزینه‌ها را انتخاب کنید.",
            reply_markup=admin_users_keyboard(),
        )
        return

    if action == "cat_vip":
        await callback.answer()
        await callback.message.edit_text(
            "💎 <b>مدیریت VIP</b>\n\nیکی از گزینه‌ها را انتخاب کنید.",
            reply_markup=admin_vip_category_keyboard(),
        )
        return

    if action == "cat_backup":
        await callback.answer()
        await callback.message.edit_text(
            "⚙️ <b>بکاپ و سیستم</b>\n\nیکی از گزینه‌ها را انتخاب کنید.",
            reply_markup=admin_backup_keyboard(),
        )
        return

    if action == "close":
        await state.clear()
        try:
            await callback.message.delete()
        except Exception:
            pass
        await callback.answer("پنل بسته شد.")
        return

    if action == "stats_detail":
        await callback.answer()
        reports_back_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="🔙 بازگشت", callback_data="admin:cat_reports", style="primary")]]
        )
        await callback.message.edit_text(await build_stats_detail_text(), reply_markup=reports_back_keyboard)
        return

    if action == "export":
        await callback.answer("⏳ در حال ساخت فایل اکسل...")
        await send_with_action(callback.message.chat.id, "upload_document", 1.0)
        file = build_export_file()
        if file is None:
            await callback.message.answer("هنوز هیچ کاربری شماره‌اش را تأیید نکرده است.")
            return
        await callback.message.answer_document(file, caption="📄 خروجی اکسل همه‌ی تأییدشده‌ها")
        return

    if action == "manual_backup":
        await callback.answer("⏳ در حال گرفتن بکاپ...")
        ok, backup_msg = await backup_data_dir_to_telegram()
        icon = "✅" if ok else "❌"
        await callback.message.answer(
            f"{icon} {backup_msg}\n🕐 {format_jalali_datetime(datetime.utcnow())}"
        )
        return

    if action == "restore_backup":
        await callback.answer()
        confirm_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✅ بله، بازیابی شود", callback_data="admin:restore_backup_confirm", style="danger")],
                [InlineKeyboardButton(text="🔙 انصراف", callback_data="admin:cat_backup", style="primary")],
            ]
        )
        await callback.message.answer(
            "⚠️ <b>بازیابیِ دستی از بکاپ</b>\n\n"
            "این کار تمامِ دیتای محلیِ فعلی را با آخرین بکاپِ پین‌شده در تلگرام جایگزین می‌کند.\n"
            "مطمئنی؟",
            reply_markup=confirm_keyboard,
        )
        return

    if action == "restore_backup_confirm":
        await callback.answer("⏳ در حال بازیابی...")
        ok, restore_msg = await restore_data_dir_from_telegram(force=True)
        if ok:
            storage.reload()  # وضعیتِ گفتگوهای درحالِ‌اجرا هم با نسخه‌ی بازیابی‌شده هماهنگ شود
        icon = "✅" if ok else "❌"
        await callback.message.answer(
            f"{icon} {restore_msg}\n🕐 {format_jalali_datetime(datetime.utcnow())}"
        )
        return

    if action == "broadcast":
        await state.set_state(BroadcastStates.waiting_for_text)
        audience_count = len(load_funnel_users())
        await callback.message.edit_text(
            f"📢 <b>ارسال پیام همگانی</b>\n\n"
            f"این پیام برای همه‌ی کسانی که ربات را استارت زده‌اند ارسال می‌شود ({to_persian_num(audience_count)} نفر).\n\n"
            "حالا می‌توانید یک پیام متنی، عکس، سند، ویدئو یا هر نوع محتوای دیگری را بفرستید.\n\n"
            "برای انصراف، دستور /cancel را بفرستید.",
            reply_markup=admin_back_keyboard(),
        )
        await callback.answer()
        return

    if action == "sendmsg":
        await state.set_state(AdminSendMsgStates.waiting_for_identifier)
        await callback.message.edit_text(
            "📨 <b>ارسال پیام مستقیم</b>\n\n"
            "شناسهٔ کاربر را وارد کنید (آیدی عددی یا @username):\n"
            "مثال: 123456789  یا  @Ali_Arch\n\n"
            "⚠️ پس از شناسایی کاربر، می‌توانید هر نوع فایلی (عکس، سند، ویدئو، استیکر و...) ارسال کنید.\n"
            "(برای لغو، /cancel بفرستید)",
            reply_markup=admin_back_keyboard()
        )
        await callback.answer()
        return

    if action == "toggle_bot":
        state_data = load_bot_state()
        new_enabled = not state_data.get("enabled", True)
        state_data["enabled"] = new_enabled
        await save_bot_state(state_data)

        status_text = "روشن ✅" if new_enabled else "خاموش 🔴"
        await callback.answer(f"ربات {status_text} شد.")

        await callback.message.edit_text(
            await build_admin_dashboard_text(),
            reply_markup=admin_panel_keyboard()
        )

        if new_enabled:
            asyncio.create_task(process_pending_requests())
        return

    if action == "menu_edit":
        await state.clear()
        await callback.message.edit_text(
            "🛠 <b>مدیریت محتوا</b>\n\n"
            "از گزینه‌های زیر برای ویرایش محتوای پویای ربات استفاده کنید:",
            reply_markup=admin_menu_edit_keyboard()
        )
        await callback.answer()
        return

    if action == "edit_announcements":
        await state.set_state(ContentEditStates.editing_announcements)
        config = load_menu_config()
        announcements = config["settings"].get("announcements", [])
        files = config["settings"].get("announcement_files", [])

        text = "📢 <b>مدیریت اطلاعیه‌ها</b>\n\n"
        if announcements:
            text += "لیست اطلاعیه‌های فعلی:\n"
            for i, ann in enumerate(announcements, 1):
                text += f"{to_persian_num(i)}. {ann}\n"
        else:
            text += "هیچ اطلاعیه‌ای وجود ندارد.\n"

        if files:
            text += f"\n📎 {to_persian_num(len(files))} فایل ضمیمه شده است."

        text += "\n\n📝 برای <b>جایگزین کردن</b> کل اطلاعیه‌ها، یک متن جدید (هر خط یک اطلاعیه) ارسال کنید.\n"
        text += "📎 می‌توانید همراه با متن، فایل یا عکس نیز ارسال کنید (ضمیمه می‌شود).\n"
        text += "برای لغو، /cancel بفرستید."

        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="🗑 حذف همه", callback_data="admin:announcement_delete_all")],
                [InlineKeyboardButton(text="🗑 حذف یک مورد (شماره را وارد کنید)", callback_data="admin:announcement_delete_one_start")],
                [InlineKeyboardButton(text="🔙 بازگشت به مدیریت محتوا", callback_data="admin:menu_edit")],
            ]
        )
        await callback.message.edit_text(text, reply_markup=keyboard)
        await callback.answer()
        return

    if action == "announcement_delete_all":
        config = load_menu_config()
        config["settings"]["announcements"] = []
        config["settings"]["announcement_files"] = []
        await save_menu_config(config)
        await callback.answer("✅ همه اطلاعیه‌ها حذف شدند.")
        await handle_all_admin_callbacks(
            CallbackQuery(
                id=callback.id,
                from_user=callback.from_user,
                chat_instance=callback.chat_instance,
                message=callback.message,
                data="admin:edit_announcements"
            ),
            state
        )
        return

    if action == "announcement_delete_one_start":
        await state.set_state(ContentEditStates.deleting_announcement)
        await callback.message.edit_text(
            "🗑 <b>حذف یک اطلاعیه</b>\n\n"
            "شمارهٔ اطلاعیه‌ای که می‌خواهید حذف کنید را وارد کنید.\n"
            "مثال: 3\n\n"
            "برای لغو، /cancel بفرستید.",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="🔙 لغو و بازگشت", callback_data="admin:edit_announcements")]]
            )
        )
        await callback.answer()
        return

    if action == "edit_faq":
        await state.set_state(ContentEditStates.editing_faq)
        config = load_menu_config()
        faq_items = config["settings"].get("faq", [])
        files = config["settings"].get("faq_files", [])

        text = "❓ <b>مدیریت سوالات متداول</b>\n\n"
        if faq_items:
            text += "لیست سوالات فعلی:\n"
            for i, item in enumerate(faq_items, 1):
                text += f"{to_persian_num(i)}. س: {item['q']}\n   ج: {item['a']}\n"
        else:
            text += "هیچ سوالی ثبت نشده است.\n"

        if files:
            text += f"\n📎 {to_persian_num(len(files))} فایل ضمیمه شده است."

        text += "\n\n📝 برای <b>جایگزین کردن</b> کل سوالات، هر سوال و پاسخ را در یک خط به‌صورت زیر وارد کنید:\n"
        text += "سوال: پاسخ\n"
        text += "مثال: چطور عضو شوم؟: روی /start کلیک کنید.\n"
        text += "📎 می‌توانید همراه با متن، فایل یا عکس نیز ارسال کنید (ضمیمه می‌شود).\n"
        text += "برای لغو، /cancel بفرستید."

        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="🗑 حذف همه", callback_data="admin:faq_delete_all")],
                [InlineKeyboardButton(text="🗑 حذف یک مورد (شماره را وارد کنید)", callback_data="admin:faq_delete_one_start")],
                [InlineKeyboardButton(text="🔙 بازگشت به مدیریت محتوا", callback_data="admin:menu_edit")],
            ]
        )
        await callback.message.edit_text(text, reply_markup=keyboard)
        await callback.answer()
        return

    if action == "faq_delete_all":
        config = load_menu_config()
        config["settings"]["faq"] = []
        config["settings"]["faq_files"] = []
        await save_menu_config(config)
        await callback.answer("✅ همه سوالات حذف شدند.")
        await handle_all_admin_callbacks(
            CallbackQuery(
                id=callback.id,
                from_user=callback.from_user,
                chat_instance=callback.chat_instance,
                message=callback.message,
                data="admin:edit_faq"
            ),
            state
        )
        return

    if action == "faq_delete_one_start":
        await state.set_state(ContentEditStates.deleting_faq)
        await callback.message.edit_text(
            "🗑 <b>حذف یک سوال</b>\n\n"
            "شمارهٔ سوالی که می‌خواهید حذف کنید را وارد کنید.\n"
            "مثال: 2\n\n"
            "برای لغو، /cancel بفرستید.",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="🔙 لغو و بازگشت", callback_data="admin:edit_faq")]]
            )
        )
        await callback.answer()
        return

    if action == "delete_user":
        await state.set_state(DeleteUserStates.waiting_for_user_id)
        await callback.message.edit_text(
            "🗑 <b>حذف کاربر از گروه</b>\n\n"
            "آیدی عددی یا @username کاربر را وارد کنید:\n"
            "مثال: 123456789  یا  @Ali_Arch\n\n"
            "⚠️ کاربر از گروه اخراج شده و تمام اطلاعاتش (فرم، شماره تلفن) حذف می‌شود.\n"
            "(برای لغو، /cancel بفرستید)",
            reply_markup=admin_back_keyboard()
        )
        await callback.answer()
        return

    if action == "vip_settings":
        await state.clear()
        await callback.message.edit_text(
            await build_vip_settings_text(),
            reply_markup=vip_settings_keyboard(),
        )
        await callback.answer()
        return

    if action == "vip_global_settings":
        await state.clear()
        await callback.message.edit_text(
            await build_vip_global_settings_text(),
            reply_markup=vip_global_settings_keyboard(),
        )
        await callback.answer()
        return

    await callback.answer("❌ گزینه نامعتبر", show_alert=True)

# ---------- هندلرهای اختصاصی برای FSM ----------
class BroadcastStates(StatesGroup):
    waiting_for_text = State()
    confirming = State()

@dp.message(BroadcastStates.waiting_for_text)
async def handle_broadcast_text_input(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(
            "ارسال همگانی لغو شد.",
            reply_markup=admin_panel_keyboard(),
        )
        return

    await state.update_data(
        broadcast_chat_id=message.chat.id,
        broadcast_message_id=message.message_id,
    )
    await state.set_state(BroadcastStates.confirming)

    preview_text = "پیش‌نمایش پیام:\n"
    if message.text:
        preview_text += message.text
    elif message.caption:
        preview_text += f"📎 {message.caption}"
    else:
        preview_text += "📎 (یک فایل یا رسانه)"

    user_ids = load_funnel_users()
    confirm_keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ ارسال شود", callback_data="admin:broadcast_confirm", style="success")],
            [InlineKeyboardButton(text="❌ انصراف", callback_data="admin:broadcast_cancel", style="danger")],
        ]
    )
    await message.answer(
        f"{preview_text}\n\n"
        f"این پیام برای <b>{to_persian_num(len(user_ids))}</b> نفر (همه‌ی کسانی که ربات را استارت زده‌اند) ارسال می‌شود. مطمئنید؟",
        reply_markup=confirm_keyboard,
    )

@dp.callback_query(F.data == "admin:broadcast_confirm", BroadcastStates.confirming)
async def cb_broadcast_confirm(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    data = await state.get_data()
    chat_id = data.get("broadcast_chat_id")
    message_id = data.get("broadcast_message_id")
    await state.clear()

    if not chat_id or not message_id:
        await callback.answer()
        await callback.message.edit_text(
            "متنی برای ارسال پیدا نشد.",
            reply_markup=admin_back_keyboard()
        )
        return

    await callback.answer("⏳ در حال ارسال...")
    user_ids = load_funnel_users()
    sent, failed = 0, 0

    for uid in user_ids:
        try:
            await bot.copy_message(
                chat_id=uid,
                from_chat_id=chat_id,
                message_id=message_id,
            )
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.05)

    await callback.message.edit_text(
        f"✅ ارسال همگانی تمام شد.\n"
        f"موفق: <b>{to_persian_num(sent)}</b>\n"
        f"ناموفق: <b>{to_persian_num(failed)}</b>",
        reply_markup=admin_back_keyboard(),
    )

@dp.callback_query(F.data == "admin:broadcast_cancel", BroadcastStates.confirming)
async def cb_broadcast_cancel(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await state.clear()
    await callback.answer()
    await callback.message.edit_text("ارسال همگانی لغو شد.", reply_markup=admin_back_keyboard())

# ---------- صندوق پیام اعضا ----------
async def relay_message_to_admin(user, text: str) -> None:
    if not NOTIFY_CHAT_ID:
        return

    display_name = html_escape(user.full_name or user.first_name or "یک عضو")
    username_part = f"@{user.username}" if user.username else f"<code>{user.id}</code>"

    try:
        sent = await bot.send_message(
            chat_id=NOTIFY_CHAT_ID,
            text=(
                "📩 پیامِ تازه از یکی از اعضای رواق\n"
                f"👤 {display_name} ({username_part})\n\n"
                f"{html_escape(text)}\n\n"
                "برای پاسخ به همین عضو، فقط روی همین پیام «ریپلای» بزنید؛ "
                "پاسخ‌تون مستقیم و بدونِ نیاز به دونستنِ آیدی، براش ارسال می‌شه."
            ),
        )
        _pending_admin_replies[sent.message_id] = user.id
    except Exception as e:
        logger.warning("ارسالِ پیامِ عضو به ادمین ممکن نشد: %s", e)

@dp.message(F.chat.id == NOTIFY_CHAT_ID_INT, F.reply_to_message)
async def handle_admin_reply_via_native_reply(message: Message):
    if not is_admin(message.from_user.id):
        return

    replied_id = message.reply_to_message.message_id
    target_user_id = _pending_admin_replies.get(replied_id)
    if target_user_id is None:
        return

    reply_text = (message.html_text or message.text or "").strip()
    if not reply_text:
        return

    try:
        await bot.send_message(
            chat_id=target_user_id,
            text=f"از سوی مدیریتِ رواق:\n\n{reply_text}",
        )
        await message.reply("✅ پاسخ شما برای همون عضو ارسال شد.")
    except Exception as e:
        logger.warning("ارسالِ پاسخِ ادمین به کاربر %s ممکن نشد: %s", target_user_id, e)
        await message.reply(f"❌ ارسال پاسخ ناموفق بود: {e}")

@dp.message(F.chat.type == "private", StateFilter(None))
async def handle_generic_member_message(message: Message):
    if is_admin(message.from_user.id):
        return

    text = message.text or message.caption
    if not text or text.startswith("/"):
        return

    await relay_message_to_admin(message.from_user, text)
    await message.answer(
        "پیام شما به ادمین‌های رواق ارسال شد.\n"
        "به‌زودی پاسخ دریافت خواهید کرد 🙏"
    )

# ==============================================================
#  بخش پنل کاربری
# ==============================================================

class ContactAdminStates(StatesGroup):
    waiting_for_message = State()

async def show_text_panel(callback: CallbackQuery, text: str, keyboard: InlineKeyboardMarkup) -> None:
    """
    نمایشِ یک پیامِ متنیِ پنل، چه پیامِ فعلی متنی باشد و چه عکس (مثلاً از داخلِ اسلایدهای VIP).
    این تابع مانعِ خطای «ادیتِ پیامِ عکس‌دار به‌صورتِ متنی» می‌شود.
    """
    msg = callback.message
    if getattr(msg, "photo", None):
        try:
            await msg.delete()
        except Exception:
            pass
        await bot.send_message(chat_id=msg.chat.id, text=text, reply_markup=keyboard)
        return
    try:
        await msg.edit_text(text, reply_markup=keyboard)
    except Exception:
        try:
            await msg.delete()
        except Exception:
            pass
        await bot.send_message(chat_id=msg.chat.id, text=text, reply_markup=keyboard)

async def show_vip_page(callback: CallbackQuery, caption: str, keyboard: InlineKeyboardMarkup, image_id: str | None) -> None:
    """
    نمایشِ یک اسلایدِ VIP، مستقل از این‌که پیامِ فعلی متنی باشد یا عکس‌دار.
    """
    msg = callback.message
    has_photo = bool(getattr(msg, "photo", None))
    try:
        if image_id and has_photo:
            await msg.edit_media(
                media=InputMediaPhoto(media=image_id, caption=caption, show_caption_above_media=True),
                reply_markup=keyboard,
            )
            return
        if not image_id and not has_photo:
            await msg.edit_text(caption, reply_markup=keyboard)
            return
    except Exception:
        pass
    # انتقال بینِ نوعِ پیام (متن↔عکس) — پیامِ قبلی حذف و پیامِ تازه ارسال می‌شود
    try:
        await msg.delete()
    except Exception:
        pass
    if image_id:
        await bot.send_photo(
            chat_id=msg.chat.id, photo=image_id, caption=caption,
            show_caption_above_media=True, reply_markup=keyboard,
        )
    else:
        await bot.send_message(chat_id=msg.chat.id, text=caption, reply_markup=keyboard)

@dp.callback_query(F.data == "menu:back")
async def cb_menu_back(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await show_text_panel(
        callback,
        "🏛 <b>به رواق خوش آمدید</b>\n\n"
        "از پنل زیر یکی از گزینه‌ها را انتخاب کنید:",
        user_panel_keyboard(),
    )
    await callback.answer()

@dp.callback_query(F.data == "menu:close")
async def cb_menu_close(callback: CallbackQuery):
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.answer("پنل بسته شد.")

@dp.callback_query(F.data.startswith("menu:"))
async def handle_user_menu(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 1)[1]
    if key in ["close", "back"]:
        return

    if key == "profile":
        user_id = callback.from_user.id
        record = _user_cache.get(str(user_id))

        # ---------- بخشِ وضعیتِ اشتراکِ VIP ----------
        vip_line = ""
        vip_buttons = []
        if VIP_GROUP_CHAT_ID is not None:
            vip_status = get_user_vip_status(user_id)
            if vip_status["is_active"]:
                end_jalali = format_jalali_datetime(vip_status["end"])
                vip_line = (
                    "\n🌟 <b>اشتراکِ VIP:</b> فعال ✅\n"
                    f"⏳ {to_persian_num(vip_status['remaining_days'])} روزِ دیگر باقی مانده (تا {end_jalali})\n"
                )
                vip_buttons.append([InlineKeyboardButton(text="🌟 تمدیدِ اشتراکِ VIP", callback_data="vip:open", style="success")])
            elif vip_status["has_subscription"]:
                vip_line = "\n🌟 <b>اشتراکِ VIP:</b> به پایان رسیده ⌛️\n"
                vip_buttons.append([InlineKeyboardButton(text="🌟 تمدیدِ اشتراکِ VIP", callback_data="vip:open", style="success")])
            else:
                vip_line = "\n🌟 <b>اشتراکِ VIP:</b> ندارید\n"
                vip_buttons.append([InlineKeyboardButton(text="🌟 مشاهده‌ی گروهِ VIP", callback_data="vip:open", style="success")])

        if record:
            interests_text = '، '.join(record.get('interests', []))
            text = (
                "🥇 <b>پروفایلِ من — کاربرِ طلایی</b>\n\n"
                f"🎓 {record.get('education_label', '')}\n"
                f"⭐️ علایق: {interests_text}\n"
                f"{vip_line}\n"
                "می‌توانید هر زمان اطلاعاتِ خود را ویرایش کنید."
            )
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ ویرایشِ اطلاعات", callback_data="profile:edit", style="primary")],
                *vip_buttons,
                [InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back", style="primary")],
            ])
        else:
            text = (
                "👤 <b>پروفایلِ من</b>\n"
                f"{vip_line}\n"
                "هنوز پروفایلِ شما تکمیل نشده است.\n"
                "با پاسخ به سه سوالِ کوتاه، به «🥇 کاربرِ طلایی» رواق ارتقا پیدا کنید."
            )
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🚀 شروعِ تکمیلِ پروفایل", callback_data="profile:start", style="success")],
                *vip_buttons,
                [InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back", style="primary")],
            ])
        await show_text_panel(callback, text, keyboard)
        await callback.answer()
        return

    config = load_menu_config()
    item = config["menu_items"].get(key)
    if not item:
        await callback.answer("گزینه‌ای یافت نشد.")
        return

    if key == "vip":
        if VIP_GROUP_CHAT_ID is None:
            await callback.answer("گروه VIP هنوز راه‌اندازی نشده است.", show_alert=True)
            return
        caption, keyboard, image_id = await render_vip_page(0)
        await show_vip_page(callback, caption, keyboard, image_id)
        await callback.answer()
        return

    if key == "topics":
        await callback.message.edit_text(
            "📚 <b>راهنمای تاپیک‌های رواق</b>\n\n"
            "لطفاً یکی از تاپیک‌های زیر را انتخاب کنید:",
            reply_markup=topics_panel_keyboard()
        )
        await callback.answer()
        return

    if key == "contact_admin":
        await state.set_state(ContactAdminStates.waiting_for_message)
        await callback.message.edit_text(
            "📞 پیام خود را تایپ کنید تا برای ادمین ارسال شود.\n"
            "(برای لغو، /cancel بفرستید)",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="🔙 بازگشت به پنل", callback_data="menu:back")]]
            )
        )
        await callback.answer()
        return

    if key == "my_status":
        user_id = callback.from_user.id
        await send_with_action(callback.message.chat.id, "typing", 1.0)
        try:
            user = await bot.get_chat(user_id)
            display_name = html_escape(user.full_name or user.first_name or "کاربر")
        except Exception:
            display_name = "کاربر"

        is_member = False
        try:
            member = await bot.get_chat_member(GROUP_CHAT_ID, user_id)
            is_member = member.status in [ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR]
        except Exception as e:
            logger.warning("گرفتن وضعیت عضویت کاربر %s ممکن نشد: %s", user_id, e)
            is_member = False

        if is_member:
            status_text = f"✅ {display_name} عزیز، شما عضو گروه هستید."
        else:
            status_text = f"❌ {display_name} عزیز، شما عضو گروه نیستید."

        try:
            await callback.message.delete()
        except Exception:
            pass

        await bot.send_message(
            chat_id=callback.message.chat.id,
            text=status_text,
            reply_markup=user_panel_keyboard(),
            message_effect_id=MESSAGE_EFFECT_PARTY_POPPER if is_member else None,
        )
        await callback.answer()
        return

    if key == "announcements":
        announcements = config["settings"].get("announcements", [])
        announcement_files = config["settings"].get("announcement_files", [])

        if announcements:
            announcements_text = "\n".join([f"▪️ {a}" for a in announcements])
        else:
            announcements_text = "📢 هیچ اطلاعیه‌ای وجود ندارد."

        response = item["response"].format(
            announcements=announcements_text
        )

        await callback.message.edit_text(response, reply_markup=user_panel_keyboard())

        for file_id in announcement_files:
            try:
                await callback.message.answer_document(document=file_id)
            except Exception:
                pass

        await callback.answer()
        return

    if key == "faq":
        faq_items = config["settings"].get("faq", [])
        faq_files = config["settings"].get("faq_files", [])

        if faq_items:
            faq_text = "\n".join([f"❓ {item['q']}\n📝 {item['a']}" for item in faq_items])
        else:
            faq_text = "هنوز سوالی ثبت نشده است."

        response = item["response"].format(
            faq_list=faq_text
        )

        await callback.message.edit_text(response, reply_markup=user_panel_keyboard())

        for file_id in faq_files:
            try:
                await callback.message.answer_document(document=file_id)
            except Exception:
                pass

        await callback.answer()
        return

    if key == "social":
        social = config["settings"]["social"]
        response = item["response"].format(
            instagram=social.get("instagram", ""),
            channel=social.get("channel", "")
        )
        await callback.message.edit_text(response, reply_markup=user_panel_keyboard())
        await callback.answer()
        return

    if key == "join":
        response = item["response"].format(
            invite_link=config["settings"]["group_invite_link"]
        )
        await callback.message.edit_text(response, reply_markup=user_panel_keyboard())
        await callback.answer()
        return

    await callback.message.edit_text(response, reply_markup=user_panel_keyboard())
    await callback.answer()

@dp.message(ContactAdminStates.waiting_for_message)
async def handle_contact_admin_message(message: Message, state: FSMContext):
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=user_panel_keyboard())
        return

    await relay_message_to_admin(message.from_user, message.text)
    await message.answer("✅ پیام شما به ادمین ارسال شد.", reply_markup=user_panel_keyboard())
    await state.clear()

# ==============================================================
#  بخش مدیریت محتوا (ادمین)
# ==============================================================

class ContentEditStates(StatesGroup):
    editing_announcements = State()
    editing_faq = State()
    deleting_announcement = State()
    deleting_faq = State()

@dp.message(ContentEditStates.deleting_announcement)
async def handle_delete_announcement_number(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    if not message.text or not message.text.strip().isdigit():
        await message.answer("❌ لطفاً یک شمارهٔ معتبر (عدد) وارد کنید.")
        return

    index = int(message.text.strip())
    config = load_menu_config()
    announcements = config["settings"].get("announcements", [])
    if 1 <= index <= len(announcements):
        deleted = announcements.pop(index - 1)
        config["settings"]["announcements"] = announcements
        await save_menu_config(config)
        await message.answer(
            f"✅ اطلاعیهٔ شمارهٔ {to_persian_num(index)} با متن:\n"
            f"«{deleted}»\n"
            "حذف شد."
        )
    else:
        await message.answer(f"❌ شمارهٔ {to_persian_num(index)} معتبر نیست. تعداد اطلاعیه‌ها: {to_persian_num(len(announcements))}")

    await state.clear()
    await message.answer("برای ادامه، روی دکمه‌ی «ویرایش اطلاعیه‌ها» کلیک کنید.", reply_markup=admin_menu_edit_keyboard())

@dp.message(ContentEditStates.deleting_faq)
async def handle_delete_faq_number(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    if not message.text or not message.text.strip().isdigit():
        await message.answer("❌ لطفاً یک شمارهٔ معتبر (عدد) وارد کنید.")
        return

    index = int(message.text.strip())
    config = load_menu_config()
    faq_items = config["settings"].get("faq", [])
    if 1 <= index <= len(faq_items):
        deleted = faq_items.pop(index - 1)
        config["settings"]["faq"] = faq_items
        await save_menu_config(config)
        await message.answer(
            f"✅ سوال شمارهٔ {to_persian_num(index)} با متن:\n"
            f"«{deleted['q']}»\n"
            "حذف شد."
        )
    else:
        await message.answer(f"❌ شمارهٔ {to_persian_num(index)} معتبر نیست. تعداد سوالات: {to_persian_num(len(faq_items))}")

    await state.clear()
    await message.answer("برای ادامه، روی دکمه‌ی «ویرایش سوالات متداول» کلیک کنید.", reply_markup=admin_menu_edit_keyboard())

@dp.message(ContentEditStates.editing_announcements)
async def handle_edit_announcements(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    config = load_menu_config()

    if message.text:
        lines = [line.strip() for line in message.text.split("\n") if line.strip()]
        if lines:
            config["settings"]["announcements"] = lines

    file_id = None
    if message.document:
        file_id = message.document.file_id
    elif message.photo:
        file_id = message.photo[-1].file_id
    elif message.video:
        file_id = message.video.file_id
    elif message.audio:
        file_id = message.audio.file_id

    if file_id:
        if "announcement_files" not in config["settings"]:
            config["settings"]["announcement_files"] = []
        config["settings"]["announcement_files"].append(file_id)
        await message.answer("✅ فایل با موفقیت ضمیمه شد.")

    await save_menu_config(config)

    if message.text:
        await message.answer(
            f"✅ {to_persian_num(len(config['settings']['announcements']))} اطلاعیه ثبت شد.",
            reply_markup=admin_menu_edit_keyboard()
        )
    else:
        await message.answer("✅ فایل ضمیمه شد.", reply_markup=admin_menu_edit_keyboard())

    await state.clear()

@dp.message(ContentEditStates.editing_faq)
async def handle_edit_faq(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    config = load_menu_config()

    if message.text:
        lines = [line.strip() for line in message.text.split("\n") if line.strip()]
        faq_items = []
        for line in lines:
            if ":" in line:
                q, a = line.split(":", 1)
                faq_items.append({"q": q.strip(), "a": a.strip()})
        if faq_items:
            config["settings"]["faq"] = faq_items

    file_id = None
    if message.document:
        file_id = message.document.file_id
    elif message.photo:
        file_id = message.photo[-1].file_id
    elif message.video:
        file_id = message.video.file_id
    elif message.audio:
        file_id = message.audio.file_id

    if file_id:
        if "faq_files" not in config["settings"]:
            config["settings"]["faq_files"] = []
        config["settings"]["faq_files"].append(file_id)
        await message.answer("✅ فایل با موفقیت ضمیمه شد.")

    await save_menu_config(config)

    if message.text:
        await message.answer(
            f"✅ {to_persian_num(len(config['settings']['faq']))} سوال و پاسخ ثبت شد.",
            reply_markup=admin_menu_edit_keyboard()
        )
    else:
        await message.answer("✅ فایل ضمیمه شد.", reply_markup=admin_menu_edit_keyboard())

    await state.clear()

# ==============================================================
#  بخش حذف کاربر (ادمین)
# ==============================================================

class DeleteUserStates(StatesGroup):
    waiting_for_user_id = State()
    confirming = State()

@dp.message(DeleteUserStates.waiting_for_user_id)
async def delete_user_identifier(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    identifier = message.text.strip()
    if not identifier:
        await message.answer("لطفاً یک شناسه معتبر وارد کنید.")
        return

    if identifier.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    user_id = None
    if identifier.isdigit():
        user_id = int(identifier)
    else:
        username = identifier.lstrip('@')
        try:
            chat = await bot.get_chat(f"@{username}")
            user_id = chat.id
        except Exception as e:
            await message.answer(
                f"❌ کاربر @{username} پیدا نشد. خطا: {e}\n"
                "توجه: کاربر باید حداقل یک‌بار ربات را استارت کرده باشد یا در گروه عضو باشد.\n"
                "لطفاً دوباره آیدی عددی یا یوزرنیم صحیح را وارد کنید."
            )
            return

    try:
        user = await bot.get_chat(user_id)
        display = user.full_name or str(user_id)
        await state.update_data(target_user_id=user_id, target_display=display)
        await state.set_state(DeleteUserStates.confirming)

        confirm_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✅ بله، حذف شود", callback_data="admin:delete_confirm", style="danger")],
                [InlineKeyboardButton(text="❌ انصراف", callback_data="admin:delete_cancel", style="primary")],
            ]
        )
        await message.answer(
            f"⚠️ آیا از حذف کاربر <b>{html_escape(display)}</b> (آیدی: <code>{user_id}</code>) مطمئنید؟\n"
            "این عملیات غیرقابل بازگشت است.",
            reply_markup=confirm_keyboard
        )
    except Exception as e:
        await message.answer(f"❌ خطا در دریافت اطلاعات کاربر: {e}\nلطفاً دوباره تلاش کنید.")
        await state.clear()

@dp.callback_query(F.data == "admin:delete_confirm", DeleteUserStates.confirming)
async def cb_delete_confirm(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    data = await state.get_data()
    user_id = data.get("target_user_id")
    display = data.get("target_display", "کاربر")
    if not user_id:
        await callback.message.edit_text("خطا: کاربر مشخص نیست.")
        await state.clear()
        return

    # ۱) تلاش برای اخراج از گروه (در صورت عضویت)
    try:
        member = await bot.get_chat_member(GROUP_CHAT_ID, user_id)
        if member.status in [ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR]:
            await bot.ban_chat_member(chat_id=GROUP_CHAT_ID, user_id=user_id)
    except Exception as e:
        logger.warning(f"اخراج کاربر {user_id} از گروه ممکن نشد (احتمالاً عضو نیست): {e}")

    # ۲) حذف اطلاعات از فایل‌ها با مدیریت خطا
    errors = []
    try:
        async with _write_lock:
            verified = load_verified()
            if str(user_id) in verified:
                del verified[str(user_id)]
                VERIFIED_FILE.write_text(json.dumps(verified, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        errors.append(f"خطا در حذف از verified: {e}")
        logger.error(f"خطا در حذف کاربر {user_id} از verified: {e}")

    try:
        if DATA_FILE.exists():
            new_lines = []
            async with _write_lock:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            record = json.loads(line)
                            if record.get("user_id") != user_id:
                                new_lines.append(line)
                        except json.JSONDecodeError:
                            continue
                with open(DATA_FILE, "w", encoding="utf-8") as f:
                    f.write("\n".join(new_lines))
                    if new_lines:
                        f.write("\n")
    except Exception as e:
        errors.append(f"خطا در حذف از submissions: {e}")
        logger.error(f"خطا در حذف کاربر {user_id} از submissions: {e}")

    try:
        if str(user_id) in _user_cache:
            del _user_cache[str(user_id)]
    except Exception as e:
        errors.append(f"خطا در حذف از کش: {e}")
        logger.error(f"خطا در حذف کاربر {user_id} از کش: {e}")

    if errors:
        await callback.message.edit_text(
            f"⚠️ کاربر <b>{html_escape(display)}</b> تا حدی حذف شد، اما خطاهایی رخ داد:\n" + "\n".join(errors) +
            "\n\nلطفاً وضعیت را دستی بررسی کنید."
        )
    else:
        await callback.message.edit_text(f"✅ کاربر <b>{html_escape(display)}</b> با موفقیت حذف شد.")

    await state.clear()
    await callback.message.answer("🛠 پنل مدیریت", reply_markup=admin_panel_keyboard())

# ==============================================================
#  بخش ارسال مستقیم به کاربر
# ==============================================================

class AdminSendMsgStates(StatesGroup):
    waiting_for_identifier = State()
    waiting_for_message = State()

@dp.message(AdminSendMsgStates.waiting_for_identifier)
async def admin_sendmsg_identifier(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    identifier = message.text.strip()
    if not identifier:
        await message.answer("لطفاً یک شناسه معتبر وارد کنید.")
        return

    if identifier.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    user_id = None
    if identifier.isdigit():
        user_id = int(identifier)
    else:
        username = identifier.lstrip('@')
        try:
            chat = await bot.get_chat(f"@{username}")
            user_id = chat.id
        except Exception as e:
            await message.answer(f"❌ کاربر @{username} پیدا نشد. خطا: {e}\nلطفاً دوباره وارد کنید.")
            return

    try:
        user = await bot.get_chat(user_id)
        display = user.full_name or str(user_id)
        await state.update_data(target_user_id=user_id, target_user_display=display)
        await message.answer(
            f"👤 کاربر: <b>{html_escape(display)}</b> (آیدی: <code>{user_id}</code>)\n\n"
            "📤 حالا <b>متن، عکس، سند، ویدئو، استیکر یا هر فایل دیگری</b> را ارسال کنید:\n"
            "(برای لغو، /cancel بفرستید)",
            reply_markup=admin_back_keyboard()
        )
        await state.set_state(AdminSendMsgStates.waiting_for_message)
    except Exception as e:
        await message.answer(f"❌ خطا در دریافت اطلاعات کاربر: {e}")

@dp.message(AdminSendMsgStates.waiting_for_message)
async def admin_sendmsg_media(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_panel_keyboard())
        return

    data = await state.get_data()
    user_id = data.get("target_user_id")
    display = data.get("target_user_display", "کاربر")
    if not user_id:
        await message.answer("خطا: کاربر مشخص نیست.")
        await state.clear()
        return

    caption = message.caption or ""

    try:
        if message.text:
            await bot.send_message(chat_id=user_id, text=message.text)
        elif message.photo:
            await bot.send_photo(chat_id=user_id, photo=message.photo[-1].file_id, caption=caption)
        elif message.document:
            await bot.send_document(chat_id=user_id, document=message.document.file_id, caption=caption, file_name=message.document.file_name)
        elif message.video:
            await bot.send_video(chat_id=user_id, video=message.video.file_id, caption=caption, supports_streaming=True)
        elif message.audio:
            await bot.send_audio(chat_id=user_id, audio=message.audio.file_id, caption=caption, performer=message.audio.performer, title=message.audio.title)
        elif message.voice:
            await bot.send_voice(chat_id=user_id, voice=message.voice.file_id, caption=caption)
        elif message.video_note:
            await bot.send_video_note(chat_id=user_id, video_note=message.video_note.file_id)
        elif message.sticker:
            await bot.send_sticker(chat_id=user_id, sticker=message.sticker.file_id)
        elif message.animation:
            await bot.send_animation(chat_id=user_id, animation=message.animation.file_id, caption=caption)
        elif message.contact:
            await bot.send_contact(chat_id=user_id, phone_number=message.contact.phone_number, first_name=message.contact.first_name, last_name=message.contact.last_name)
        elif message.location:
            await bot.send_location(chat_id=user_id, latitude=message.location.latitude, longitude=message.location.longitude)
        elif message.poll:
            await bot.send_poll(chat_id=user_id, question=message.poll.question, options=[opt.text for opt in message.poll.options], is_anonymous=message.poll.is_anonymous, type=message.poll.type)
        else:
            await message.answer("❌ این نوع فایل پشتیبانی نمی‌شود.")
            return

        await message.answer(f"✅ پیام به <b>{html_escape(display)}</b> ارسال شد.")
    except Exception as e:
        logger.error(f"خطا در ارسال پیام به {user_id}: {e}")
        await message.answer(f"❌ خطا در ارسال پیام: {e}")
    await state.clear()

# ==============================================================
#  بخش حضور و غیاب
# ==============================================================

def load_attendance_data() -> dict:
    if not ATTENDANCE_FILE.exists():
        return {"active": None}
    try:
        return json.loads(ATTENDANCE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"active": None}

async def save_attendance_data(data: dict) -> None:
    async with _write_lock:
        ATTENDANCE_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

async def build_attendance_excel(participants: list) -> BufferedInputFile:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "لیست حاضرین"
    sheet.sheet_view.rightToLeft = True

    headers = ["ردیف", "نام کامل", "آیدی عددی", "یوزرنیم", "تاریخ ثبت (شمسی)"]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="14532F", end_color="14532F", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, p in enumerate(participants, 1):
        uid = p["user_id"]
        time_jalali = p.get("time", "نامشخص")

        try:
            member = await bot.get_chat_member(GROUP_CHAT_ID, uid)
            name = member.user.full_name or str(uid)
            username = f"@{member.user.username}" if member.user.username else ""
        except Exception:
            name = f"کاربر {uid}"
            username = ""

        sheet.append([i, name, uid, username, time_jalali])

    col_widths = [8, 30, 18, 20, 22]
    for idx, width in enumerate(col_widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return BufferedInputFile(buffer.read(), filename="لیست_حاضرین_نهایی.xlsx")

async def _send_attendance_reminder(chat_id: int, main_message_id: int, data: dict):
    active = data.get("active")
    if not active:
        return
    last_msg_id = active.get("last_reminder_message_id")
    if last_msg_id:
        try:
            await bot.delete_message(chat_id=chat_id, message_id=last_msg_id)
        except Exception as e:
            logger.warning("حذف پیام یادآوری قبلی ممکن نشد: %s", e)

    text = (
        "⚠️ <b>یادآوری حضور هفتگی</b>\n\n"
        "هنوز حضور خود را ثبت نکرده‌اید.\n"
        "لطفاً با کلیک روی دکمه‌ی «✅ من اینجام» در پیام بالایی، حضور خود را تأیید کنید.\n\n"
        "🔴 یادآوری: اگر نامتان در لیست نهایی نباشد، از گروه حذف خواهید شد."
    )
    try:
        sent = await bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_to_message_id=main_message_id,
        )
        active["last_reminder_message_id"] = sent.message_id
        await save_attendance_data(data)
    except Exception as e:
        logger.error("ارسال یادآوری ممکن نشد: %s", e)

async def _attendance_reminder_task(chat_id: int, main_message_id: int):
    data = load_attendance_data()
    active = data.get("active")
    if not active:
        return

    for i in range(1, 9):
        await asyncio.sleep(3 * 3600)

        data = load_attendance_data()
        active = data.get("active")
        if not active:
            break

        await _send_attendance_reminder(chat_id, main_message_id, data)
        active["reminder_count"] = i
        await save_attendance_data(data)

    data = load_attendance_data()
    active = data.get("active")
    if active:
        await _finish_attendance(chat_id)

async def _finish_attendance(chat_id: int):
    data = load_attendance_data()
    active = data.get("active")
    if not active:
        return

    admin_id = active.get("admin_id", chat_id)
    participants = active.get("participants", [])
    total = len(participants)

    last_reminder = active.get("last_reminder_message_id")
    if last_reminder:
        try:
            await bot.delete_message(chat_id=chat_id, message_id=last_reminder)
        except Exception:
            pass

    if total == 0:
        await bot.send_message(
            chat_id=admin_id,
            text="📋 در این دوره هیچ‌کس حضور ثبت نکرد."
        )
    else:
        await bot.send_message(
            chat_id=admin_id,
            text=f"📋 <b>دوره‌ی حضور و غیاب به پایان رسید.</b>\nتعداد کل شرکت‌کنندگان: <b>{to_persian_num(total)}</b> نفر"
        )

        if total > 30:
            excel_file = await build_attendance_excel(participants)
            await bot.send_document(
                chat_id=admin_id,
                document=excel_file,
                caption=f"📄 لیست کامل {to_persian_num(total)} نفر شرکت‌کننده"
            )
        else:
            lines = ["👤 <b>شرکت‌کنندگان:</b>"]
            for p in participants:
                uid = p["user_id"]
                time_jalali = p.get("time", "نامشخص")
                try:
                    member = await bot.get_chat_member(GROUP_CHAT_ID, uid)
                    name = member.user.full_name or str(uid)
                    username = f" @{member.user.username}" if member.user.username else ""
                    lines.append(f"• {name} (ID: <code>{uid}</code>){username} — ثبت در {time_jalali}")
                except Exception:
                    lines.append(f"• کاربر با آیدی <code>{uid}</code> — ثبت در {time_jalali}")

            text = "\n".join(lines)
            if len(text) < 4000:
                await bot.send_message(chat_id=admin_id, text=text)
            else:
                plain = [l.replace("<b>", "").replace("</b>", "").replace("<code>", "").replace("</code>", "") for l in lines]
                with BytesIO() as f:
                    f.write("\n".join(plain).encode("utf-8"))
                    f.seek(0)
                    await bot.send_document(
                        chat_id=admin_id,
                        document=BufferedInputFile(f.read(), filename="لیست_حاضرین.txt"),
                        caption=f"📄 لیست کامل حاضرین (تعداد {to_persian_num(total)} نفر)"
                    )

    data["active"] = None
    await save_attendance_data(data)

    task_key = f"reminder_{chat_id}"
    if task_key in _attendance_tasks:
        _attendance_tasks[task_key].cancel()
        del _attendance_tasks[task_key]

    await bot.send_message(
        chat_id=admin_id,
        text="✅ گزارش نهایی حضور و غیاب ارسال شد."
    )

@dp.message(Command("attendance_start"))
async def cmd_attendance_start(message: Message):
    if not is_admin(message.from_user.id):
        return
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("این دستور فقط در گروه قابل استفاده است.")
        return

    data = load_attendance_data()
    if data.get("active") is not None:
        await message.answer("یک دوره‌ی حضور و غیاب هم‌اکنون فعال است. ابتدا آن را با /attendance_end پایان دهید.")
        return

    text = (
        "📋 <b>ثبت حضور هفتگی</b>\n\n"
        "⚠️ این پیام یک <b>دورهٔ ۲۴ ساعته</b> برای شناسایی اعضای فعال است.\n"
        "اگر در گروه فعال هستید، حتماً روی دکمهٔ زیر کلیک کنید.\n\n"
        "🔴 <b>توجه:</b> اعضایی که نامشان در لیست نهایی نباشد، از گروه <b>حذف</b> خواهند شد.\n\n"
        "پس از پایان ۲۴ ساعت، لیست نهایی اعلام می‌شود."
    )
    try:
        sent_msg = await bot.send_message(
            chat_id=message.chat.id,
            text=text,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="✅ من اینجام", callback_data="attendance:yes")]
                ]
            )
        )
    except Exception as e:
        logger.error("خطا در ارسال پیام حضور: %s", e)
        await message.answer("خطا در ارسال پیام حضور.")
        return

    active_data = {
        "started_at": datetime.utcnow().isoformat(),
        "message_id": sent_msg.message_id,
        "chat_id": message.chat.id,
        "admin_id": message.from_user.id,
        "participants": [],
        "reminder_count": 0,
        "last_reminder_message_id": None,
    }
    data["active"] = active_data
    await save_attendance_data(data)

    task_key = f"reminder_{message.chat.id}"
    if task_key in _attendance_tasks:
        _attendance_tasks[task_key].cancel()
    task = asyncio.create_task(_attendance_reminder_task(message.chat.id, sent_msg.message_id))
    _attendance_tasks[task_key] = task

    await message.answer("✅ دوره‌ی حضور و غیاب شروع شد. اولین یادآوری ۳ ساعت دیگر ارسال می‌شود.")

@dp.message(Command("attendance_status"))
async def cmd_attendance_status(message: Message):
    if not is_admin(message.from_user.id):
        return
    data = load_attendance_data()
    active = data.get("active")
    if active is None:
        await message.answer("هیچ دوره‌ی فعالی وجود ندارد.")
        return

    started = format_jalali_datetime(datetime.fromisoformat(active["started_at"]))
    participants_count = len(active["participants"])
    reminder_count = active.get("reminder_count", 0)
    remaining = 8 - reminder_count
    if remaining < 0:
        remaining = 0
    await message.answer(
        f"📊 <b>وضعیت حضور و غیاب</b>\n\n"
        f"⏰ شروع: <code>{started}</code>\n"
        f"👤 شرکت‌کنندگان: <b>{to_persian_num(participants_count)}</b> نفر\n"
        f"🔔 تعداد یادآوری‌های ارسال‌شده: <b>{to_persian_num(reminder_count)}</b>\n"
        f"⏳ یادآوری‌های باقی‌مانده: <b>{to_persian_num(remaining)}</b>"
    )

@dp.message(Command("attendance_end"))
async def cmd_attendance_end(message: Message):
    if not is_admin(message.from_user.id):
        return
    data = load_attendance_data()
    if data.get("active") is None:
        await message.answer("هیچ دوره‌ی فعالی برای پایان دادن وجود ندارد.")
        return

    await _finish_attendance(message.chat.id)
    await message.answer("✅ دوره‌ی حضور و غیاب به‌صورت دستی پایان یافت و گزارش برای شما ارسال شد.")

@dp.message(Command("attendance_report"))
async def cmd_attendance_report(message: Message):
    if not is_admin(message.from_user.id):
        return
    data = load_attendance_data()
    active = data.get("active")
    if active is None:
        await message.answer("هیچ دوره‌ی فعالی وجود ندارد.")
        return

    participants = active.get("participants", [])
    if not participants:
        await message.answer("📋 تا الان کسی حضور خود را ثبت نکرده است.")
        return

    total = len(participants)
    if total > 30:
        excel_file = await build_attendance_excel(participants)
        await message.answer_document(
            document=excel_file,
            caption=f"📄 لیست حاضرین تا این لحظه (تعداد {to_persian_num(total)} نفر)"
        )
    else:
        lines = ["📋 لیست حاضرین تا این لحظه:"]
        for p in participants:
            uid = p["user_id"]
            time_jalali = p.get("time", "نامشخص")
            try:
                member = await bot.get_chat_member(GROUP_CHAT_ID, uid)
                name = member.user.full_name or str(uid)
                username = f" @{member.user.username}" if member.user.username else ""
                lines.append(f"• {name} (ID: {uid}){username} — ثبت در {time_jalali}")
            except Exception:
                lines.append(f"• کاربر با آیدی {uid} — ثبت در {time_jalali}")
        await message.answer("\n".join(lines))

@dp.callback_query(F.data == "attendance:yes")
async def cb_attendance_yes(callback: CallbackQuery):
    user_id = callback.from_user.id
    data = load_attendance_data()
    active = data.get("active")
    if active is None:
        await callback.answer("دوره‌ی حضور و غیاب فعال نیست.", show_alert=True)
        return

    for p in active["participants"]:
        if p["user_id"] == user_id:
            await callback.answer("✅ حضور شما قبلاً ثبت شده است.")
            return

    now_utc = datetime.utcnow()
    time_jalali = format_jalali_datetime(now_utc)
    active["participants"].append({"user_id": user_id, "time": time_jalali})
    await save_attendance_data(data)
    await callback.answer("✅ حضور شما ثبت شد.", show_alert=False)

async def restore_attendance_tasks():
    data = load_attendance_data()
    active = data.get("active")
    if not active:
        return
    chat_id = active["chat_id"]
    message_id = active["message_id"]
    started_at = datetime.fromisoformat(active["started_at"])
    now = datetime.utcnow()
    elapsed = (now - started_at).total_seconds()

    if elapsed >= 24 * 3600:
        await _finish_attendance(chat_id)
        return

    reminder_count = active.get("reminder_count", 0)
    if reminder_count < 8:
        next_reminder_seconds = (reminder_count + 1) * 3 * 3600 - elapsed
        if next_reminder_seconds < 0:
            next_reminder_seconds = 0
        async def delayed_start():
            await asyncio.sleep(next_reminder_seconds)
            await _attendance_reminder_task(chat_id, message_id)
        task = asyncio.create_task(delayed_start())
        _attendance_tasks[f"reminder_{chat_id}"] = task
        logger.info("دوره‌ی حضور و غیاب بازیابی شد. یادآوری بعدی در %s ثانیه.", next_reminder_seconds)
    else:
        remaining_to_end = 24 * 3600 - elapsed
        if remaining_to_end > 0:
            async def delayed_end():
                await asyncio.sleep(remaining_to_end)
                await _finish_attendance(chat_id)
            task = asyncio.create_task(delayed_end())
            _attendance_tasks[f"reminder_{chat_id}"] = task
            logger.info("تسک پایان دوره بازیابی شد. پایان در %s ثانیه.", remaining_to_end)

# ==============================================================
#  ماژول گروه VIP (نسخهٔ نهایی با اشتراک کامل)
# ==============================================================

def _cancel_vip_intro(user_id: int) -> None:
    task = _vip_intro_tasks.pop(user_id, None)
    if task and not task.done():
        task.cancel()

def _schedule_vip_intro(user_id: int) -> None:
    _cancel_vip_intro(user_id)

    async def _intro():
        await asyncio.sleep(VIP_INTRO_DELAY_MINUTES * 60)
        try:
            await send_vip_intro_message(user_id)
        except Exception as e:
            logger.warning("ارسالِ پیامِ معرفیِ VIP به کاربر %s ممکن نشد: %s", user_id, e)

    _vip_intro_tasks[user_id] = asyncio.create_task(_intro())

async def send_vip_intro_message(user_id: int) -> None:
    if VIP_GROUP_CHAT_ID is None:
        return
    text = sign(
        "🌟 <b>یک قدم فراتر از رواق — گروه VIP</b>\n\n"
        "برای کسانی که می‌خواهند مسیرِ حرفه‌ای‌شان را جدی‌تر دنبال کنند، رواقِ VIP را راه‌اندازی کرده‌ایم:\n\n"
        "▪️ هر نرم‌افزار، تاپیکِ اختصاصیِ خودش را دارد؛ بدونِ قاطی‌شدنِ موضوعات.\n"
        "▪️ آموزشِ ویدئوییِ کامل برای هر نرم‌افزار — از صفر تا حرفه‌ای.\n"
        "▪️ آرشیوِ کاملِ پلاگین‌ها، فمیلی‌ها، آبجکت‌ها و متریال‌های به‌روز.\n"
        "▪️ همه‌چیز در یک‌جا؛ دیگر نیازی به خریدِ دوره‌های پراکنده‌ی دیگر نیست.\n\n"
        "اشتراک‌ها به‌صورتِ ۳، ۶ و ۱۲ ماهه ارائه می‌شوند.\n\n"
        "برای دیدنِ دسته‌بندی‌ها و تعرفه‌ها، دکمه‌ی زیر را بزنید 👇"
    )
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🌟 مشاهده‌ی گروه VIP", callback_data="vip:open", style="success")],
        ]
    )
    await bot.send_message(chat_id=user_id, text=text, reply_markup=keyboard)

# ---------- پنل مرور دسته‌بندی‌های VIP (کاربر) ----------
async def render_vip_page(index: int):
    categories = load_vip_categories()
    if not categories:
        return (
            "🌟 <b>گروه VIP</b>\n\nهنوز هیچ دسته‌بندی‌ای اضافه نشده. به‌زودی تکمیل می‌شود.",
            InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🔙 بازگشت به پنل اصلی", callback_data="menu:back", style="danger")]]),
            None
        )
    total = len(categories)
    index = max(0, min(index, total - 1))
    cat = categories[index]

    desc_lines = cat.get('description', '').split('\n')
    desc_text = '\n'.join([html_escape(line) for line in desc_lines])

    caption = (
        f"🌟 <b>گروه VIP</b>\n"
        f"({to_persian_num(index + 1)}/{to_persian_num(total)})\n\n"
        f"📦 <b>{html_escape(cat['name'])}</b>\n\n"
        f"{desc_text}\n\n"
        "💎 با خریدِ اشتراک، به تمامِ محتوای گروهِ VIP یک‌جا دسترسی پیدا می‌کنید.\n"
        "⠀"
    )

    rows = []
    nav_row = []
    # اسلایدِ اول فقط «بعدی» و اسلایدِ آخر فقط «قبلی» دارد؛ در میانه هر دو دکمه نمایش داده می‌شود.
    if total > 1:
        if index > 0:
            nav_row.append(InlineKeyboardButton(text="◀️ قبلی", callback_data=f"vipnav:{index - 1}", style="primary"))
        if index < total - 1:
            nav_row.append(InlineKeyboardButton(text="بعدی ▶️", callback_data=f"vipnav:{index + 1}", style="primary"))
    if nav_row:
        rows.append(nav_row)

    if total > 1:
        rows.append([InlineKeyboardButton(text="📋 فهرست کامل دسته‌بندی‌ها", callback_data="vip:list", style="primary")])

    rows.append([InlineKeyboardButton(text="💎 خرید اشتراک کامل", callback_data="vip:buy_subscription", style="success")])
    rows.append([InlineKeyboardButton(text="🔙 بازگشت به پنل اصلی", callback_data="menu:back", style="danger")])
    keyboard = InlineKeyboardMarkup(inline_keyboard=rows)

    image_file_id = cat.get("image_file_id")
    return caption, keyboard, image_file_id

async def render_vip_list_page() -> tuple[str, InlineKeyboardMarkup]:
    categories = load_vip_categories()
    intro = (
        "🎓 <b>گروه VIP رواق</b>\n\n"
        "یک جهشِ واقعی در مسیرِ معماری و عمران: هر نرم‌افزار با تاپیکِ اختصاصیِ خودش، "
        "آموزش‌های ویدئوییِ کامل از صفر تا حرفه‌ای، آرشیوِ به‌روزِ پلاگین‌ها و متریال‌های نایاب، "
        "کتابخانه‌ی ضوابط و بانکِ پروژه، و آموزشِ اصولیِ هوش مصنوعی در معماری — "
        "هر چیزی که برای پیشرفت لازم دارید، یک‌جا.\n\n"
        "👇 یکی از دسته‌بندی‌های زیر رو انتخاب کنید:"
    )
    rows = []
    row = []
    for i, cat in enumerate(categories):
        row.append(InlineKeyboardButton(text=cat["name"], callback_data=f"vipnav:{i}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="🔙 بازگشت به پنل اصلی", callback_data="menu:back", style="danger")])
    return intro, InlineKeyboardMarkup(inline_keyboard=rows)

@dp.callback_query(F.data == "vip:list")
async def cb_vip_list(callback: CallbackQuery):
    text, keyboard = await render_vip_list_page()
    await show_text_panel(callback, text, keyboard)
    await callback.answer()

@dp.callback_query(F.data == "vip:open")
async def cb_vip_open(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    if VIP_GROUP_CHAT_ID is None:
        await callback.answer("گروه VIP هنوز راه‌اندازی نشده است.", show_alert=True)
        return
    caption, keyboard, image_id = await render_vip_page(0)
    await show_vip_page(callback, caption, keyboard, image_id)
    await callback.answer()

@dp.callback_query(F.data.startswith("vipnav:"))
async def cb_vip_nav(callback: CallbackQuery):
    try:
        index = int(callback.data.split(":", 1)[1])
    except ValueError:
        index = 0
    caption, keyboard, image_id = await render_vip_page(index)
    await show_vip_page(callback, caption, keyboard, image_id)
    await callback.answer()

# ---------- انتخاب مدت اشتراک ----------
class VipSubscriptionStates(StatesGroup):
    choosing_duration = State()
    waiting_for_receipt = State()

@dp.callback_query(F.data == "vip:buy_subscription")
async def cb_vip_buy_subscription(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        settings = load_vip_global_settings()
        prices = settings.get("prices", {})
        discount = settings.get("discount_percent", 0)
        
        text = "💎 <b>انتخاب مدت اشتراک VIP</b>\n\n"
        for months in (3, 6, 12):
            price = prices.get(str(months), 0)
            if discount > 0:
                final_price = int(price * (1 - discount / 100))
                text += (
                    f"▫️ {to_persian_num(months)} ماهه: "
                    f"<s>{format_toman(price)}</s> → {format_toman(final_price)} "
                    f"(تخفیف {to_persian_num(discount)}%)\n"
                )
            else:
                text += f"▫️ {to_persian_num(months)} ماهه: {format_toman(price)}\n"
        
        text += "\nلطفاً یکی از گزینه‌های بالا را انتخاب کنید."
        
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    # تلگرام کیبورد را برای زبان‌های راست‌به‌چپ خودکار آینه نمی‌کند؛
                    # ترتیبِ آرایه همیشه چپ‌به‌راستِ روی صفحه است. برای اینکه «۳ ماهه»
                    # سمت راست (نزدیک‌تر به شست) بیفتد، باید آخرین آیتمِ آرایه باشد.
                    InlineKeyboardButton(text=f"{to_persian_num(12)} ماهه", callback_data="vip:duration:12", style="primary"),
                    InlineKeyboardButton(text=f"{to_persian_num(6)} ماهه", callback_data="vip:duration:6", style="primary"),
                    InlineKeyboardButton(text=f"{to_persian_num(3)} ماهه", callback_data="vip:duration:3", style="primary"),
                ],
                [InlineKeyboardButton(text="❌ انصراف", callback_data="vip:cancel_payment", style="danger")],
            ]
        )
        await callback.message.delete()
        await callback.message.answer(text, reply_markup=keyboard)
        await state.set_state(VipSubscriptionStates.choosing_duration)
        await callback.answer()
    except Exception as e:
        logger.error(f"خطا در باز کردن صفحه خرید: {e}")
        await callback.answer("متأسفانه خطایی رخ داد. لطفاً دوباره تلاش کنید.", show_alert=True)

@dp.callback_query(F.data.startswith("vip:duration:"), VipSubscriptionStates.choosing_duration)
async def cb_vip_duration_chosen(callback: CallbackQuery, state: FSMContext):
    months = int(callback.data.split(":", 2)[2])
    settings = load_vip_global_settings()
    price = settings.get("prices", {}).get(str(months), 0)
    discount = settings.get("discount_percent", 0)
    if discount > 0:
        final_price = int(price * (1 - discount / 100))
    else:
        final_price = price
    
    await state.update_data(vip_months=months, vip_price=final_price)
    await state.set_state(VipSubscriptionStates.waiting_for_receipt)
    
    price_line = ""
    if discount > 0:
        price_line = (
            f"💰 مبلغ اصلی: <s>{format_toman(price)}</s>\n"
            f"💰 مبلغ پس از تخفیف {to_persian_num(discount)}%: <b>{format_toman(final_price)}</b>"
        )
    else:
        price_line = f"💰 مبلغ: <b>{format_toman(final_price)}</b>"
    
    text = sign(
        f"💳 <b>پرداختِ اشتراکِ VIP</b>\n\n"
        f"🗓 مدت: <b>{to_persian_num(months)} ماهه</b>\n"
        f"{price_line}\n\n"
        f"لطفاً مبلغ فوق را به شماره‌کارتِ زیر واریز کنید:\n\n"
        f"<code>{VIP_CARD_NUMBER}</code>\n"
        f"به نام: {html_escape(VIP_CARD_HOLDER)}\n\n"
        "(روی شماره‌کارت بزنید تا کپی شود)\n\n"
        "📸 پس از واریز، عکسِ فیش یا رسیدِ پرداخت را همین‌جا ارسال کنید.\n"
        "پس از تاییدِ ادمین، لینکِ ورود به گروهِ VIP برایتان ارسال می‌شود."
    )
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="❌ انصراف", callback_data="vip:cancel_payment", style="danger")]]
    )
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()

@dp.callback_query(F.data == "vip:cancel_payment")
async def cb_vip_cancel_payment(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("عملیاتِ خرید لغو شد.", reply_markup=user_panel_keyboard())
    await callback.answer()

def vip_admin_decision_keyboard(payment_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(text="✅ تایید و ارسالِ لینک", callback_data=f"vipadm:approve:{payment_id}", style="success"),
            InlineKeyboardButton(text="❌ رد کردن", callback_data=f"vipadm:reject:{payment_id}", style="danger"),
        ]]
    )

@dp.message(VipSubscriptionStates.waiting_for_receipt)
async def handle_vip_receipt(message: Message, state: FSMContext):
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=user_panel_keyboard())
        return

    photo_file_id = None
    if message.photo:
        photo_file_id = message.photo[-1].file_id
    elif message.document and (message.document.mime_type or "").startswith("image/"):
        photo_file_id = message.document.file_id

    if not photo_file_id:
        await message.answer("لطفاً فقط عکسِ فیش یا رسیدِ پرداخت را ارسال کنید.")
        return

    data = await state.get_data()
    months = data.get("vip_months")
    price = data.get("vip_price")
    if not months or not price:
        await state.clear()
        await message.answer("مسیرِ خرید قطع شده. لطفاً دوباره از پنلِ VIP اقدام کنید.", reply_markup=user_panel_keyboard())
        return

    user = message.from_user
    payment_id = f"pay_{uuid.uuid4().hex[:10]}"
    payment_record = {
        "id": payment_id,
        "user_id": user.id,
        "user_display": user.full_name or str(user.id),
        "username": user.username,
        "category_id": "all",
        "category_name": "اشتراک کامل VIP",
        "months": months,
        "price": price,
        "photo_file_id": photo_file_id,
        "status": "pending",
        "requested_at": datetime.utcnow().isoformat(),
        "admin_messages": [],
    }

    payments = load_vip_payments()
    payments[payment_id] = payment_record
    await save_vip_payments(payments)
    await state.clear()

    await message.answer(
        sign(
            "✅ <b>رسید شما دریافت شد</b>\n\n"
            "فیشِ پرداخت برای بررسیِ ادمین ارسال شد.\n"
            "به‌محضِ تایید، لینکِ ورود به گروهِ VIP برایتان ارسال می‌شود."
        )
    )

    username_part = f"@{user.username}" if user.username else f"<code>{user.id}</code>"
    caption = (
        f"💳 <b>درخواستِ جدیدِ اشتراکِ VIP</b>\n\n"
        f"👤 {html_escape(user.full_name or str(user.id))} ({username_part})\n"
        f"🗓 مدت: <b>{to_persian_num(months)} ماهه</b>\n"
        f"💰 مبلغ: <b>{format_toman(price)}</b>\n\n"
        f"شناسه پرداخت: <code>{payment_id}</code>"
    )
    for admin_id in ADMIN_IDS:
        try:
            sent = await bot.send_photo(
                chat_id=admin_id,
                photo=photo_file_id,
                caption=caption,
                reply_markup=vip_admin_decision_keyboard(payment_id),
            )
            payment_record["admin_messages"].append({"chat_id": admin_id, "message_id": sent.message_id})
        except Exception as e:
            logger.warning("ارسالِ درخواستِ پرداخت به ادمین %s ممکن نشد: %s", admin_id, e)

    payments = load_vip_payments()
    if payment_id in payments:
        payments[payment_id]["admin_messages"] = payment_record["admin_messages"]
        await save_vip_payments(payments)

@dp.callback_query(F.data.startswith("vipadm:"))
async def cb_vip_admin_decision(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return

    _, action, payment_id = callback.data.split(":", 2)
    payments = load_vip_payments()
    payment = payments.get(payment_id)
    if not payment:
        await callback.answer("این درخواست دیگر یافت نشد.", show_alert=True)
        return
    if payment["status"] != "pending":
        await callback.answer("این درخواست قبلاً بررسی شده است.", show_alert=True)
        return

    user_id = payment["user_id"]
    admin_name = callback.from_user.full_name or str(callback.from_user.id)

    if action == "reject":
        payment["status"] = "rejected"
        payment["decided_by"] = callback.from_user.id
        payment["decided_at"] = datetime.utcnow().isoformat()
        payments[payment_id] = payment
        await save_vip_payments(payments)

        for ref in payment.get("admin_messages", []):
            try:
                await bot.edit_message_reply_markup(chat_id=ref["chat_id"], message_id=ref["message_id"], reply_markup=None)
            except Exception:
                pass
        try:
            await callback.message.edit_caption(caption=callback.message.caption + f"\n\n❌ رد شد توسط {html_escape(admin_name)}")
        except Exception:
            pass

        try:
            await bot.send_message(
                chat_id=user_id,
                text=sign(
                    "❌ <b>پرداختِ شما تایید نشد</b>\n\n"
                    "ممکن است رسیدِ ارسالی خوانا نبوده یا مبلغ مطابقت نداشته باشد.\n"
                    "برای پیگیری، از «📞 ارتباط با ادمین» استفاده کنید یا دوباره از پنلِ VIP اقدام کنید."
                ),
            )
        except Exception as e:
            logger.warning("اطلاع‌رسانیِ ردِ پرداخت به کاربر %s ممکن نشد: %s", user_id, e)

        await callback.answer("درخواست رد شد.")
        return

    if action == "approve":
        if VIP_GROUP_CHAT_ID is None:
            await callback.answer("آیدیِ گروهِ VIP تنظیم نشده است.", show_alert=True)
            return

        try:
            invite = await bot.create_chat_invite_link(
                chat_id=VIP_GROUP_CHAT_ID,
                member_limit=1,
                name=f"vip-{user_id}-{payment['months']}m",
            )
        except Exception as e:
            logger.error("ساختِ لینکِ دعوتِ VIP ناموفق بود: %s", e)
            await callback.answer(f"خطا در ساختِ لینکِ دعوت: {e}", show_alert=True)
            return

        now = datetime.utcnow()
        days = VIP_MONTHS_TO_DAYS.get(payment["months"], payment["months"] * 30)

        subs = load_vip_subscriptions()
        user_subs = subs.setdefault(str(user_id), [])

        # اگر کاربر یک اشتراکِ «فعالِ» قبلی داشته باشد (تمدیدِ زودهنگام)، مدتِ
        # جدید را از تاریخِ پایانِ همان اشتراک اضافه می‌کنیم، نه از همین لحظه؛
        # در غیرِ این صورت چند روزِ باقی‌مانده از خریدِ قبلی کاربر هدر می‌رفت.
        # هم‌زمان رکوردهای «active» قبلی را به «renewed» تغییر می‌دهیم تا
        # حلقه‌ی بررسیِ انقضا (که فقط رکوردهای active را پردازش می‌کند) با
        # چند رکوردِ هم‌پوشان اشتباه نکند و کاربر را زودتر از موعد از گروه حذف نکند.
        previous_active_end = None
        for sub in user_subs:
            if sub.get("status") == "active":
                try:
                    sub_end = datetime.fromisoformat(sub["end"])
                except (KeyError, ValueError):
                    sub_end = None
                if sub_end and sub_end > now and (previous_active_end is None or sub_end > previous_active_end):
                    previous_active_end = sub_end
                sub["status"] = "renewed"

        start = previous_active_end if previous_active_end else now
        end = start + timedelta(days=days)

        user_subs.append({
            "category_id": "all",
            "category_name": "اشتراک کامل VIP",
            "months": payment["months"],
            "price": payment["price"],
            "start": start.isoformat(),
            "end": end.isoformat(),
            "status": "active",
            "reminded": False,
        })
        await save_vip_subscriptions(subs)

        payment["status"] = "approved"
        payment["decided_by"] = callback.from_user.id
        payment["decided_at"] = now.isoformat()
        payments[payment_id] = payment
        await save_vip_payments(payments)

        for ref in payment.get("admin_messages", []):
            try:
                await bot.edit_message_reply_markup(chat_id=ref["chat_id"], message_id=ref["message_id"], reply_markup=None)
            except Exception:
                pass
        try:
            await callback.message.edit_caption(caption=callback.message.caption + f"\n\n✅ تایید شد توسط {html_escape(admin_name)}")
        except Exception:
            pass

        end_jalali = format_jalali_datetime(end)
        try:
            await bot.send_message(
                chat_id=user_id,
                text=sign(
                    "✅ <b>پرداختِ شما تایید شد</b>\n\n"
                    f"🗓 مدت: <b>{to_persian_num(payment['months'])} ماهه</b>\n"
                    f"⏳ تا تاریخِ: <b>{end_jalali}</b>\n\n"
                    "برای ورود به گروهِ VIP از لینکِ زیر استفاده کنید (این لینک فقط یک‌بار قابلِ استفاده است):"
                ),
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[InlineKeyboardButton(text="🌟 ورود به گروهِ VIP", url=invite.invite_link)]]
                ),
            )
        except Exception as e:
            logger.warning("ارسالِ لینکِ VIP به کاربر %s ممکن نشد: %s", user_id, e)

        await callback.answer("تایید شد و لینک ارسال شد.")
        return

# ---------- بررسیِ دوره‌ایِ انقضای اشتراک ----------
async def vip_expiry_checker_loop() -> None:
    while True:
        try:
            await _check_vip_expirations()
        except Exception as e:
            logger.error("خطا در بررسیِ انقضای اشتراک‌های VIP: %s", e)
        await asyncio.sleep(3 * 3600)

async def _check_vip_expirations() -> None:
    subs = load_vip_subscriptions()
    now = datetime.utcnow()
    changed = False

    for user_id_str, user_subs in subs.items():
        for sub in user_subs:
            if sub.get("status") != "active":
                continue
            end = datetime.fromisoformat(sub["end"])
            remaining_days = (end - now).total_seconds() / 86400

            if 0 < remaining_days <= 3 and not sub.get("reminded"):
                sub["reminded"] = True
                changed = True
                try:
                    await bot.send_message(
                        chat_id=int(user_id_str),
                        text=sign(
                            f"⏳ <b>یادآوریِ اشتراکِ VIP</b>\n\n"
                            f"اشتراکِ VIP شما تا "
                            f"<b>{to_persian_num(int(remaining_days) + 1)}</b> روزِ دیگر به پایان می‌رسد.\n"
                            "برای تمدید، از پنلِ VIP اقدام کنید."
                        ),
                        reply_markup=InlineKeyboardMarkup(
                            inline_keyboard=[[InlineKeyboardButton(text="🌟 تمدیدِ اشتراک", callback_data="vip:open", style="success")]]
                        ),
                    )
                except Exception:
                    pass

            elif remaining_days <= 0:
                sub["status"] = "expired"
                changed = True

                # محافظِ ایمنی: اگر رکوردِ دیگری (مثلاً یک تمدیدِ ثبت‌شده با داده‌های
                # قدیمی‌تر از این اصلاح) هنوز تا آینده معتبر است، کاربر نباید حذف شود.
                other_active_end = None
                for other in user_subs:
                    if other is sub or other.get("status") not in ("active", "renewed"):
                        continue
                    try:
                        other_end = datetime.fromisoformat(other["end"])
                    except (KeyError, ValueError):
                        continue
                    if other_end > now:
                        other_active_end = other_end
                        break
                if other_active_end is not None:
                    continue

                if VIP_GROUP_CHAT_ID is not None:
                    try:
                        await bot.ban_chat_member(chat_id=VIP_GROUP_CHAT_ID, user_id=int(user_id_str))
                        await bot.unban_chat_member(chat_id=VIP_GROUP_CHAT_ID, user_id=int(user_id_str), only_if_banned=True)
                    except Exception as e:
                        logger.warning("حذفِ خودکارِ کاربرِ %s از گروهِ VIP ممکن نشد: %s", user_id_str, e)
                try:
                    await bot.send_message(
                        chat_id=int(user_id_str),
                        text=sign(
                            f"⌛️ اشتراکِ VIP شما به پایان رسید.\n"
                            "برای تمدید، از پنلِ VIP اقدام کنید."
                        ),
                        reply_markup=InlineKeyboardMarkup(
                            inline_keyboard=[[InlineKeyboardButton(text="🌟 تمدیدِ اشتراک", callback_data="vip:open", style="success")]]
                        ),
                    )
                except Exception:
                    pass
                if NOTIFY_CHAT_ID_INT:
                    try:
                        await bot.send_message(
                            chat_id=NOTIFY_CHAT_ID_INT,
                            text=f"⌛️ اشتراکِ VIP کاربر <code>{user_id_str}</code> به پایان رسید.",
                        )
                    except Exception:
                        pass

    if changed:
        await save_vip_subscriptions(subs)

# ==============================================================
#  مدیریتِ ادمین روی مشترکینِ VIP — لیست، جزئیات، تمدید و لغو
# ==============================================================

VIP_SUBSCRIBERS_PAGE_SIZE = 6

class VipAdminManageStates(StatesGroup):
    waiting_extend_days = State()

async def _display_name_for(user_id: int) -> str:
    """نامِ نمایشیِ کاربر برای پنلِ ادمین: اول از کشِ پروفایل، وگرنه از تلگرام."""
    record = _user_cache.get(str(user_id))
    if record and record.get("full_name"):
        return record["full_name"]
    try:
        chat = await bot.get_chat(user_id)
        return chat.full_name or (f"@{chat.username}" if chat.username else str(user_id))
    except Exception:
        return str(user_id)

def _vip_subscriber_ids(subs: dict) -> list[int]:
    """
    فهرستِ آیدیِ همه‌ی کسانی که حداقل یک رکوردِ اشتراک (فعال، تمدیدشده، منقضی یا
    لغوشده) داشته‌اند، مرتب‌شده: فعال‌ها بر اساسِ نزدیک‌ترین تاریخِ پایان اول،
    سپس بقیه بر اساسِ آخرین تاریخِ پایان (نزولی).
    """
    now = datetime.utcnow()
    active_rows: list[tuple[datetime, int]] = []
    other_rows: list[tuple[datetime, int]] = []
    for uid_str in subs.keys():
        try:
            uid = int(uid_str)
        except ValueError:
            continue
        status = get_user_vip_status(uid, subs)
        if not status["has_subscription"]:
            continue
        if status["is_active"]:
            active_rows.append((status["end"], uid))
        else:
            other_rows.append((status["end"] or datetime.min, uid))
    active_rows.sort(key=lambda t: t[0])
    other_rows.sort(key=lambda t: t[0], reverse=True)
    return [uid for _, uid in active_rows] + [uid for _, uid in other_rows]

async def render_vip_subscribers_page(page: int) -> tuple[str, InlineKeyboardMarkup]:
    subs = load_vip_subscriptions()
    ids = _vip_subscriber_ids(subs)

    if not ids:
        text = "📋 <b>مشترکینِ VIP</b>\n\nهنوز هیچ کاربری اشتراکِ VIP نداشته است."
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 بازگشت", callback_data="admin:cat_vip", style="primary")]
        ])
        return text, keyboard

    total_pages = max(1, (len(ids) + VIP_SUBSCRIBERS_PAGE_SIZE - 1) // VIP_SUBSCRIBERS_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    page_ids = ids[page * VIP_SUBSCRIBERS_PAGE_SIZE: (page + 1) * VIP_SUBSCRIBERS_PAGE_SIZE]

    text = (
        f"📋 <b>مشترکینِ VIP</b> "
        f"({to_persian_num(page + 1)}/{to_persian_num(total_pages)})\n\n"
        f"مجموع: {to_persian_num(len(ids))} نفر\n"
        "برای مدیریتِ هرکاربر روی نامش بزنید 👇"
    )

    rows = []
    for uid in page_ids:
        status = get_user_vip_status(uid, subs)
        name = await _display_name_for(uid)
        if status["is_active"]:
            label = f"✅ {name} — {to_persian_num(status['remaining_days'])} روزِ دیگر"
        else:
            label = f"⌛️ {name} — منقضی"
        rows.append([InlineKeyboardButton(text=label, callback_data=f"vipadmin:user:{page}:{uid}")])

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton(text="◀️ قبلی", callback_data=f"vipadmin:list:{page - 1}", style="primary"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton(text="بعدی ▶️", callback_data=f"vipadmin:list:{page + 1}", style="primary"))
    if nav_row:
        rows.append(nav_row)

    rows.append([InlineKeyboardButton(text="🔙 بازگشت", callback_data="admin:cat_vip", style="primary")])
    return text, InlineKeyboardMarkup(inline_keyboard=rows)

async def render_vip_subscriber_detail(user_id: int, back_page: int) -> tuple[str, InlineKeyboardMarkup]:
    subs = load_vip_subscriptions()
    status = get_user_vip_status(user_id, subs)
    name = await _display_name_for(user_id)
    user_subs = subs.get(str(user_id), [])

    if status["is_active"]:
        status_line = (
            f"✅ فعال — {to_persian_num(status['remaining_days'])} روزِ دیگر باقی مانده\n"
            f"⏳ تا تاریخِ: {format_jalali_datetime(status['end'])}"
        )
    elif status["has_subscription"]:
        end_str = format_jalali_datetime(status["end"]) if status["end"] else "نامشخص"
        status_line = f"⌛️ منقضی‌شده — پایان: {end_str}"
    else:
        status_line = "بدونِ سابقه‌ی اشتراک"

    history_lines = []
    for sub in sorted(user_subs, key=lambda s: s.get("start", ""), reverse=True)[:5]:
        # >>> تغییر برای تشخیص پاداش VIP <<<
        if sub.get("category_id") == "reward":
            months_label = "🎁 پاداشِ مدیریت"
        elif sub.get("months"):
            months_label = f"{to_persian_num(sub['months'])} ماهه"
        else:
            months_label = "تمدیدِ دستیِ ادمین"
        try:
            end_h = format_jalali_datetime(datetime.fromisoformat(sub["end"]))
        except (KeyError, ValueError):
            end_h = "-"
        status_icon = {
            "active": "🟢", "renewed": "🔁", "expired": "⌛️", "cancelled": "❌",
        }.get(sub.get("status"), "•")
        history_lines.append(f"{status_icon} {months_label} — تا {end_h}")
    history_text = "\n".join(history_lines) if history_lines else "رکوردی موجود نیست."

    text = (
        f"👤 <b>{html_escape(name)}</b>\n"
        f"🆔 <code>{user_id}</code>\n\n"
        f"{status_line}\n\n"
        f"🗂 <b>تاریخچه:</b>\n{history_text}"
    )

    rows = [
        [InlineKeyboardButton(text="➕ تمدید / افزودنِ اعتبار", callback_data=f"vipadmin:extend:{back_page}:{user_id}", style="success")],
    ]
    if status["is_active"]:
        rows.append([InlineKeyboardButton(text="❌ لغوِ اشتراک", callback_data=f"vipadmin:revoke:{back_page}:{user_id}", style="danger")])
    rows.append([InlineKeyboardButton(text="🔙 بازگشت به لیست", callback_data=f"vipadmin:list:{back_page}", style="primary")])

    return text, InlineKeyboardMarkup(inline_keyboard=rows)

async def _grant_or_extend_vip(user_id: int, days: int, granted_by: int) -> tuple[bool, str, datetime | None]:
    """
    به کاربر days روز اعتبارِ VIP اضافه می‌کند (اگر اشتراکِ فعالی داشته باشد،
    از تاریخِ پایانِ همان اضافه می‌شود؛ وگرنه از همین لحظه). اگر کاربر عضوِ
    گروهِ VIP نباشد، لینکِ دعوتِ یک‌بارمصرف می‌سازد و برایش می‌فرستد.
    خروجی: (موفقیت، پیامِ توضیحی، تاریخِ پایانِ جدید).
    """
    if VIP_GROUP_CHAT_ID is None:
        return False, "آیدیِ گروهِ VIP تنظیم نشده است.", None

    now = datetime.utcnow()
    subs = load_vip_subscriptions()
    user_subs = subs.setdefault(str(user_id), [])

    previous_active_end = None
    for sub in user_subs:
        if sub.get("status") == "active":
            try:
                sub_end = datetime.fromisoformat(sub["end"])
            except (KeyError, ValueError):
                sub_end = None
            if sub_end and sub_end > now and (previous_active_end is None or sub_end > previous_active_end):
                previous_active_end = sub_end
            sub["status"] = "renewed"

    start = previous_active_end if previous_active_end else now
    end = start + timedelta(days=days)

    user_subs.append({
        "category_id": "admin_grant",
        "category_name": "اعطای دستیِ ادمین",
        "months": None,
        "price": 0,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "status": "active",
        "reminded": False,
        "granted_by": granted_by,
    })
    await save_vip_subscriptions(subs)

    is_member = False
    try:
        member = await bot.get_chat_member(VIP_GROUP_CHAT_ID, user_id)
        is_member = member.status in (ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR)
    except Exception:
        is_member = False

    end_jalali = format_jalali_datetime(end)
    if is_member:
        try:
            await bot.send_message(
                chat_id=user_id,
                text=sign(
                    f"🌟 <b>اشتراکِ VIP شما به‌روزرسانی شد</b>\n\n"
                    f"⏳ تا تاریخِ: <b>{end_jalali}</b>"
                ),
            )
        except Exception as e:
            logger.warning("اطلاع‌رسانیِ تمدید به کاربر %s ممکن نشد: %s", user_id, e)
        return True, f"اعتبار تا {end_jalali} تمدید شد (کاربر از قبل عضوِ گروه بود).", end
    else:
        try:
            invite = await bot.create_chat_invite_link(
                chat_id=VIP_GROUP_CHAT_ID, member_limit=1, name=f"vip-admin-{user_id}",
            )
        except Exception as e:
            logger.error("ساختِ لینکِ دعوتِ VIP ناموفق بود: %s", e)
            return True, f"اعتبار تا {end_jalali} ثبت شد اما ساختِ لینکِ دعوت ناموفق بود: {e}", end
        try:
            await bot.send_message(
                chat_id=user_id,
                text=sign(
                    f"🌟 <b>دسترسیِ VIP برایتان فعال شد</b>\n\n"
                    f"⏳ تا تاریخِ: <b>{end_jalali}</b>\n\n"
                    "برای ورود به گروهِ VIP از لینکِ زیر استفاده کنید (این لینک فقط یک‌بار قابلِ استفاده است):"
                ),
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[InlineKeyboardButton(text="🌟 ورود به گروهِ VIP", url=invite.invite_link)]]
                ),
            )
        except Exception as e:
            logger.warning("ارسالِ لینکِ VIP به کاربر %s ممکن نشد: %s", user_id, e)
        return True, f"اعتبار تا {end_jalali} ثبت و لینکِ ورود برای کاربر ارسال شد.", end

async def _grant_vip_reward(
    user_id: int, days: int, reason: str, granted_by: int
) -> tuple[bool, str, datetime | None]:
    """
    نسخه‌ی مستقلِ اعطای اعتبارِ VIP، برای وقتی که ادمین می‌خواهد به‌عنوانِ
    «پاداش/تشویق» (نه تمدیدِ خرید، نه تاییدِ فیش) به کاربری دسترسیِ VIP بدهد.

    تفاوت با _grant_or_extend_vip:
      ۱) رکوردِ اشتراک با category_id="reward" ثبت می‌شود تا در تاریخچه‌ی
         مشترکین با آیکنِ 🎁 از خرید/تمدیدِ عادی قابلِ تشخیص باشد.
      ۲) متنِ پیامِ ارسالی به کاربر کاملاً مجزا و با لحنِ «هدیه از طرفِ
         مدیریت» است، نه «تاییدِ پرداخت» یا «به‌روزرسانیِ اشتراک».

    منطقِ محاسبه‌ی روزها (اگر کاربر از قبل اشتراکِ فعال داشته باشد، از
    تاریخِ پایانِ همان اضافه می‌شود؛ وگرنه از همین لحظه) دقیقاً مثلِ بقیه‌ی
    مسیرهای VIP است تا هیچ روزی از قلم نیفتد.

    خروجی: (موفقیت، پیامِ توضیحی برای ادمین، تاریخِ پایانِ جدید).
    """
    if VIP_GROUP_CHAT_ID is None:
        return False, "آیدیِ گروهِ VIP تنظیم نشده است.", None

    now = datetime.utcnow()
    subs = load_vip_subscriptions()
    user_subs = subs.setdefault(str(user_id), [])

    previous_active_end = None
    for sub in user_subs:
        if sub.get("status") == "active":
            try:
                sub_end = datetime.fromisoformat(sub["end"])
            except (KeyError, ValueError):
                sub_end = None
            if sub_end and sub_end > now and (previous_active_end is None or sub_end > previous_active_end):
                previous_active_end = sub_end
            sub["status"] = "renewed"

    start = previous_active_end if previous_active_end else now
    end = start + timedelta(days=days)

    user_subs.append({
        "category_id": "reward",
        "category_name": "🎁 پاداشِ مدیریت",
        "months": None,
        "price": 0,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "status": "active",
        "reminded": False,
        "granted_by": granted_by,
        "reason": reason or None,
    })
    await save_vip_subscriptions(subs)

    reason_line = f"\n\n📝 <i>{html_escape(reason)}</i>" if reason else ""
    end_jalali = format_jalali_datetime(end)
    reward_text = sign(
        "🎁 <b>یک خبرِ خوب برایتان داریم!</b>\n\n"
        f"به‌پاسِ فعالیتِ خوبِ شما در رواق، تیمِ مدیریت تصمیم گرفت "
        f"<b>{to_persian_num(days)} روز</b> دسترسیِ رایگانِ VIP را به‌عنوانِ پاداش برایتان فعال کند 🌟"
        f"{reason_line}\n\n"
        f"⏳ این دسترسی تا تاریخِ <b>{end_jalali}</b> معتبر است.\n\n"
        "این هدیه‌ای از طرفِ مدیریتِ رواق است و ربطی به خریدِ اشتراک ندارد؛ "
        "امیدواریم لذت ببرید 🏛"
    )

    is_member = False
    try:
        member = await bot.get_chat_member(VIP_GROUP_CHAT_ID, user_id)
        is_member = member.status in (ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR)
    except Exception:
        is_member = False

    if is_member:
        try:
            await bot.send_message(chat_id=user_id, text=reward_text)
        except Exception as e:
            logger.warning("اطلاع‌رسانیِ پاداشِ VIP به کاربر %s ممکن نشد: %s", user_id, e)
            return True, f"اعتبار تا {end_jalali} ثبت شد اما ارسالِ پیام به کاربر ناموفق بود: {e}", end
        return True, f"پاداش ثبت شد؛ کاربر از قبل عضوِ گروهِ VIP بود (اعتبار تا {end_jalali}).", end

    try:
        invite = await bot.create_chat_invite_link(
            chat_id=VIP_GROUP_CHAT_ID, member_limit=1, name=f"vip-reward-{user_id}",
        )
    except Exception as e:
        logger.error("ساختِ لینکِ دعوتِ VIP (پاداش) ناموفق بود: %s", e)
        try:
            await bot.send_message(chat_id=user_id, text=reward_text)
        except Exception:
            pass
        return True, f"اعتبار تا {end_jalali} ثبت شد اما ساختِ لینکِ دعوت ناموفق بود: {e}", end

    try:
        await bot.send_message(
            chat_id=user_id,
            text=reward_text + "\n\nبرای ورود به گروهِ VIP از لینکِ زیر استفاده کنید (این لینک فقط یک‌بار قابلِ استفاده است):",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="🌟 ورود به گروهِ VIP", url=invite.invite_link)]]
            ),
        )
    except Exception as e:
        logger.warning("ارسالِ لینکِ VIP (پاداش) به کاربر %s ممکن نشد: %s", user_id, e)
        return True, f"اعتبار تا {end_jalali} ثبت شد اما ارسالِ پیام به کاربر ناموفق بود: {e}", end

    return True, f"پاداش ثبت و لینکِ ورودِ گروهِ VIP برای کاربر ارسال شد (اعتبار تا {end_jalali}).", end

async def _revoke_vip(user_id: int, revoked_by: int) -> tuple[bool, str]:
    if VIP_GROUP_CHAT_ID is None:
        return False, "آیدیِ گروهِ VIP تنظیم نشده است."

    subs = load_vip_subscriptions()
    user_subs = subs.get(str(user_id))
    if not user_subs:
        return False, "این کاربر سابقه‌ی اشتراکی ندارد."

    had_active = False
    for sub in user_subs:
        if sub.get("status") in ("active", "renewed"):
            sub["status"] = "cancelled"
            sub["cancelled_by"] = revoked_by
            sub["cancelled_at"] = datetime.utcnow().isoformat()
            had_active = True
    await save_vip_subscriptions(subs)

    try:
        await bot.ban_chat_member(chat_id=VIP_GROUP_CHAT_ID, user_id=user_id)
        await bot.unban_chat_member(chat_id=VIP_GROUP_CHAT_ID, user_id=user_id, only_if_banned=True)
    except Exception as e:
        logger.warning("حذفِ کاربرِ %s توسط ادمین از گروهِ VIP ممکن نشد: %s", user_id, e)

    try:
        await bot.send_message(
            chat_id=user_id,
            text=sign("⚠️ اشتراکِ VIP شما توسط ادمین لغو شد و از گروه حذف شدید."),
        )
    except Exception:
        pass

    if not had_active:
        return True, "کاربر اشتراکِ فعالی نداشت؛ فقط از گروه حذف شد (برای اطمینان)."
    return True, "اشتراک لغو شد و کاربر از گروهِ VIP حذف شد."

@dp.callback_query(F.data.startswith("vipadmin:list:"))
async def cb_vipadmin_list(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    await state.clear()
    page = int(callback.data.split(":")[2])
    text, keyboard = await render_vip_subscribers_page(page)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()

@dp.callback_query(F.data.startswith("vipadmin:user:"))
async def cb_vipadmin_user(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    await state.clear()
    _, _, back_page, uid = callback.data.split(":")
    text, keyboard = await render_vip_subscriber_detail(int(uid), int(back_page))
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()

@dp.callback_query(F.data.startswith("vipadmin:extend:"))
async def cb_vipadmin_extend(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    _, _, back_page, uid = callback.data.split(":")
    await state.set_state(VipAdminManageStates.waiting_extend_days)
    await state.update_data(target_user_id=int(uid), back_page=int(back_page))
    await callback.message.edit_text(
        "➕ <b>تمدید / افزودنِ اعتبار</b>\n\n"
        "تعدادِ روزی که می‌خواهید اضافه شود را وارد کنید (فقط عدد).\n"
        "مثال: برای یک ماه بنویسید 30\n\n"
        "اگر کاربر همین الان هم اشتراکِ فعال داشته باشد، این روزها به پایانِ اشتراکِ فعلی‌اش اضافه می‌شود.\n"
        "(برای لغو، /cancel بفرستید)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data=f"vipadmin:user:{back_page}:{uid}", style="primary")]
        ])
    )
    await callback.answer()

@dp.message(VipAdminManageStates.waiting_extend_days)
async def handle_vipadmin_extend_days(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        data = await state.get_data()
        await state.clear()
        text, keyboard = await render_vip_subscriber_detail(data["target_user_id"], data.get("back_page", 0))
        await message.answer("لغو شد.", reply_markup=keyboard)
        return

    days, ok = await _parse_price_or_discount(message)
    if not ok or not days or days <= 0:
        await message.answer("❌ لطفاً یک عددِ صحیحِ مثبت وارد کنید.")
        return

    data = await state.get_data()
    target_user_id = data["target_user_id"]
    back_page = data.get("back_page", 0)
    await state.clear()

    ok, result_msg, _ = await _grant_or_extend_vip(target_user_id, days, granted_by=message.from_user.id)
    icon = "✅" if ok else "❌"
    text, keyboard = await render_vip_subscriber_detail(target_user_id, back_page)
    await message.answer(f"{icon} {result_msg}")
    await message.answer(text, reply_markup=keyboard)

@dp.callback_query(F.data.startswith("vipadmin:revoke:"))
async def cb_vipadmin_revoke(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    _, _, back_page, uid = callback.data.split(":")
    await callback.message.edit_text(
        "❌ <b>لغوِ اشتراکِ VIP</b>\n\n"
        "آیا مطمئنید؟ کاربر بلافاصله از گروهِ VIP حذف می‌شود.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ بله، لغو شود", callback_data=f"vipadmin:revoke_confirm:{back_page}:{uid}", style="danger")],
            [InlineKeyboardButton(text="🔙 انصراف", callback_data=f"vipadmin:user:{back_page}:{uid}", style="primary")],
        ])
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("vipadmin:revoke_confirm:"))
async def cb_vipadmin_revoke_confirm(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    _, _, back_page, uid = callback.data.split(":")
    uid = int(uid)
    back_page = int(back_page)
    ok, result_msg = await _revoke_vip(uid, revoked_by=callback.from_user.id)
    icon = "✅" if ok else "❌"
    await callback.answer(f"{icon} {result_msg}", show_alert=True)
    text, keyboard = await render_vip_subscriber_detail(uid, back_page)
    await callback.message.edit_text(text, reply_markup=keyboard)

# ==============================================================
#  پنل تنظیماتِ VIP (ادمین) — دسته‌بندی‌ها و قیمت‌های جهانی
# ==============================================================

class VipCategoryStates(StatesGroup):
    waiting_new_name = State()
    waiting_new_desc = State()
    waiting_edit_value = State()
    waiting_banner = State()

class VipGlobalSettingsStates(StatesGroup):
    waiting_price3 = State()
    waiting_price6 = State()
    waiting_price12 = State()
    waiting_discount = State()


async def build_vip_settings_text() -> str:
    categories = load_vip_categories()
    if not categories:
        return "💎 <b>تنظیماتِ VIP</b>\n\nهنوز هیچ دسته‌بندی‌ای اضافه نشده است."
    lines = ["💎 <b>تنظیماتِ VIP</b>\n", "دسته‌بندی‌های فعلی:\n"]
    for i, cat in enumerate(categories, 1):
        lines.append(f"{to_persian_num(i)}. <b>{html_escape(cat['name'])}</b>\n   {html_escape(cat.get('description', ''))[:60]}...")
    return "\n".join(lines)

def vip_settings_keyboard() -> InlineKeyboardMarkup:
    categories = load_vip_categories()
    rows = []
    for i in range(0, len(categories), 2):
        row = [InlineKeyboardButton(text=f"✏️ {categories[i]['name']}", callback_data=f"vipset:edit:{categories[i]['id']}", style="primary")]
        if i + 1 < len(categories):
            row.append(InlineKeyboardButton(text=f"✏️ {categories[i+1]['name']}", callback_data=f"vipset:edit:{categories[i+1]['id']}", style="primary"))
        rows.append(row)
    rows.append([InlineKeyboardButton(text="➕ افزودنِ دسته‌بندیِ جدید", callback_data="vipset:add", style="success")])
    rows.append([InlineKeyboardButton(text="🔙 بازگشت", callback_data="admin:cat_vip", style="primary")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def vip_category_edit_keyboard(cat_id: str) -> InlineKeyboardMarkup:
    categories = load_vip_categories()
    index = next((i for i, c in enumerate(categories) if c["id"] == cat_id), None)
    reorder_row = []
    if index is not None:
        if index > 0:
            reorder_row.append(InlineKeyboardButton(text="⬆️ جابه‌جایی به بالا", callback_data=f"vipset:moveup:{cat_id}", style="primary"))
        if index < len(categories) - 1:
            reorder_row.append(InlineKeyboardButton(text="⬇️ جابه‌جایی به پایین", callback_data=f"vipset:movedown:{cat_id}", style="primary"))
    rows = [
        [InlineKeyboardButton(text="✏️ ویرایشِ نام", callback_data=f"vipset:field:{cat_id}:name", style="primary")],
        [InlineKeyboardButton(text="✏️ ویرایشِ توضیحات", callback_data=f"vipset:field:{cat_id}:description", style="primary")],
        [InlineKeyboardButton(text="🖼 آپلود/تغییر بنر", callback_data=f"vipset:banner:{cat_id}", style="primary")],
        [InlineKeyboardButton(text="🗑 حذف بنر", callback_data=f"vipset:delete_banner:{cat_id}", style="danger")],
    ]
    if reorder_row:
        rows.append(reorder_row)
    rows.append([InlineKeyboardButton(text="🗑 حذفِ این دسته‌بندی", callback_data=f"vipset:delete:{cat_id}", style="danger")])
    rows.append([InlineKeyboardButton(text="🔙 بازگشت به لیست", callback_data="admin:vip_settings", style="danger")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

@dp.callback_query(F.data == "vipset:add")
async def cb_vipset_add(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await state.set_state(VipCategoryStates.waiting_new_name)
    await callback.message.edit_text(
        "➕ <b>افزودنِ دسته‌بندیِ جدید</b>\n\nنامِ دسته‌بندی را ارسال کنید:\n(برای لغو، /cancel بفرستید)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data="admin:vip_settings", style="primary")]
        ]),
    )
    await callback.answer()

@dp.message(VipCategoryStates.waiting_new_name)
async def handle_vipset_new_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_settings_text(), reply_markup=vip_settings_keyboard())
        return
    await state.update_data(new_cat_name=message.text.strip())
    await state.set_state(VipCategoryStates.waiting_new_desc)
    await message.answer("توضیحاتِ دسته‌بندی را ارسال کنید (می‌توانید از خط جدید استفاده کنید):")

@dp.message(VipCategoryStates.waiting_new_desc)
async def handle_vipset_new_desc(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_settings_text(), reply_markup=vip_settings_keyboard())
        return
    
    data = await state.get_data()
    categories = load_vip_categories()
    new_cat = {
        "id": f"cat_{uuid.uuid4().hex[:8]}",
        "name": data["new_cat_name"],
        "description": message.text.strip(),
        "created_at": datetime.utcnow().isoformat(),
        "image_file_id": None,
    }
    categories.append(new_cat)
    await save_vip_categories(categories)
    await state.clear()

    # به‌جای برگشتن به منوی اصلی، مستقیم می‌رویم روی صفحه‌ی ویرایشِ همین دسته‌بندیِ
    # تازه‌ساخته‌شده تا ادمین بلافاصله بتواند بنرش را آپلود کند، بدونِ اینکه لازم
    # باشد دوباره از اول (منو ← VIP ← تنظیمات ← پیداکردنِ دسته‌بندی) مسیر را طی کند.
    await message.answer(
        f"✅ دسته‌بندیِ «{html_escape(new_cat['name'])}» با موفقیت اضافه شد.\n\n"
        f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
        f"نام: <b>{html_escape(new_cat['name'])}</b>\n"
        f"توضیحات:\n{html_escape(new_cat.get('description', ''))}\n\n"
        "می‌توانید همین حالا بنرش را آپلود کنید یا دسته‌بندیِ بعدی را اضافه کنید 👇",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            *vip_category_edit_keyboard(new_cat["id"]).inline_keyboard[:-1],
            [InlineKeyboardButton(text="➕ افزودنِ دسته‌بندیِ دیگر", callback_data="vipset:add", style="success")],
            [InlineKeyboardButton(text="🔙 بازگشت به لیست", callback_data="admin:vip_settings", style="primary")],
        ]),
    )

@dp.callback_query(F.data.startswith("vipset:edit:"))
async def cb_vipset_edit(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    cat = get_vip_category(cat_id)
    if not cat:
        await callback.answer("این دسته‌بندی دیگر موجود نیست.", show_alert=True)
        return
    text = (
        f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
        f"نام: <b>{html_escape(cat['name'])}</b>\n"
        f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
        "کدام مورد را می‌خواهید ویرایش کنید؟"
    )
    await callback.message.edit_text(text, reply_markup=vip_category_edit_keyboard(cat_id))
    await callback.answer()

@dp.callback_query(F.data.startswith("vipset:moveup:"))
async def cb_vipset_moveup(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    categories = load_vip_categories()
    index = next((i for i, c in enumerate(categories) if c["id"] == cat_id), None)
    if index is None or index == 0:
        await callback.answer("امکانِ جابه‌جایی نیست.", show_alert=True)
        return
    categories[index - 1], categories[index] = categories[index], categories[index - 1]
    await save_vip_categories(categories)
    await callback.answer("✅ جابه‌جا شد.")
    await callback.message.edit_reply_markup(reply_markup=vip_category_edit_keyboard(cat_id))

@dp.callback_query(F.data.startswith("vipset:movedown:"))
async def cb_vipset_movedown(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    categories = load_vip_categories()
    index = next((i for i, c in enumerate(categories) if c["id"] == cat_id), None)
    if index is None or index == len(categories) - 1:
        await callback.answer("امکانِ جابه‌جایی نیست.", show_alert=True)
        return
    categories[index + 1], categories[index] = categories[index], categories[index + 1]
    await save_vip_categories(categories)
    await callback.answer("✅ جابه‌جا شد.")
    await callback.message.edit_reply_markup(reply_markup=vip_category_edit_keyboard(cat_id))

@dp.callback_query(F.data.startswith("vipset:field:"))
async def cb_vipset_field(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    _, _, cat_id, field = callback.data.split(":", 3)
    cat = get_vip_category(cat_id)
    if not cat:
        await callback.answer("این دسته‌بندی دیگر موجود نیست.", show_alert=True)
        return

    await state.update_data(edit_cat_id=cat_id, edit_field=field)
    await state.set_state(VipCategoryStates.waiting_edit_value)

    prompts = {
        "name": "نامِ جدید را ارسال کنید:",
        "description": "توضیحاتِ جدید را ارسال کنید (می‌توانید چند خط باشد):",
    }
    await callback.message.edit_text(
        prompts.get(field, "مقدارِ جدید را ارسال کنید:"),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data=f"vipset:edit:{cat_id}", style="primary")]
        ]),
    )
    await callback.answer()

@dp.message(VipCategoryStates.waiting_edit_value)
async def handle_vipset_edit_value(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    cat_id = data.get("edit_cat_id")

    if message.text and message.text.startswith("/"):
        await state.clear()
        cat = get_vip_category(cat_id) if cat_id else None
        if cat:
            await message.answer(
                f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
                f"نام: <b>{html_escape(cat['name'])}</b>\n"
                f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
                "کدام مورد را می‌خواهید ویرایش کنید؟",
                reply_markup=vip_category_edit_keyboard(cat_id),
            )
        else:
            await message.answer(await build_vip_settings_text(), reply_markup=vip_settings_keyboard())
        return

    field = data.get("edit_field")
    categories = load_vip_categories()
    cat = next((c for c in categories if c["id"] == cat_id), None)
    if not cat:
        await state.clear()
        await message.answer("این دسته‌بندی دیگر موجود نیست.", reply_markup=vip_settings_keyboard())
        return

    if field == "name":
        cat["name"] = message.text.strip()
    elif field == "description":
        cat["description"] = message.text.strip()

    await save_vip_categories(categories)
    await state.clear()
    # بعد از ثبتِ تغییر، دوباره همان صفحه‌ی ویرایشِ همین دسته‌بندی را نشان می‌دهیم
    # تا ادمین بتواند بدونِ رفتن به منوی اصلی، مستقیم فیلدِ بعدی یا بنر را هم ویرایش کند.
    await message.answer(
        f"✅ با موفقیت به‌روزرسانی شد.\n\n"
        f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
        f"نام: <b>{html_escape(cat['name'])}</b>\n"
        f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
        "کدام مورد را می‌خواهید ویرایش کنید؟",
        reply_markup=vip_category_edit_keyboard(cat_id),
    )

# ---------- هندلرهای بنر VIP ----------
@dp.callback_query(F.data.startswith("vipset:banner:"))
async def cb_vipset_banner(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    cat = get_vip_category(cat_id)
    if not cat:
        await callback.answer("دسته‌بندی یافت نشد.", show_alert=True)
        return
    await state.update_data(edit_cat_id=cat_id)
    await state.set_state(VipCategoryStates.waiting_banner)
    await callback.message.edit_text(
        "🖼 لطفاً یک عکس برای بنر این دسته‌بندی ارسال کنید.\n"
        "عکس می‌تواند هر فرمتی داشته باشد (JPEG, PNG و غیره).\n"
        "(برای لغو، /cancel بفرستید)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data=f"vipset:edit:{cat_id}", style="primary")]
        ]),
    )
    await callback.answer()

@dp.message(VipCategoryStates.waiting_banner)
async def handle_vipset_banner_message(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text and message.text.startswith("/"):
        data = await state.get_data()
        cat_id = data.get("edit_cat_id")
        await state.clear()
        cat = get_vip_category(cat_id) if cat_id else None
        if cat:
            await message.answer(
                f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
                f"نام: <b>{html_escape(cat['name'])}</b>\n"
                f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
                "کدام مورد را می‌خواهید ویرایش کنید؟",
                reply_markup=vip_category_edit_keyboard(cat_id),
            )
        else:
            await message.answer(await build_vip_settings_text(), reply_markup=vip_settings_keyboard())
        return

    if not message.photo:
        await message.answer("❌ لطفاً یک عکس ارسال کنید (یا برای لغو /cancel بفرستید).")
        return

    data = await state.get_data()
    cat_id = data.get("edit_cat_id")
    if not cat_id:
        await state.clear()
        await message.answer("خطا: شناسه دسته‌بندی مشخص نیست.", reply_markup=vip_settings_keyboard())
        return
    file_id = message.photo[-1].file_id
    categories = load_vip_categories()
    cat = next((c for c in categories if c["id"] == cat_id), None)
    if not cat:
        await state.clear()
        await message.answer("دسته‌بندی دیگر موجود نیست.", reply_markup=vip_settings_keyboard())
        return
    cat["image_file_id"] = file_id
    await save_vip_categories(categories)
    await state.clear()
    # بعد از آپلودِ بنر، دوباره صفحه‌ی ویرایشِ همین دسته‌بندی را نشان می‌دهیم تا
    # ادمین بتواند بلافاصله ادامه بدهد (مثلاً دسته‌بندیِ بعدی یا فیلدِ دیگری را ویرایش کند).
    await message.answer(
        f"✅ بنر برای دسته‌بندی «{html_escape(cat['name'])}» با موفقیت آپلود شد.\n\n"
        f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
        f"نام: <b>{html_escape(cat['name'])}</b>\n"
        f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
        "کدام مورد را می‌خواهید ویرایش کنید؟",
        reply_markup=vip_category_edit_keyboard(cat_id),
    )

@dp.callback_query(F.data.startswith("vipset:delete_banner:"))
async def cb_vipset_delete_banner(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    categories = load_vip_categories()
    cat = next((c for c in categories if c["id"] == cat_id), None)
    if not cat:
        await callback.answer("دسته‌بندی یافت نشد.", show_alert=True)
        return
    cat.pop("image_file_id", None)
    await save_vip_categories(categories)
    await callback.answer("✅ بنر حذف شد.")
    await callback.message.edit_text(
        f"✏️ <b>ویرایشِ دسته‌بندی</b>\n\n"
        f"نام: <b>{html_escape(cat['name'])}</b>\n"
        f"توضیحات:\n{html_escape(cat.get('description', ''))}\n\n"
        "کدام مورد را می‌خواهید ویرایش کنید؟",
        reply_markup=vip_category_edit_keyboard(cat_id),
    )

@dp.callback_query(F.data.startswith("vipset:delete:"))
async def cb_vipset_delete(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    cat = get_vip_category(cat_id)
    if not cat:
        await callback.answer("این دسته‌بندی دیگر موجود نیست.", show_alert=True)
        return
    await callback.message.edit_text(
        f"⚠️ آیا از حذفِ دسته‌بندیِ «{html_escape(cat['name'])}» مطمئنید؟",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ بله، حذف شود", callback_data=f"vipset:delete_confirm:{cat_id}", style="danger")],
            [InlineKeyboardButton(text="❌ انصراف", callback_data="admin:vip_settings", style="primary")],
        ]),
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("vipset:delete_confirm:"))
async def cb_vipset_delete_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    cat_id = callback.data.split(":", 2)[2]
    categories = [c for c in load_vip_categories() if c["id"] != cat_id]
    await save_vip_categories(categories)
    await callback.answer("✅ حذف شد.")
    await callback.message.edit_text(await build_vip_settings_text(), reply_markup=vip_settings_keyboard())

# ---------- تنظیمات قیمت‌های جهانی ----------
async def build_vip_global_settings_text() -> str:
    settings = load_vip_global_settings()
    prices = settings.get("prices", {})
    discount = settings.get("discount_percent", 0)
    return (
        f"💰 <b>تنظیم قیمت اشتراک VIP</b>\n\n"
        f"قیمت ۳ ماهه: {format_toman(prices.get('3', 0))}\n"
        f"قیمت ۶ ماهه: {format_toman(prices.get('6', 0))}\n"
        f"قیمت ۱۲ ماهه: {format_toman(prices.get('12', 0))}\n"
        f"تخفیف درصدی: {to_persian_num(discount)}%\n"
        f"آخرین بروزرسانی: {format_jalali_datetime(datetime.fromisoformat(settings.get('updated_at', datetime.utcnow().isoformat())))}"
    )

def vip_global_settings_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✏️ قیمت ۳ ماهه", callback_data="vipglob:price3", style="primary")],
            [InlineKeyboardButton(text="✏️ قیمت ۶ ماهه", callback_data="vipglob:price6", style="primary")],
            [InlineKeyboardButton(text="✏️ قیمت ۱۲ ماهه", callback_data="vipglob:price12", style="primary")],
            [InlineKeyboardButton(text="✏️ تخفیف درصدی", callback_data="vipglob:discount", style="primary")],
            [InlineKeyboardButton(text="🔙 بازگشت", callback_data="admin:cat_vip", style="danger")],
        ]
    )

@dp.callback_query(F.data.startswith("vipglob:"))
async def cb_vipglob(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    field = callback.data.split(":", 1)[1]
    state_map = {
        "price3": VipGlobalSettingsStates.waiting_price3,
        "price6": VipGlobalSettingsStates.waiting_price6,
        "price12": VipGlobalSettingsStates.waiting_price12,
        "discount": VipGlobalSettingsStates.waiting_discount,
    }
    await state.set_state(state_map.get(field))
    prompts = {
        "price3": "قیمت جدیدِ ۳ ماهه را (فقط عدد، تومان) وارد کنید:",
        "price6": "قیمت جدیدِ ۶ ماهه را (فقط عدد، تومان) وارد کنید:",
        "price12": "قیمت جدیدِ ۱۲ ماهه را (فقط عدد، تومان) وارد کنید:",
        "discount": "تخفیف درصدی جدید را (فقط عدد، مثل 10 برای ۱۰٪) وارد کنید:",
    }
    await callback.message.edit_text(
        prompts.get(field, "مقدار جدید را وارد کنید:"),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data="admin:vip_global_settings", style="primary")]
        ]),
    )
    await callback.answer()

async def _parse_price_or_discount(message: Message) -> tuple[int | None, bool]:
    if not message.text:
        return None, False
    digits = message.text.strip().replace(",", "").replace("٬", "")
    persian_to_en = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
    digits = digits.translate(persian_to_en)
    if not digits.isdigit():
        return None, False
    return int(digits), True

@dp.message(VipGlobalSettingsStates.waiting_price3)
async def handle_vipglob_price3(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_global_settings_text(), reply_markup=vip_global_settings_keyboard())
        return
    price, ok = await _parse_price_or_discount(message)
    if not ok:
        await message.answer("❌ لطفاً فقط عدد وارد کنید.")
        return
    settings = load_vip_global_settings()
    settings["prices"]["3"] = price
    await save_vip_global_settings(settings)
    await state.clear()
    await message.answer(
        f"✅ قیمت ۳ ماهه به‌روزرسانی شد.\n\n{await build_vip_global_settings_text()}",
        reply_markup=vip_global_settings_keyboard(),
    )

@dp.message(VipGlobalSettingsStates.waiting_price6)
async def handle_vipglob_price6(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_global_settings_text(), reply_markup=vip_global_settings_keyboard())
        return
    price, ok = await _parse_price_or_discount(message)
    if not ok:
        await message.answer("❌ لطفاً فقط عدد وارد کنید.")
        return
    settings = load_vip_global_settings()
    settings["prices"]["6"] = price
    await save_vip_global_settings(settings)
    await state.clear()
    await message.answer(
        f"✅ قیمت ۶ ماهه به‌روزرسانی شد.\n\n{await build_vip_global_settings_text()}",
        reply_markup=vip_global_settings_keyboard(),
    )

@dp.message(VipGlobalSettingsStates.waiting_price12)
async def handle_vipglob_price12(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_global_settings_text(), reply_markup=vip_global_settings_keyboard())
        return
    price, ok = await _parse_price_or_discount(message)
    if not ok:
        await message.answer("❌ لطفاً فقط عدد وارد کنید.")
        return
    settings = load_vip_global_settings()
    settings["prices"]["12"] = price
    await save_vip_global_settings(settings)
    await state.clear()
    await message.answer(
        f"✅ قیمت ۱۲ ماهه به‌روزرسانی شد.\n\n{await build_vip_global_settings_text()}",
        reply_markup=vip_global_settings_keyboard(),
    )

@dp.message(VipGlobalSettingsStates.waiting_discount)
async def handle_vipglob_discount(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(await build_vip_global_settings_text(), reply_markup=vip_global_settings_keyboard())
        return
    discount, ok = await _parse_price_or_discount(message)
    if not ok or discount < 0 or discount > 100:
        await message.answer("❌ لطفاً عددی بین ۰ تا ۱۰۰ وارد کنید.")
        return
    settings = load_vip_global_settings()
    settings["discount_percent"] = discount
    await save_vip_global_settings(settings)
    await state.clear()
    await message.answer(
        f"✅ تخفیف به {to_persian_num(discount)}% تنظیم شد.\n\n{await build_vip_global_settings_text()}",
        reply_markup=vip_global_settings_keyboard(),
    )

# ==============================================================
#  بخش پاداش VIP (مسیر مستقل از خرید/تمدید)
# ==============================================================

class VipRewardStates(StatesGroup):
    waiting_user_id = State()
    waiting_days = State()
    waiting_reason = State()
    confirming = State()

@dp.callback_query(F.data == "vipreward:start")
async def cb_vipreward_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    await state.set_state(VipRewardStates.waiting_user_id)
    await callback.message.edit_text(
        "🎁 <b>اهدای پاداشِ VIP</b>\n\n"
        "این مسیر کاملاً مستقل از خریدِ اشتراک است؛ پیامی که برای کاربر ارسال می‌شود "
        "به‌جای «تاییدِ پرداخت»، با لحنِ «هدیه/تشویقِ مدیریت» نوشته می‌شود.\n\n"
        "آیدیِ عددیِ کاربر را ارسال کنید:\n"
        "(کاربر باید حداقل یک‌بار ربات را استارت کرده باشد تا شناسایی شود)\n\n"
        "(برای لغو، /cancel بفرستید)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 انصراف", callback_data="admin:cat_vip", style="primary")]
        ]),
    )
    await callback.answer()

@dp.message(VipRewardStates.waiting_user_id)
async def handle_vipreward_user_id(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_vip_category_keyboard())
        return

    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("❌ لطفاً فقط آیدیِ عددیِ کاربر را ارسال کنید (نه یوزرنیم).")
        return

    target_user_id = int(text)
    try:
        chat = await bot.get_chat(target_user_id)
        display_name = chat.full_name or (f"@{chat.username}" if chat.username else str(target_user_id))
    except Exception as e:
        await message.answer(
            f"❌ کاربری با این آیدی پیدا نشد یا ربات دسترسی ندارد: {e}\n"
            "توجه: کاربر باید حداقل یک‌بار ربات را استارت کرده باشد.\n"
            "لطفاً دوباره آیدی را ارسال کنید یا /cancel بزنید."
        )
        return

    await state.update_data(reward_user_id=target_user_id, reward_user_display=display_name)
    await state.set_state(VipRewardStates.waiting_days)
    await message.answer(
        f"👤 کاربر شناسایی شد: <b>{html_escape(display_name)}</b> (<code>{target_user_id}</code>)\n\n"
        "حالا تعدادِ روزِ پاداش را وارد کنید (فقط عدد).\n"
        "مثال: برای یک هفته بنویسید 7\n\n"
        "(برای لغو، /cancel بفرستید)"
    )

@dp.message(VipRewardStates.waiting_days)
async def handle_vipreward_days(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_vip_category_keyboard())
        return

    days, ok = await _parse_price_or_discount(message)  # همین تابعِ کمکیِ موجود، برای پارسِ عددِ فارسی/انگلیسی
    if not ok or not days or days <= 0:
        await message.answer("❌ لطفاً یک عددِ صحیحِ مثبت وارد کنید.")
        return

    await state.update_data(reward_days=days)
    await state.set_state(VipRewardStates.waiting_reason)
    await message.answer(
        "📝 در صورتِ تمایل، یک دلیل/توضیحِ کوتاه برای این پاداش بنویسید "
        "(مثلاً «به‌خاطرِ معرفیِ رواق در استوری»).\n"
        "این متن داخلِ پیامی که برای کاربر فرستاده می‌شود نمایش داده خواهد شد.\n\n"
        "اگر نمی‌خواهید دلیلی نوشته شود، فقط علامتِ «-» را ارسال کنید.\n"
        "(برای لغو، /cancel بفرستید)"
    )

@dp.message(VipRewardStates.waiting_reason)
async def handle_vipreward_reason(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer("لغو شد.", reply_markup=admin_vip_category_keyboard())
        return

    reason = (message.text or "").strip()
    if reason == "-":
        reason = ""

    await state.update_data(reward_reason=reason)
    data = await state.get_data()
    await state.set_state(VipRewardStates.confirming)

    reason_line = f"\n📝 دلیل: {html_escape(reason)}" if reason else ""
    summary = (
        "🎁 <b>تاییدِ نهاییِ پاداشِ VIP</b>\n\n"
        f"👤 کاربر: <b>{html_escape(data['reward_user_display'])}</b>\n"
        f"🆔 <code>{data['reward_user_id']}</code>\n"
        f"🗓 مدت: <b>{to_persian_num(data['reward_days'])} روز</b>"
        f"{reason_line}\n\n"
        "با تاییدِ زیر، این اعتبار برای کاربر فعال و پیامِ تشویقی برایش ارسال می‌شود."
    )
    await message.answer(
        summary,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ تایید و ارسال", callback_data="vipreward:confirm", style="success")],
            [InlineKeyboardButton(text="❌ انصراف", callback_data="vipreward:cancel", style="danger")],
        ]),
    )

@dp.callback_query(F.data == "vipreward:cancel")
async def cb_vipreward_cancel(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await state.clear()
    await callback.message.edit_text("عملیاتِ پاداش لغو شد.", reply_markup=admin_vip_category_keyboard())
    await callback.answer()

@dp.callback_query(F.data == "vipreward:confirm")
async def cb_vipreward_confirm(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ دسترسی ندارید.", show_alert=True)
        return
    data = await state.get_data()
    await state.clear()

    target_user_id = data.get("reward_user_id")
    days = data.get("reward_days")
    reason = data.get("reward_reason", "")
    display_name = data.get("reward_user_display", str(target_user_id))
    if not target_user_id or not days:
        await callback.message.edit_text("خطا: اطلاعاتِ پاداش ناقص است. دوباره تلاش کنید.")
        await callback.answer()
        return

    await callback.answer("⏳ در حال ثبت...")
    ok, result_msg, _end_dt = await _grant_vip_reward(
        target_user_id, days, reason=reason, granted_by=callback.from_user.id
    )
    icon = "✅" if ok else "❌"
    await callback.message.edit_text(
        f"{icon} {result_msg}",
        reply_markup=admin_vip_category_keyboard(),
    )

    if NOTIFY_CHAT_ID_INT:
        try:
            reason_line = f"\n📝 دلیل: {html_escape(reason)}" if reason else ""
            await bot.send_message(
                chat_id=NOTIFY_CHAT_ID_INT,
                text=(
                    "🎁 <b>پاداشِ VIP ثبت شد</b>\n\n"
                    f"👤 {html_escape(display_name)} (<code>{target_user_id}</code>)\n"
                    f"🗓 مدت: {to_persian_num(days)} روز\n"
                    f"👮 توسطِ: {html_escape(callback.from_user.full_name)}"
                    f"{reason_line}"
                ),
                disable_notification=True,
            )
        except Exception as e:
            logger.warning("اطلاع‌رسانیِ پاداشِ VIP به ادمین ممکن نشد: %s", e)

# ---------- مسیر سلامت و پینگ خودکار ----------
async def handle_health(request: web.Request) -> web.Response:
    return web.json_response({"ok": True})

async def self_ping_loop(app: web.Application) -> None:
    import aiohttp
    ping_url = f"{WEBHOOK_HOST}/health"
    async with aiohttp.ClientSession() as session:
        while True:
            await asyncio.sleep(PING_INTERVAL_SECONDS)
            try:
                async with session.get(ping_url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                    logger.info("پینگِ خودکار به %s — کدِ پاسخ: %s", ping_url, resp.status)
            except Exception as e:
                logger.warning("پینگِ خودکار ناموفق بود: %s", e)

async def start_self_ping(app: web.Application) -> None:
    app["self_ping_task"] = asyncio.create_task(self_ping_loop(app))

async def stop_self_ping(app: web.Application) -> None:
    task = app.get("self_ping_task")
    if task:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass

# ---------- بکاپِ دوره‌ایِ خودکار ----------
# چون FSM (وضعیتِ گفتگو، از جمله وسطِ خریدِ VIP) هم حالا داخلِ DATA_DIR ذخیره
# می‌شود، این حلقه با فاصله‌ی منظم آخرین نسخه‌ی همه‌چیز (دیتای اصلی + وضعیتِ
# گفتگوهای در حال انجام) را روی تلگرام پین می‌کند؛ تا اگر Render دقیقاً وسطِ
# یک فرآیندِ حساس (مثلاً آپلودِ رسیدِ پرداخت) ری‌استارت/دیپلوی شود، بیشترین
# چیزی که از دست می‌رود، تغییراتِ همان چند دقیقه‌ی آخر باشد، نه همه‌چیز.
async def auto_backup_loop(app: web.Application) -> None:
    while True:
        await asyncio.sleep(AUTO_BACKUP_INTERVAL_SECONDS)
        try:
            ok, msg = await backup_data_dir_to_telegram()
            if not ok:
                logger.info(f"بکاپِ خودکارِ دوره‌ای انجام نشد: {msg}")
        except Exception as e:
            logger.error(f"بکاپِ خودکارِ دوره‌ای با خطا مواجه شد: {e}", exc_info=True)

async def start_auto_backup(app: web.Application) -> None:
    app["auto_backup_task"] = asyncio.create_task(auto_backup_loop(app))

async def stop_auto_backup(app: web.Application) -> None:
    task = app.get("auto_backup_task")
    if task:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass

async def stop_vip_expiry_checker(app: web.Application) -> None:
    task = app.get("vip_expiry_task")
    if task:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass

async def stop_pending_join_checker(app: web.Application) -> None:
    task = app.get("pending_join_task")
    if task:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass

# ==============================================================
#  مدیریت خطاهای سراسری
# ==============================================================

@dp.errors()
async def global_error_handler(update: Update, exception: Exception):
    logger.error(f"❌ خطای سراسری: {exception}", exc_info=True)

    user_id = None
    chat_id = None

    if update.message:
        user_id = update.message.from_user.id
        chat_id = update.message.chat.id
    elif update.callback_query:
        user_id = update.callback_query.from_user.id
        chat_id = update.callback_query.message.chat.id

    if chat_id:
        try:
            await bot.send_message(
                chat_id=chat_id,
                text=(
                    "⚠️ <b>متأسفانه مشکلی پیش آمد!</b>\n\n"
                    "تیم فنی رواق مطلع شد و در حال رفع مشکل است.\n"
                    "لطفاً چند دقیقه بعد دوباره تلاش کنید.\n\n"
                    "اگر مشکل ادامه داشت، از طریق دکمه‌ی «📞 ارتباط با ادمین» به ما اطلاع دهید.\n"
                    "🙏 پوزش از بابت مزاحمت"
                ),
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [InlineKeyboardButton(text="📞 ارتباط با ادمین", callback_data="menu:contact_admin")]
                    ]
                )
            )
        except Exception as e:
            logger.error(f"ارسال پیام خطا به کاربر ممکن نشد: {e}")

    if NOTIFY_CHAT_ID_INT and user_id:
        try:
            error_summary = f"{exception.__class__.__name__}: {str(exception)[:100]}"

            context_bits = []
            profile_data = _pending_profile.get(user_id)
            if profile_data is not None:
                if not profile_data.get("education"):
                    context_bits.append("در پروفایل — سوال ۱ از ۳ (سطحِ تحصیلی)")
                elif not profile_data.get("referral"):
                    context_bits.append("در پروفایل — سوال ۲ از ۳ (نحوه‌ی آشنایی)")
                else:
                    context_bits.append("در پروفایل — سوال ۳ از ۳ (علایق)")
            try:
                fsm_context = FSMContext(storage=storage, key=StorageKey(bot_id=bot.id, chat_id=chat_id or user_id, user_id=user_id))
                current_fsm_state = await fsm_context.get_state()
                if current_fsm_state:
                    context_bits.append(f"وضعیتِ FSM: {current_fsm_state}")
            except Exception:
                pass
            context_line = f"مرحله: {' | '.join(context_bits)}\n" if context_bits else ""

            await bot.send_message(
                chat_id=NOTIFY_CHAT_ID_INT,
                text=(
                    f"🚨 <b>خطا در ربات</b>\n"
                    f"کاربر: <code>{user_id}</code>\n"
                    f"{context_line}"
                    f"خطا: <code>{error_summary}</code>\n"
                    f"زمان: {format_jalali_datetime(datetime.utcnow())}"
                )
            )
        except Exception as e:
            logger.error(f"ارسال گزارش خطا به ادمین ممکن نشد: {e}")

    return True

# ---------- راه‌اندازی وب‌سرور ----------
async def on_startup(app: web.Application):
    restored_ok, restore_msg = await restore_data_dir_from_telegram()
    storage.reload()  # چون storage قبل از این restore ساخته و از دیسک خوانده شده بود
    status_icon = "✅" if restored_ok else "⚠️"
    await _notify_backup_admin(f"{status_icon} بازیابیِ خودکارِ دیتا در استارتاپ:\n{restore_msg}")
    cache_users()

    await bot.set_webhook(
        WEBHOOK_URL,
        drop_pending_updates=True,
        allowed_updates=[
            "message",
            "chat_join_request",
            "chat_member",
            "poll_answer",
            "callback_query",
        ],
    )
    logger.info("Webhook تنظیم شد روی: %s", WEBHOOK_URL)

    await restore_attendance_tasks()
    app["vip_expiry_task"] = asyncio.create_task(vip_expiry_checker_loop())
    app["pending_join_task"] = asyncio.create_task(pending_join_checker_loop())
    logger.info("ربات «رواق» با موفقیت راه‌اندازی شد! 🏛")

def create_app() -> web.Application:
    app = web.Application()

    app.router.add_get("/health", handle_health)

    SimpleRequestHandler(dispatcher=dp, bot=bot).register(app, path=WEBHOOK_PATH)
    setup_application(app, dp, bot=bot)
    app.on_startup.append(on_startup)
    app.on_startup.append(start_self_ping)
    # بکاپِ خودکارِ دوره‌ای (هر ۱۵ دقیقه) به‌درخواستِ ادمین غیرفعال شد؛
    # بکاپ‌گیری از این پس فقط دستی و از طریقِ دکمه‌ی «📥 گرفتن بکاپ» در پنلِ
    # ادمین انجام می‌شود. توابعِ auto_backup_loop/start_auto_backup/stop_auto_backup
    # برای فعال‌سازیِ احتمالیِ دوباره در آینده همچنان در کد باقی مانده‌اند.
    app.on_cleanup.append(stop_self_ping)
    app.on_cleanup.append(stop_vip_expiry_checker)
    app.on_cleanup.append(stop_pending_join_checker)
    return app

if __name__ == "__main__":
    web.run_app(create_app(), host="0.0.0.0", port=PORT)